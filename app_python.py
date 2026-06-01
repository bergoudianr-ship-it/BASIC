import cgi
from copy import deepcopy
from datetime import datetime
from html import escape
from http.server import BaseHTTPRequestHandler, HTTPServer
import io
import json
from math import erf, exp, isfinite, pi, sqrt
from pathlib import Path
import re
import unicodedata
from urllib.parse import parse_qs, urlparse

from openpyxl import Workbook, load_workbook

HOST = "127.0.0.1"
PORT = 8899
MONTHS = ["M+0", "M+1", "M+2", "M+3", "M+4", "M+5", "M+6"]
DEFAULT_HOURS = [744, 720, 744, 744, 720, 744, 720]
EPS = 1e-6

FIXED_SOURCES = [
    "SE/CO",
    "SUL",
    "NORDESTE",
    "NORTE",
    "CONVENCIONAL",
    "I0",
    "I5",
    "CQ5",
    "I8",
    "I1",
]
DERIVATIVE_SOURCES = ["SE/CO", "SUL", "NORDESTE", "NORTE", "CONVENCIONAL"]
RISK_SOURCES = FIXED_SOURCES[:]
SOURCE_DISPLAY = {
    "SE/CO": "SUDESTE/CENTRO-OESTE",
    "SUL": "SUL",
    "NORDESTE": "NORDESTE",
    "NORTE": "NORTE",
    "CONVENCIONAL": "CONVENCIONAL",
    "I0": "INCENTIVADA 0%",
    "I5": "INCENTIVADA 50%",
    "CQ5": "CQ5",
    "I8": "INCENTIVADA 80%",
    "I1": "INCENTIVADA 100%",
}

SOURCE_FIELDS = {
    "SE/CO": "SECO",
    "SUL": "SUL",
    "NORDESTE": "NE",
    "NORTE": "N",
    "CONVENCIONAL": "CONV",
    "I0": "I0",
    "I5": "I5",
    "CQ5": "CQ5",
    "I8": "I8",
    "I1": "I1",
}

SECTION_DEFS = [
    {
        "key": "fixed",
        "title": "Portfolio - Preco Fixo, Consumo e Geracao",
        "sources": FIXED_SOURCES,
        "has_pm_requirement": True,
        "sheet": {
            "source_start": 6,
            "resource": 16,
            "pm_resource": 17,
            "requirement": 18,
            "pm_requirement": 19,
            "net": 20,
        },
    },
    {
        "key": "variable",
        "title": "Portfolio - Preco Variavel",
        "sources": FIXED_SOURCES,
        "has_pm_requirement": True,
        "sheet": {
            "source_start": 22,
            "resource": 32,
            "pm_resource": 33,
            "requirement": 34,
            "pm_requirement": 35,
            "net": 36,
        },
    },
    {
        "key": "derivatives",
        "title": "Portfolio - Derivativos",
        "sources": DERIVATIVE_SOURCES,
        "has_pm_requirement": False,
        "sheet": {
            "source_start": 38,
            "resource": 43,
            "pm_resource": 44,
            "requirement": 45,
            "pm_requirement": None,
            "net": 46,
        },
    },
]

HISTORY_PATH = Path(__file__).with_name("historico_analises.json")
RANKING_HISTORY_PATH = Path(__file__).with_name("historico_fa_ccee.json")
UPLOADS_DIR = Path(__file__).with_name("uploads_portfolio")
DEFAULT_XLSX_CANDIDATES = [
    Path.home() / "Downloads" / "2026.05.27 - Arquivo de apoio para simulação (declaração semanal).xlsx",
    Path(__file__).with_name("simulacao_semanal.xlsx"),
]
DEFAULT_XLSX_PATH = str(next((path for path in DEFAULT_XLSX_CANDIDATES if path.exists()), DEFAULT_XLSX_CANDIDATES[-1]))


def make_section_state(source_labels, has_pm_requirement):
    return {
        "sources": {label: [0.0] * 7 for label in source_labels},
        "resource": [0.0] * 7,
        "pmResource": [0.0] * 7,
        "requirement": [0.0] * 7,
        "pmRequirement": ([0.0] * 7) if has_pm_requirement else None,
        "netLine": [0.0] * 7,
    }


def new_default_state():
    return {
        "xlsxPath": DEFAULT_XLSX_PATH,
        "company": {
            "name": "",
            "cnpj": "",
            "analyst": "",
            "note": "",
        },
        "pla": 0.0,
        "parameters": {
            "faReference": 1.5,
            "confidence": 95.0,
            "phiZ": 1.6448536269,
            "liquidationDays": 1.0,
            "correlation": 1.0,
            "correlationMatrix": [[1.0 for _ in range(7)] for _ in range(7)],
            "theta": 0.0,
            "pldMin": 57.31,
            "pldMax": 785.27,
            "pldMinCurve": [57.31] * 7,
            "pldMaxCurve": [785.27] * 7,
        },
        "vertex": {
            "hours": DEFAULT_HOURS[:],
            "volatility": [0.0] * 7,
            "stressLongPrice": [0.0] * 7,
            "stressShortPrice": [0.0] * 7,
        },
        "portfolio": {
            "fixed": make_section_state(FIXED_SOURCES, True),
            "variable": make_section_state(FIXED_SOURCES, True),
            "derivatives": make_section_state(DERIVATIVE_SOURCES, False),
            "regulatedRevenue": [0.0] * 7,
        },
        "forward": {key: [0.0] * 7 for key in RISK_SOURCES},
    }


def clamp(value, min_value, max_value):
    return min(max_value, max(min_value, value))


def parse_number(value, default=0.0):
    if value is None:
        return default
    text = str(value).strip()
    if text == "":
        return default
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return default


def norm_text(value):
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in text if not unicodedata.combining(ch))


def br_number(value, decimals=2):
    if not isfinite(value):
        return "n.a."
    txt = f"{value:,.{decimals}f}"
    return txt.replace(",", "X").replace(".", ",").replace("X", ".")


def br_money(value):
    if not isfinite(value):
        return "R$ n.a."
    return f"R$ {br_number(value, 0)}"


def ratio(value):
    if not isfinite(value):
        return "n.a."
    return f"{br_number(value, 2)}x"


def pct(value):
    if not isfinite(value):
        return "n.a."
    return f"{br_number(value * 100, 1)}%"


def normal_pdf(z):
    return exp(-0.5 * z * z) / sqrt(2 * pi)


def normal_cdf(z):
    return (1 + erf(z / sqrt(2))) / 2


def get_sheet_by_name(wb, target_name):
    target = norm_text(target_name)
    for sheet in wb.sheetnames:
        if norm_text(sheet) == target:
            return wb[sheet]
    return None


def section_def_by_key(key):
    for item in SECTION_DEFS:
        if item["key"] == key:
            return item
    raise KeyError(key)


def forward_field(source):
    return SOURCE_FIELDS[source]


def get_post_state(params):
    state = new_default_state()
    state["xlsxPath"] = params.get("xlsx_path", [state["xlsxPath"]])[0].strip() or state["xlsxPath"]
    state["pla"] = parse_number(params.get("pla", [state["pla"]])[0], state["pla"])

    state["company"]["name"] = params.get("company_name", [""])[0].strip()
    state["company"]["cnpj"] = params.get("company_cnpj", [""])[0].strip()
    state["company"]["analyst"] = params.get("company_analyst", [""])[0].strip()
    state["company"]["note"] = params.get("company_note", [""])[0].strip()

    for key, value in state["parameters"].items():
        if isinstance(value, list):
            continue
        state["parameters"][key] = parse_number(params.get(f"parameters_{key}", [value])[0], value)

    for i in range(7):
        state["parameters"]["pldMinCurve"][i] = parse_number(
            params.get(f"parameters_pldMin_{i}", [state["parameters"]["pldMinCurve"][i]])[0],
            state["parameters"]["pldMinCurve"][i],
        )
        state["parameters"]["pldMaxCurve"][i] = parse_number(
            params.get(f"parameters_pldMax_{i}", [state["parameters"]["pldMaxCurve"][i]])[0],
            state["parameters"]["pldMaxCurve"][i],
        )
        for j in range(7):
            state["parameters"]["correlationMatrix"][i][j] = parse_number(
                params.get(
                    f"parameters_corr_{i}_{j}",
                    [state["parameters"]["correlationMatrix"][i][j]],
                )[0],
                state["parameters"]["correlationMatrix"][i][j],
            )

    state["parameters"]["pldMin"] = state["parameters"]["pldMinCurve"][0]
    state["parameters"]["pldMax"] = state["parameters"]["pldMaxCurve"][0]
    corr_vals = [state["parameters"]["correlationMatrix"][i][j] for i in range(7) for j in range(7)]
    state["parameters"]["correlation"] = sum(corr_vals) / len(corr_vals) if corr_vals else state["parameters"]["correlation"]

    for i in range(7):
        state["vertex"]["hours"][i] = parse_number(
            params.get(f"vertex_hours_{i}", [state["vertex"]["hours"][i]])[0], state["vertex"]["hours"][i]
        )
        state["vertex"]["volatility"][i] = parse_number(
            params.get(f"vertex_volatility_{i}", [state["vertex"]["volatility"][i]])[0],
            state["vertex"]["volatility"][i],
        )
        state["vertex"]["stressLongPrice"][i] = parse_number(
            params.get(f"vertex_stress_long_{i}", [state["vertex"]["stressLongPrice"][i]])[0],
            state["vertex"]["stressLongPrice"][i],
        )
        state["vertex"]["stressShortPrice"][i] = parse_number(
            params.get(f"vertex_stress_short_{i}", [state["vertex"]["stressShortPrice"][i]])[0],
            state["vertex"]["stressShortPrice"][i],
        )

    for section in SECTION_DEFS:
        key = section["key"]
        data = state["portfolio"][key]
        for source in section["sources"]:
            source_key = forward_field(source)
            for i in range(7):
                form_key = f"sec_{key}_{source_key}_{i}"
                data["sources"][source][i] = parse_number(params.get(form_key, [data["sources"][source][i]])[0], 0.0)

        for i in range(7):
            data["resource"][i] = parse_number(params.get(f"sec_{key}_resource_{i}", [data["resource"][i]])[0], 0.0)
            data["pmResource"][i] = parse_number(
                params.get(f"sec_{key}_pm_resource_{i}", [data["pmResource"][i]])[0], 0.0
            )
            data["requirement"][i] = parse_number(
                params.get(f"sec_{key}_requirement_{i}", [data["requirement"][i]])[0], 0.0
            )
            if section["has_pm_requirement"]:
                data["pmRequirement"][i] = parse_number(
                    params.get(f"sec_{key}_pm_requirement_{i}", [data["pmRequirement"][i]])[0], 0.0
                )

    for i in range(7):
        state["portfolio"]["regulatedRevenue"][i] = parse_number(
            params.get(f"regulated_revenue_{i}", [state["portfolio"]["regulatedRevenue"][i]])[0],
            state["portfolio"]["regulatedRevenue"][i],
        )

    for source in RISK_SOURCES:
        field = forward_field(source)
        for i in range(7):
            state["forward"][source][i] = parse_number(
                params.get(f"forward_{field}_{i}", [state["forward"][source][i]])[0],
                state["forward"][source][i],
            )
    return state


def read_row_values(ws, row):
    return [parse_number(ws.cell(row, 4 + i).value, 0.0) for i in range(7)]


def recompute_net_lines(state):
    for section in SECTION_DEFS:
        data = state["portfolio"][section["key"]]
        for i in range(7):
            data["netLine"][i] = data["resource"][i] - data["requirement"][i]
    return state


def load_calc_state_from_xlsx(path_text):
    path = Path(path_text).expanduser()
    if not path.exists():
        raise FileNotFoundError(f"Planilha nao encontrada: {path}")

    imported = new_default_state()
    imported["xlsxPath"] = str(path)
    wb = load_workbook(str(path), data_only=True)

    ws_premissas = get_sheet_by_name(wb, "Premissas")
    ws_forward = get_sheet_by_name(wb, "Curva Forward")
    ws_portfolio = get_sheet_by_name(wb, "Declaracao Portfolio")
    ws_pla = get_sheet_by_name(wb, "Patrimonio Liquido Ajustado")
    ws_consolidado = get_sheet_by_name(wb, "Consolidado")
    ws_var = get_sheet_by_name(wb, "Calculo VaR e Teste de Estresse")

    if ws_premissas:
        conf = parse_number(ws_premissas.cell(4, 2).value, 0.95)
        imported["parameters"]["confidence"] = conf * 100 if conf <= 1.5 else conf
        imported["parameters"]["liquidationDays"] = parse_number(
            ws_premissas.cell(5, 2).value, imported["parameters"]["liquidationDays"]
        )
        imported["parameters"]["pldMinCurve"] = [
            parse_number(ws_premissas.cell(13, 2 + i).value, imported["parameters"]["pldMinCurve"][i]) for i in range(7)
        ]
        imported["parameters"]["pldMaxCurve"] = [
            parse_number(ws_premissas.cell(14, 2 + i).value, imported["parameters"]["pldMaxCurve"][i]) for i in range(7)
        ]
        imported["parameters"]["pldMin"] = imported["parameters"]["pldMinCurve"][0]
        imported["parameters"]["pldMax"] = imported["parameters"]["pldMaxCurve"][0]

        for i in range(7):
            col = 2 + i
            vol_raw = parse_number(ws_premissas.cell(9, col).value, 0.0)
            imported["vertex"]["volatility"][i] = vol_raw * 100 if vol_raw <= 1.5 else vol_raw
            imported["vertex"]["stressLongPrice"][i] = parse_number(ws_premissas.cell(10, col).value, 0.0)
            imported["vertex"]["stressShortPrice"][i] = parse_number(ws_premissas.cell(11, col).value, 0.0)

        corr_vals = []
        for r in range(17, 24):
            for c in range(2, 9):
                value = parse_number(ws_premissas.cell(r, c).value, 1.0)
                imported["parameters"]["correlationMatrix"][r - 17][c - 2] = value
                corr_vals.append(value)
        if corr_vals:
            imported["parameters"]["correlation"] = sum(corr_vals) / len(corr_vals)

    if ws_var:
        phi = parse_number(ws_var.cell(5, 12).value, imported["parameters"]["phiZ"])
        if phi > 0:
            imported["parameters"]["phiZ"] = phi

    if ws_consolidado:
        imported["parameters"]["theta"] = parse_number(ws_consolidado.cell(4, 3).value, imported["parameters"]["theta"])
        imported["vertex"]["hours"] = [
            parse_number(ws_consolidado.cell(15, 3 + i).value, imported["vertex"]["hours"][i]) for i in range(7)
        ]

    if ws_forward:
        for i in range(7):
            col = 2 + i
            base = parse_number(ws_forward.cell(15, col).value, 0.0)
            spread_s = parse_number(ws_forward.cell(16, col).value, 0.0)
            spread_ne = parse_number(ws_forward.cell(17, col).value, 0.0)
            spread_n = parse_number(ws_forward.cell(18, col).value, 0.0)
            spread_i0 = parse_number(ws_forward.cell(19, col).value, 0.0)
            spread_i5 = parse_number(ws_forward.cell(20, col).value, 0.0)
            spread_i8 = parse_number(ws_forward.cell(21, col).value, 0.0)
            spread_i1 = parse_number(ws_forward.cell(22, col).value, 0.0)

            imported["forward"]["SE/CO"][i] = base
            imported["forward"]["CONVENCIONAL"][i] = base
            imported["forward"]["SUL"][i] = base + spread_s
            imported["forward"]["NORDESTE"][i] = base + spread_ne
            imported["forward"]["NORTE"][i] = base + spread_n
            imported["forward"]["I0"][i] = base + spread_i0
            imported["forward"]["I5"][i] = base + spread_i5
            imported["forward"]["CQ5"][i] = base + spread_i5
            imported["forward"]["I8"][i] = base + spread_i8
            imported["forward"]["I1"][i] = base + spread_i1

    if ws_portfolio:
        imported["pla"] = parse_number(ws_portfolio.cell(3, 2).value, imported["pla"])
        for section in SECTION_DEFS:
            section_data = imported["portfolio"][section["key"]]
            row_start = section["sheet"]["source_start"]
            for offset, source in enumerate(section["sources"]):
                section_data["sources"][source] = read_row_values(ws_portfolio, row_start + offset)
            section_data["resource"] = read_row_values(ws_portfolio, section["sheet"]["resource"])
            section_data["pmResource"] = read_row_values(ws_portfolio, section["sheet"]["pm_resource"])
            section_data["requirement"] = read_row_values(ws_portfolio, section["sheet"]["requirement"])
            if section["has_pm_requirement"]:
                section_data["pmRequirement"] = read_row_values(ws_portfolio, section["sheet"]["pm_requirement"])
            section_data["netLine"] = read_row_values(ws_portfolio, section["sheet"]["net"])

        imported["portfolio"]["regulatedRevenue"] = read_row_values(ws_portfolio, 48)

    if ws_pla:
        pla_sheet = parse_number(ws_pla.cell(4, 3).value, 0)
        if abs(pla_sheet) > EPS:
            imported["pla"] = pla_sheet

    return recompute_net_lines(imported)


def apply_imported_calc_data(base_state, imported_state):
    merged = deepcopy(base_state)
    merged["xlsxPath"] = imported_state["xlsxPath"]
    merged["parameters"] = imported_state["parameters"]
    merged["vertex"] = imported_state["vertex"]
    merged["portfolio"] = imported_state["portfolio"]
    merged["forward"] = imported_state["forward"]
    return merged


def apply_imported_portfolio_only(base_state, imported_state):
    merged = deepcopy(base_state)
    merged["xlsxPath"] = imported_state["xlsxPath"]
    merged["portfolio"] = imported_state["portfolio"]
    merged["vertex"]["hours"] = imported_state["vertex"]["hours"]
    return merged


def apply_imported_forward_only(base_state, imported_state):
    merged = deepcopy(base_state)
    merged["xlsxPath"] = imported_state["xlsxPath"]
    merged["forward"] = imported_state["forward"]
    return merged


def section_checks(section_key, section_data):
    checks = []
    for i, month in enumerate(MONTHS):
        src = section_data["sources"]
        source_sum = sum(src[source][i] for source in src.keys())
        net_line = section_data["netLine"][i]
        resource_minus_requirement = section_data["resource"][i] - section_data["requirement"][i]
        issue = []
        if abs(resource_minus_requirement - net_line) > 0.0001:
            issue.append("linha net difere de recurso-requisito")
        if abs(source_sum - net_line) > 0.0001:
            issue.append("linha net difere da soma das fontes")
        if section_key == "derivatives":
            sub_sum = (
                src.get("SE/CO", [0] * 7)[i]
                + src.get("SUL", [0] * 7)[i]
                + src.get("NORDESTE", [0] * 7)[i]
                + src.get("NORTE", [0] * 7)[i]
            )
            if abs(sub_sum - net_line) > 0.0001:
                issue.append("linha net derivativos difere da soma SE/CO+SUL+NE+N")
        checks.append({"month": month, "ok": len(issue) == 0, "issues": issue})
    return checks


def source_exposure_by_month(state, source, i):
    fixed_val = state["portfolio"]["fixed"]["sources"].get(source, [0.0] * 7)[i]
    variable_val = state["portfolio"]["variable"]["sources"].get(source, [0.0] * 7)[i]
    deriv_val = state["portfolio"]["derivatives"]["sources"].get(source, [0.0] * 7)[i]
    return fixed_val + variable_val + deriv_val


def build_source_month_summary(state, month_results, params):
    summaries = []
    for i, m in enumerate(month_results):
        pld_min_i = params.get("pldMinCurve", [params["pldMin"]] * 7)[i]
        pld_max_i = params.get("pldMaxCurve", [params["pldMax"]] * 7)[i]
        entries = []
        for source in RISK_SOURCES:
            exposure = source_exposure_by_month(state, source, i)
            forward = state["forward"].get(source, [0.0] * 7)[i]
            if abs(forward) <= EPS:
                forward = state["forward"]["SE/CO"][i]
            sigma = max(0, m["volatility"] / 100)
            mtm = exposure * forward * m["hours"]
            var_value = abs(mtm) * params["phiZ"] * sigma * sqrt(max(0, params["liquidationDays"]))
            if exposure >= 0:
                stress_price = m["stressLongPrice"] or max(pld_min_i, forward * (1 - 0.24))
            else:
                stress_price = m["stressShortPrice"] or min(pld_max_i, forward * (1 + 0.32))
            stress = abs(exposure) * m["hours"] * abs(forward - stress_price)
            risk = var_value + params["theta"] * stress
            entries.append(
                {
                    "source": source,
                    "exposure": exposure,
                    "forward": forward,
                    "mtm": mtm,
                    "var": var_value,
                    "stress": stress,
                    "risk": risk,
                }
            )

        entries.sort(key=lambda x: abs(x["risk"]), reverse=True)
        summaries.append(
            {
                "month": m["month"],
                "totalRisk": sum(abs(x["risk"]) for x in entries),
                "top": entries[:4],
                "all": entries,
            }
        )
    return summaries


def build_source_totals(source_month_summary):
    totals = {source: {"source": source, "exposure": 0.0, "mtm": 0.0, "var": 0.0, "stress": 0.0, "risk": 0.0} for source in RISK_SOURCES}
    for month in source_month_summary:
        for item in month["all"]:
            row = totals[item["source"]]
            row["exposure"] += item["exposure"]
            row["mtm"] += item["mtm"]
            row["var"] += item["var"]
            row["stress"] += item["stress"]
            row["risk"] += item["risk"]
    result = list(totals.values())
    result.sort(key=lambda x: abs(x["risk"]), reverse=True)
    return result


def calculate(state):
    recompute_net_lines(state)
    p = state["parameters"]
    pla = state["pla"]
    sqrt_d = sqrt(max(0, p["liquidationDays"]))
    checks = {section["key"]: section_checks(section["key"], state["portfolio"][section["key"]]) for section in SECTION_DEFS}

    month_results = []
    for i, month in enumerate(MONTHS):
        fixed_fin = (
            state["portfolio"]["fixed"]["requirement"][i] * state["portfolio"]["fixed"]["pmRequirement"][i]
            - state["portfolio"]["fixed"]["resource"][i] * state["portfolio"]["fixed"]["pmResource"][i]
        ) * state["vertex"]["hours"][i]
        variable_fin = (
            state["portfolio"]["variable"]["requirement"][i] * state["portfolio"]["variable"]["pmRequirement"][i]
            - state["portfolio"]["variable"]["resource"][i] * state["portfolio"]["variable"]["pmResource"][i]
        ) * state["vertex"]["hours"][i]
        res_contr = fixed_fin + variable_fin
        month_results.append(
            {
                "month": month,
                "hours": state["vertex"]["hours"][i],
                "volatility": state["vertex"]["volatility"][i],
                "stressLongPrice": state["vertex"]["stressLongPrice"][i],
                "stressShortPrice": state["vertex"]["stressShortPrice"][i],
                "resContr": res_contr,
                "regulatedRevenue": state["portfolio"]["regulatedRevenue"][i],
                "mtm": 0.0,
                "varValue": 0.0,
                "stressLoss": 0.0,
            }
        )

    source_month = build_source_month_summary(state, month_results, p)
    for i, detail in enumerate(source_month):
        month_results[i]["mtm"] = sum(item["mtm"] for item in detail["all"])
        month_results[i]["varValue"] = sum(item["var"] for item in detail["all"])
        month_results[i]["stressLoss"] = sum(item["stress"] for item in detail["all"])

    var_vector = [row["varValue"] for row in month_results]
    rho = clamp(p["correlation"], -1, 1)
    var_quadratic = 0.0
    for i in range(len(var_vector)):
        for j in range(len(var_vector)):
            var_quadratic += var_vector[i] * (1 if i == j else rho) * var_vector[j]
    var_total = sqrt(max(0, var_quadratic))

    stress_total = sum(row["stressLoss"] for row in month_results)
    rwa = var_total + p["theta"] * stress_total
    res_fin = sum(row["resContr"] + row["regulatedRevenue"] for row in month_results)
    pnl = sum(row["mtm"] + row["resContr"] for row in month_results)
    acr_revenue = sum(row["regulatedRevenue"] for row in month_results)
    fa_risk = rwa / pla if abs(pla) > EPS else float("inf")
    fa = max(0.0, (rwa - res_fin) / pla) if pla > 0 else float("inf")

    source_totals = build_source_totals(source_month)
    top_source_risk = abs(source_totals[0]["risk"]) if source_totals else 0.0
    top_source_ratio = top_source_risk / pla if pla > EPS else float("inf")
    stress_ratio = stress_total / pla if pla > EPS else float("inf")

    if pla <= EPS:
        score = 0.0
    else:
        m = max(0.001, p["faReference"])
        score = 100.0
        score -= min(60.0, (fa / m) * 40.0)
        score -= min(20.0, (fa_risk / m) * 15.0)
        score -= min(15.0, stress_ratio * 45.0)
        score -= min(20.0, top_source_ratio * 80.0)
        score = clamp(score, 0.0, 100.0)

    if score >= 90:
        rating = "AAA"
    elif score >= 80:
        rating = "AA"
    elif score >= 70:
        rating = "A"
    elif score >= 60:
        rating = "BBB"
    elif score >= 50:
        rating = "BB"
    elif score >= 40:
        rating = "B"
    else:
        rating = "CCC"

    notes = []
    if pla <= EPS:
        verdict = "Critica"
        notes.append("PLA <= 0. Sem base prudencial para sustentar o risco.")
    else:
        m = p["faReference"]
        if fa > m:
            notes.append(f"FA ({ratio(fa)}) acima da referencia M ({ratio(m)}).")
        elif fa > m * 0.8:
            notes.append(f"FA ({ratio(fa)}) em zona de atencao (>80% de M).")
        if fa_risk > m:
            notes.append("FA de risco acima da referencia.")
        if stress_ratio > 0.25:
            notes.append("Stress / PLA acima de 25%.")
        if top_source_ratio > 0.35:
            notes.append("Concentracao de risco por fonte/submercado acima de 35% do PLA.")
        if res_fin < 0:
            notes.append("Resultado financeiro prudencial negativo.")

        invalid_checks = []
        for key, check_rows in checks.items():
            for c in check_rows:
                if not c["ok"]:
                    invalid_checks.append(f"{key.upper()} {c['month']}: {', '.join(c['issues'])}")
        if invalid_checks:
            notes.append("Validacao do portfolio com divergencias de consistencia em linhas do template.")
        if not notes:
            notes.append("Sem gatilho critico nos parametros atuais.")

        if fa > m or fa_risk > m * 1.1:
            verdict = "Alavancada"
        elif fa > m * 0.8 or stress_ratio > 0.25 or top_source_ratio > 0.35:
            verdict = "Atencao"
        else:
            verdict = "Saudavel"

    return {
        "pla": pla,
        "monthResults": month_results,
        "varTotal": var_total,
        "stressTotal": stress_total,
        "rwa": rwa,
        "pnl": pnl,
        "acrRevenue": acr_revenue,
        "resFin": res_fin,
        "faRisk": fa_risk,
        "fa": fa,
        "score": score,
        "rating": rating,
        "stressRatio": stress_ratio,
        "topSourceRatio": top_source_ratio,
        "verdict": verdict,
        "notes": notes,
        "sourceMonthSummary": source_month,
        "sourceTotalSummary": source_totals,
        "checks": checks,
    }


def load_history():
    if not HISTORY_PATH.exists():
        return []
    try:
        return json.loads(HISTORY_PATH.read_text(encoding="utf-8"))
    except Exception:
        return []


def save_history(records):
    HISTORY_PATH.write_text(json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8")


def append_history_record(state, metrics):
    records = load_history()
    now = datetime.now().isoformat(timespec="seconds")
    record = {
        "timestamp": now,
        "company_name": state["company"]["name"],
        "company_cnpj": state["company"]["cnpj"],
        "analyst": state["company"]["analyst"],
        "note": state["company"]["note"],
        "xlsx_path": state.get("xlsxPath", ""),
        "verdict": metrics["verdict"],
        "fa": metrics["fa"],
        "faRisk": metrics["faRisk"],
        "pla": metrics["pla"],
        "rwa": metrics["rwa"],
        "varTotal": metrics["varTotal"],
        "stressTotal": metrics["stressTotal"],
        "resFin": metrics["resFin"],
        "pnl": metrics["pnl"],
        "acrRevenue": metrics["acrRevenue"],
        "stressRatio": metrics["stressRatio"],
        "topSourceRatio": metrics["topSourceRatio"],
        "score": metrics["score"],
        "rating": metrics["rating"],
        "notes": metrics["notes"],
        "parameters": {
            "confidence": state["parameters"]["confidence"],
            "liquidationDays": state["parameters"]["liquidationDays"],
            "phiZ": state["parameters"]["phiZ"],
            "correlation": state["parameters"]["correlation"],
            "theta": state["parameters"]["theta"],
            "faReference": state["parameters"]["faReference"],
        },
    }
    records.append(record)
    records = records[-500:]
    save_history(records)
    return record


def load_ranking_history():
    if not RANKING_HISTORY_PATH.exists():
        return []
    try:
        data = json.loads(RANKING_HISTORY_PATH.read_text(encoding="utf-8"))
        return data if isinstance(data, list) else []
    except Exception:
        return []


def save_ranking_history(records):
    RANKING_HISTORY_PATH.write_text(json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8")


def parse_week_key(value):
    text = str(value or "").strip()
    if not text:
        return ""
    m = re.search(r"(\d{2})/(\d{2})/(\d{4})", text)
    if m:
        return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
    m = re.search(r"(\d{2})/(\d{4})", text)
    if m:
        return f"{m.group(2)}-{m.group(1)}"
    return text


def parse_fa_ranking_xlsx(path):
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    header_row = None
    header_map = {}
    for row_idx in range(1, min(ws.max_row, 40) + 1):
        cells = [ws.cell(row_idx, col).value for col in range(1, min(ws.max_column, 40) + 1)]
        norm = [norm_text(v) for v in cells]
        if any("fator de alavancagem" in c for c in norm) and any("classe" in c for c in norm):
            header_row = row_idx
            for col_idx, val in enumerate(norm, start=1):
                if val:
                    header_map[val] = col_idx
            break

    if not header_row:
        raise ValueError("Nao encontrei cabecalho da CCEE (Classe/Fator de Alavancagem) no arquivo.")

    def find_col(options):
        for key, idx in header_map.items():
            for opt in options:
                if opt in key:
                    return idx
        return None

    col_class = find_col(["classe"])
    col_fa = find_col(["fator de alavancagem"])
    col_name = find_col(["razao social", "agente"])
    col_cnpj = find_col(["cnpj"])
    col_sigla = find_col(["sigla"])
    col_week = find_col(["inicio do periodo", "mes/ano evento", "mês/ano evento"])

    if not col_class or not col_fa or not col_name:
        raise ValueError("Arquivo sem colunas minimas: Classe, Razao Social/Agente e Fator de Alavancagem.")

    rows = []
    week_key = ""
    for row_idx in range(header_row + 1, ws.max_row + 1):
        cls = ws.cell(row_idx, col_class).value
        if not cls:
            continue
        cls_norm = norm_text(cls)
        if "comercial" not in cls_norm:
            continue

        fa = parse_number(ws.cell(row_idx, col_fa).value, float("nan"))
        if not isfinite(fa):
            continue

        name = str(ws.cell(row_idx, col_name).value or "").strip()
        if not name:
            continue

        cnpj = str(ws.cell(row_idx, col_cnpj).value or "").strip() if col_cnpj else ""
        sigla = str(ws.cell(row_idx, col_sigla).value or "").strip() if col_sigla else ""
        week_raw = ws.cell(row_idx, col_week).value if col_week else ""
        wk = parse_week_key(week_raw)
        if wk and not week_key:
            week_key = wk

        rows.append(
            {
                "name": name,
                "cnpj": cnpj,
                "sigla": sigla,
                "classe": str(cls).strip(),
                "fa": fa,
            }
        )

    if not rows:
        raise ValueError("Nenhuma comercializadora encontrada no arquivo.")

    if not week_key:
        week_key = datetime.now().strftime("%Y-%m-%d")
    return week_key, rows


def update_ranking_snapshot(path):
    week_key, rows = parse_fa_ranking_xlsx(path)
    history = load_ranking_history()
    now = datetime.now().isoformat(timespec="seconds")
    snapshot = {"week": week_key, "uploaded_at": now, "source_path": path, "records": rows}
    replaced = False
    for i, item in enumerate(history):
        if item.get("week") == week_key:
            history[i] = snapshot
            replaced = True
            break
    if not replaced:
        history.append(snapshot)
    history.sort(key=lambda x: x.get("week", ""))
    history = history[-120:]
    save_ranking_history(history)
    return snapshot, history


def build_ranking_view_data():
    history = load_ranking_history()
    if not history:
        return None
    current = history[-1]
    previous = history[-2] if len(history) > 1 else None
    prev_map = {}
    if previous:
        for rec in previous.get("records", []):
            key = (rec.get("cnpj") or "").strip() or norm_text(rec.get("name"))
            prev_map[key] = parse_number(rec.get("fa"), 0.0)

    ranking = []
    for rec in current.get("records", []):
        key = (rec.get("cnpj") or "").strip() or norm_text(rec.get("name"))
        fa_now = parse_number(rec.get("fa"), 0.0)
        fa_prev = prev_map.get(key)
        delta = fa_now - fa_prev if fa_prev is not None else float("nan")
        ranking.append({**rec, "delta": delta})

    ranking.sort(key=lambda x: parse_number(x.get("fa"), 0.0), reverse=True)
    ups = sorted([r for r in ranking if isfinite(r["delta"]) and r["delta"] > 0], key=lambda x: x["delta"], reverse=True)[:5]
    downs = sorted([r for r in ranking if isfinite(r["delta"]) and r["delta"] < 0], key=lambda x: x["delta"])[:5]
    critical = [r for r in ranking if parse_number(r.get("fa"), 0.0) >= 1.5]
    return {"history": history, "current": current, "previous": previous, "ranking": ranking, "ups": ups, "downs": downs, "critical": critical}


def initial_state():
    state = new_default_state()
    try:
        imported = load_calc_state_from_xlsx(state["xlsxPath"])
        return apply_imported_calc_data(state, imported)
    except Exception:
        return state


APP_STATE = initial_state()


def render_history_page():
    records = list(reversed(load_history()))
    total_analyses = len(records)
    unique_companies = len({(rec.get("company_name", "").strip().lower(), rec.get("company_cnpj", "").strip()) for rec in records})
    avg_score = br_number(
        sum(parse_number(rec.get("score"), 0.0) for rec in records) / total_analyses if total_analyses else 0.0,
        1,
    )
    last_timestamp = records[0].get("timestamp", "-") if records else "-"

    grouped = {}
    for rec in records:
        name = rec.get("company_name", "").strip() or "Empresa sem nome"
        cnpj = rec.get("company_cnpj", "").strip() or "CNPJ nao informado"
        key = f"{name}|||{cnpj}"
        if key not in grouped:
            grouped[key] = {"name": name, "cnpj": cnpj, "records": []}
        grouped[key]["records"].append(rec)

    company_blocks = []
    for group in grouped.values():
        rows = []
        for rec in group["records"]:
            params = rec.get("parameters", {})
            notes_text = " | ".join(rec.get("notes", []))
            params_text = (
                f"Conf {br_number(parse_number(params.get('confidence'), 0.0), 1)}% | "
                f"Dias {br_number(parse_number(params.get('liquidationDays'), 0.0), 2)} | "
                f"Theta {br_number(parse_number(params.get('theta'), 0.0), 4)} | "
                f"Corr {br_number(parse_number(params.get('correlation'), 0.0), 3)}"
            )
            rows.append(
                f"""
                <tr>
                  <td>{escape(rec.get("timestamp", ""))}</td>
                  <td>{escape(rec.get("verdict", ""))}</td>
                  <td>{escape(rec.get("rating", ""))} ({br_number(parse_number(rec.get("score"), 0.0), 1)})</td>
                  <td>{ratio(parse_number(rec.get("fa"), 0.0))}</td>
                  <td>{ratio(parse_number(rec.get("faRisk"), 0.0))}</td>
                  <td>{br_money(parse_number(rec.get("rwa"), 0.0))}</td>
                  <td>{br_money(parse_number(rec.get("pla"), 0.0))}</td>
                  <td>{br_money(parse_number(rec.get("varTotal"), 0.0))}</td>
                  <td>{br_money(parse_number(rec.get("stressTotal"), 0.0))}</td>
                  <td>{pct(parse_number(rec.get("stressRatio"), 0.0))}</td>
                  <td>{pct(parse_number(rec.get("topSourceRatio"), 0.0))}</td>
                  <td>{escape(rec.get("analyst", ""))}</td>
                  <td>{escape(rec.get("note", ""))}</td>
                  <td>{escape(params_text)}</td>
                  <td>{escape(notes_text)}</td>
                </tr>
                """
            )

        company_blocks.append(
            f"""
            <details class="company-card" open>
              <summary>
                <div>
                  <strong>{escape(group["name"])}</strong>
                  <small>{escape(group["cnpj"])}</small>
                </div>
                <span>{len(group["records"])} analise(s)</span>
              </summary>
              <div class="table-wrap">
                <table>
                  <thead>
                    <tr>
                      <th>Data/Hora</th><th>Parecer</th><th>Rating</th><th>FA</th><th>FA Risco</th>
                      <th>RWA</th><th>PLA</th><th>VaR</th><th>Stress</th><th>Stress/PLA</th>
                      <th>Maior Fonte/PLA</th><th>Analista</th><th>Observacao</th><th>Parametros</th><th>Notas</th>
                    </tr>
                  </thead>
                  <tbody>{''.join(rows)}</tbody>
                </table>
              </div>
            </details>
            """
        )

    if not company_blocks:
        company_blocks.append('<div class="card empty">Sem historico salvo ate o momento.</div>')

    return f"""<!doctype html>
<html lang="pt-BR">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>Historico de Analises</title>
  <style>
    body{{margin:0;background:#f6f6f2;font-family:Segoe UI,Arial,sans-serif;color:#1c221f}}
    .wrap{{max-width:1440px;margin:0 auto;padding:18px}}
    .top{{display:flex;justify-content:space-between;gap:10px;align-items:center;flex-wrap:wrap}}
    .card{{background:#fff;border:1px solid #d9ddd8;border-radius:8px;padding:14px}}
    .empty{{margin-top:12px;color:#67706b}}
    a.btn{{display:inline-block;padding:8px 12px;background:#12231e;color:#fff;border-radius:7px;text-decoration:none}}
    .summary-grid{{display:grid;grid-template-columns:repeat(4,minmax(180px,1fr));gap:10px;margin-top:12px}}
    .summary-grid .card span{{display:block;color:#67706b;font-size:.76rem;text-transform:uppercase}}
    .summary-grid .card strong{{display:block;margin-top:6px;font-size:1.35rem}}
    .company-list{{display:grid;gap:12px;margin-top:12px}}
    details.company-card{{background:#fff;border:1px solid #d9ddd8;border-radius:8px;padding:10px}}
    details.company-card summary{{list-style:none;display:flex;justify-content:space-between;align-items:center;cursor:pointer;gap:10px}}
    details.company-card summary::-webkit-details-marker{{display:none}}
    details.company-card summary strong{{display:block;font-size:1rem}}
    details.company-card summary small{{display:block;color:#67706b;margin-top:3px}}
    details.company-card summary span{{font-size:.82rem;color:#43524a;background:#eaf1ed;border-radius:999px;padding:6px 8px}}
    .table-wrap{{overflow:auto;border:1px solid #d9ddd8;border-radius:8px;margin-top:10px}}
    table{{width:100%;border-collapse:collapse;min-width:1640px}}
    th,td{{padding:8px;border-bottom:1px solid #d9ddd8;text-align:left;font-size:.86rem;vertical-align:top}}
    th{{font-size:.72rem;color:#67706b;text-transform:uppercase;background:#f9faf8;position:sticky;top:0}}
    @media(max-width:1080px){{.summary-grid{{grid-template-columns:1fr 1fr}}}}
    @media(max-width:680px){{.summary-grid{{grid-template-columns:1fr}}}}
  </style>
</head>
<body>
  <div class="wrap">
    <div class="top">
      <div>
        <h1 style="margin:0">Historico de Analises</h1>
        <p style="margin:6px 0 0 0;color:#67706b">Todas as analises salvas, agrupadas por empresa.</p>
      </div>
      <a class="btn" href="/">Voltar para calculadora</a>
    </div>

    <section class="summary-grid">
      <article class="card"><span>Total de analises</span><strong>{total_analyses}</strong></article>
      <article class="card"><span>Empresas unicas</span><strong>{unique_companies}</strong></article>
      <article class="card"><span>Score medio</span><strong>{avg_score}</strong></article>
      <article class="card"><span>Ultima analise</span><strong>{escape(last_timestamp)}</strong></article>
    </section>

    <section class="company-list">{''.join(company_blocks)}</section>
  </div>
</body>
</html>"""


def render_section_table(section, state):
    key = section["key"]
    data = state["portfolio"][key]
    month_headers = "".join(f"<th>{month}</th>" for month in MONTHS)

    def input_cells(name_prefix, values, locked=False):
        if locked:
            return "".join(
                f'<td><input class="locked-input" value="{values[i]}" readonly tabindex="-1" title="Calculado automaticamente: RECURSO - REQUISITO" /></td>'
                for i in range(7)
            )
        return "".join(f'<td><input name="{name_prefix}_{i}" value="{values[i]}" /></td>' for i in range(7))

    sheet_rows = []
    for source in section["sources"]:
        field = forward_field(source)
        sheet_rows.append(
            f"""
            <tr>
              <td>{escape(SOURCE_DISPLAY[source])}</td>
              <td>{escape(section["title"])}</td>
              <td>MWm</td>
              {input_cells(f"sec_{key}_{field}", data["sources"][source])}
            </tr>
            """
        )

    aggregate_specs = [
        ("RECURSO", "MWm", "resource", data["resource"]),
        ("PREÇO MÉDIO RECURSO", "R$/MWh", "pm_resource", data["pmResource"]),
        ("REQUISITO", "MWm", "requirement", data["requirement"]),
    ]
    if section["has_pm_requirement"]:
        aggregate_specs.append(("PREÇO MÉDIO REQUISITO", "R$/MWh", "pm_requirement", data["pmRequirement"]))
    aggregate_specs.append(("NET ENERGÉTICO", "MWm", "net", data["netLine"]))

    for label, unit, field, values in aggregate_specs:
        sheet_rows.append(
            f"""
            <tr>
              <td>{label}</td>
              <td>{escape(section["title"])}</td>
              <td>{unit}</td>
              {input_cells(f"sec_{key}_{field}", values, field == "net")}
            </tr>
            """
        )

    return f"""
    <article class="card">
      <h2>{escape(section["title"])}</h2>
      <div class="table-wrap">
        <table>
          <thead>
            <tr>
              <th>Exposições</th><th>Portifólio</th><th>Unid.</th>{month_headers}
            </tr>
          </thead>
          <tbody>{''.join(sheet_rows)}</tbody>
        </table>
      </div>
    </article>
    """


def render_ranking_page(flash_msg=""):
    data = build_ranking_view_data()
    if not data:
        empty = """
        <div class="card">
          <h2>Sem base semanal ainda</h2>
          <p>Suba o arquivo .xlsx da CCEE para gerar o ranking de comercializadoras e as variacoes semanais.</p>
        </div>
        """
        rows_html = ""
        cards = ""
        week_label = "-"
    else:
        week_label = data["current"].get("week", "-")
        cards = f"""
        <div class="kpi-grid">
          <div class="kpi"><small>Comercializadoras</small><strong>{len(data['ranking'])}</strong></div>
          <div class="kpi"><small>FA >= 1,5x</small><strong>{len(data['critical'])}</strong></div>
          <div class="kpi"><small>Semana atual</small><strong>{escape(week_label)}</strong></div>
        </div>
        """
        def bullet_rows(items):
            if not items:
                return "<li>Sem variacao relevante.</li>"
            return "".join(
                f"<li><strong>{escape(it['name'])}</strong>: {ratio(parse_number(it['fa'], 0.0))} ({br_number(it['delta'], 2)}x)</li>"
                for it in items
            )
        empty = f"""
        {cards}
        <div class="card">
          <h2>Principais alteracoes da semana</h2>
          <div class="split">
            <div><h3>Altas de FA</h3><ul>{bullet_rows(data['ups'])}</ul></div>
            <div><h3>Quedas de FA</h3><ul>{bullet_rows(data['downs'])}</ul></div>
          </div>
        </div>
        """
        rows = []
        for idx, rec in enumerate(data["ranking"], start=1):
            delta = rec["delta"]
            delta_text = br_number(delta, 2) + "x" if isfinite(delta) else "n.a."
            rows.append(
                f"<tr><td>{idx}</td><td>{escape(rec.get('name',''))}</td><td>{escape(rec.get('sigla',''))}</td><td>{escape(rec.get('cnpj',''))}</td><td>{ratio(parse_number(rec.get('fa'),0.0))}</td><td>{delta_text}</td></tr>"
            )
        rows_html = "".join(rows)

    return f"""<!doctype html>
<html lang="pt-BR">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>Ranking FA CCEE - Trading</title>
  <style>
    body{{margin:0;background:#f6f6f2;font-family:Segoe UI,Arial,sans-serif;color:#1c221f}}
    .wrap{{max-width:1200px;margin:0 auto;padding:18px}}
    .top{{display:flex;justify-content:space-between;align-items:center;gap:12px;flex-wrap:wrap}}
    .btn{{display:inline-block;background:#1f2a25;color:#fff;border-radius:8px;padding:9px 12px;text-decoration:none;border:none;cursor:pointer;font-weight:700}}
    .muted{{color:#66716b}}
    .card{{background:#fff;border:1px solid #d7dfd8;border-radius:8px;padding:14px;margin-bottom:12px}}
    .kpi-grid{{display:grid;grid-template-columns:repeat(3,minmax(160px,1fr));gap:10px;margin-bottom:12px}}
    .kpi{{background:#fff;border:1px solid #d7dfd8;border-radius:8px;padding:10px}}
    .kpi small{{display:block;color:#67706b}} .kpi strong{{font-size:1.3rem}}
    .split{{display:grid;grid-template-columns:1fr 1fr;gap:14px}}
    table{{width:100%;border-collapse:collapse;font-size:.92rem}} th,td{{border:1px solid #d8e1db;padding:7px;text-align:left}}
    th{{background:#f1f5f2}} ul{{margin:0;padding-left:18px}}
  </style>
</head>
<body>
  <div class="wrap">
    <div class="top">
      <div>
        <h1 style="margin:0">Ranking Semanal FA - Comercializadoras</h1>
        <p class="muted" style="margin:4px 0 0 0">Base da CCEE, com foco nas maiores variacoes para decisao de trading.</p>
      </div>
      <div style="display:flex;gap:8px;flex-wrap:wrap">
        <a class="btn" href="/">Voltar ao simulador</a>
        <a class="btn" href="/historico">Historico analises</a>
      </div>
    </div>
    <div class="card">
      <form method="post" enctype="multipart/form-data" action="/ranking" style="display:flex;gap:8px;flex-wrap:wrap;align-items:center">
        <input type="file" name="ranking_file" accept=".xlsx,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" required />
        <button class="btn" type="submit" name="action" value="upload_ranking_xlsx">Atualizar Semana CCEE</button>
        <span class="muted">Semana atual: {escape(week_label)}</span>
      </form>
      <p class="muted" style="margin:8px 0 0 0">{escape(flash_msg)}</p>
    </div>
    {empty}
    <div class="card">
      <h2 style="margin-top:0">Ranking por FA</h2>
      <table>
        <thead><tr><th>#</th><th>Empresa</th><th>Sigla</th><th>CNPJ</th><th>FA</th><th>Var. vs semana anterior</th></tr></thead>
        <tbody>{rows_html}</tbody>
      </table>
    </div>
  </div>
</body>
</html>"""


def render_main_page(state, metrics, flash_msg):
    p = state["parameters"]
    pla_input = "" if abs(state["pla"]) <= EPS else str(state["pla"])
    corr_month_headers = "".join(f"<th>M{i}</th>" for i in range(7))
    corr_rows = []
    for i in range(7):
        corr_cells = "".join(
            f'<td><input name="parameters_corr_{i}_{j}" value="{p["correlationMatrix"][i][j]}" /></td>' for j in range(7)
        )
        corr_rows.append(f"<tr><td>M{i}</td>{corr_cells}</tr>")

    checks_lines = []
    for section in SECTION_DEFS:
        key = section["key"]
        issue_months = [f'{c["month"]}: {", ".join(c["issues"])}' for c in metrics["checks"][key] if not c["ok"]]
        if issue_months:
            checks_lines.append(f'{section["title"]}: ' + " | ".join(issue_months))
    if not checks_lines:
        checks_text = "Checks do template: OK para todos os meses."
    else:
        checks_text = "Checks do template com divergencias: " + " || ".join(checks_lines)

    rows_forward = []
    for source in RISK_SOURCES:
        field = forward_field(source)
        cells = "".join(
            f'<td><input name="forward_{field}_{i}" value="{state["forward"][source][i]}" /></td>' for i in range(7)
        )
        rows_forward.append(f"<tr><td>{escape(source)}</td>{cells}</tr>")
    forward_headers = "".join(f"<th>{month}</th>" for month in MONTHS)

    vertex_rows = []
    vertex_rows.append(
        "<tr><td>Horas</td>"
        + "".join(f'<td><input name="vertex_hours_{i}" value="{state["vertex"]["hours"][i]}" /></td>' for i in range(7))
        + "</tr>"
    )
    vertex_rows.append(
        "<tr><td>Volatilidade %</td>"
        + "".join(
            f'<td><input name="vertex_volatility_{i}" value="{state["vertex"]["volatility"][i]}" /></td>' for i in range(7)
        )
        + "</tr>"
    )
    vertex_rows.append(
        "<tr><td>Preco Stress Long</td>"
        + "".join(
            f'<td><input name="vertex_stress_long_{i}" value="{state["vertex"]["stressLongPrice"][i]}" /></td>' for i in range(7)
        )
        + "</tr>"
    )
    vertex_rows.append(
        "<tr><td>Preco Stress Short</td>"
        + "".join(
            f'<td><input name="vertex_stress_short_{i}" value="{state["vertex"]["stressShortPrice"][i]}" /></td>' for i in range(7)
        )
        + "</tr>"
    )
    vertex_rows.append(
        "<tr><td>PLD Min</td>"
        + "".join(
            f'<td><input name="parameters_pldMin_{i}" value="{state["parameters"]["pldMinCurve"][i]}" /></td>' for i in range(7)
        )
        + "</tr>"
    )
    vertex_rows.append(
        "<tr><td>PLD Max</td>"
        + "".join(
            f'<td><input name="parameters_pldMax_{i}" value="{state["parameters"]["pldMaxCurve"][i]}" /></td>' for i in range(7)
        )
        + "</tr>"
    )

    regulated_row = (
        "<tr><td>Efeitos Financeiros do Mercado Regulado (R$)</td>"
        + "".join(
            f'<td><input name="regulated_revenue_{i}" value="{state["portfolio"]["regulatedRevenue"][i]}" /></td>'
            for i in range(7)
        )
        + "</tr>"
    )

    month_detail_rows = []
    for month in metrics["sourceMonthSummary"]:
        top_lines = []
        for item in month["top"]:
            direction = "Long" if item["exposure"] >= 0 else "Short"
            top_lines.append(
                f'{item["source"]}: {direction} {br_number(item["exposure"],2)} MWm | Forward {br_money(item["forward"])} | VaR {br_money(item["var"])} | Stress {br_money(item["stress"])}'
            )
        month_detail_rows.append(
            f"""
            <tr>
              <td>{escape(month["month"])}</td>
              <td>{br_money(month["totalRisk"])}</td>
              <td>{'<br>'.join(escape(x) for x in top_lines)}</td>
            </tr>
            """
        )

    source_total_rows = []
    for item in metrics["sourceTotalSummary"]:
        direction = "Long" if item["exposure"] >= 0 else "Short"
        source_total_rows.append(
            f"""
            <tr>
              <td>{escape(item["source"])}</td>
              <td>{direction}</td>
              <td>{br_number(item["exposure"],2)}</td>
              <td>{br_money(item["mtm"])}</td>
              <td>{br_money(item["var"])}</td>
              <td>{br_money(item["stress"])}</td>
              <td>{br_money(item["risk"])}</td>
            </tr>
            """
        )

    sections_html = "".join(render_section_table(section, state) for section in SECTION_DEFS)
    return f"""<!doctype html>
<html lang="pt-BR">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>Calculadora FA Prudencial</title>
  <style>
    :root {{
      --bg:#f6f6f2; --panel:#fff; --ink:#1c221f; --line:#d9ddd8; --muted:#67706b; --accent:#12231e;
    }}
    *{{box-sizing:border-box}}
    body{{margin:0;background:var(--bg);font-family:Segoe UI,Arial,sans-serif;color:var(--ink)}}
    .wrap{{max-width:1320px;margin:0 auto;padding:18px}}
    .top{{display:flex;justify-content:space-between;gap:10px;align-items:flex-end;flex-wrap:wrap;margin-bottom:12px}}
    h1{{margin:0;font-size:2rem}} h2{{margin:0 0 8px 0;font-size:1.18rem}}
    .muted{{color:var(--muted)}}
    .btns{{display:flex;gap:8px;flex-wrap:wrap;align-items:center}}
    button,a.btn{{background:var(--accent);color:#fff;border:1px solid var(--accent);border-radius:7px;padding:9px 12px;cursor:pointer;text-decoration:none;display:inline-block}}
    .upload-inline{{display:flex;gap:8px;align-items:center;flex-wrap:wrap}}
    .upload-inline input[type=file]{{max-width:320px;padding:6px}}
    .upload-inline button,.upload-inline a.btn{{white-space:nowrap}}
    .card{{background:var(--panel);border:1px solid var(--line);border-radius:8px;padding:14px}}
    .grid5{{display:grid;grid-template-columns:repeat(5,minmax(170px,1fr));gap:10px}}
    .big{{font-size:1.85rem;font-weight:700;margin-top:6px}}
    .layout{{display:grid;grid-template-columns:1fr;gap:12px;margin-top:12px}}
    .table-wrap{{overflow:auto;border:1px solid var(--line);border-radius:8px}}
    table{{width:100%;border-collapse:collapse;min-width:1100px}}
    th,td{{padding:8px;border-bottom:1px solid var(--line);text-align:left;font-size:.9rem}}
    th{{font-size:.72rem;color:var(--muted);text-transform:uppercase}}
    input,textarea{{width:100%;padding:7px;border:1px solid #c8cfca;border-radius:6px;background:#fff}}
    input.locked-input{{background:#eef2ef;color:#48524d;border-color:#d4ddd7;font-weight:700;cursor:not-allowed}}
    .formgrid{{display:grid;grid-template-columns:repeat(4,minmax(170px,1fr));gap:10px}}
    .field label{{display:block;font-size:.75rem;color:var(--muted);text-transform:uppercase;margin-bottom:4px}}
    .kpis{{display:grid;grid-template-columns:repeat(4,minmax(170px,1fr));gap:10px;margin-top:10px}}
    .tag{{display:inline-block;padding:4px 8px;border-radius:999px;background:#e8f0ec;color:#1f6d58;font-size:.8rem;font-weight:700}}
    @media(max-width:1080px){{.grid5,.formgrid,.kpis{{grid-template-columns:1fr 1fr}}}}
    @media(max-width:680px){{.grid5,.formgrid,.kpis{{grid-template-columns:1fr}}}}
  </style>
</head>
<body>
  <div class="wrap">
    <form method="post" enctype="multipart/form-data">
      <div class="top">
        <div>
          <p class="muted" style="margin:0">Monitoramento prudencial CCEE</p>
          <h1>Calculadora de Alavancagem</h1>
          <p class="muted" style="margin:6px 0 0 0">Plataforma Python espelhada na planilha semanal: portfolio, curva forward, premissas, PLA, RWA, FA e parecer.</p>
        </div>
        <div class="btns">
          <button type="submit" name="action" value="calculate">Recalcular</button>
          <button type="submit" name="action" value="import_xlsx">Importar da planilha</button>
          <button type="submit" name="action" value="save_analysis">Salvar analise</button>
          <a class="btn" href="/historico">Historico salvo</a>
        </div>
      </div>
      <div class="card" style="margin-bottom:12px"><p class="muted" style="margin:0">{escape(flash_msg)}</p></div>

      <section class="card" style="margin-bottom:12px">
        <h2>Empresa, importacao e modelos</h2>
        <div class="upload-inline" style="margin-bottom:10px">
          <input type="file" name="new_sheet_file" accept=".xlsx,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" />
          <button type="submit" name="action" value="upload_new_sheet_xlsx">Subir Nova Planilha Semanal</button>
          <a class="btn" href="/download/modelo_portfolio.xlsx">Baixar Modelo Portfolio</a>
          <a class="btn" href="/download/modelo_forward.xlsx">Baixar Modelo Curva Forward</a>
        </div>
        <div class="upload-inline" style="margin-bottom:10px">
          <input type="file" name="portfolio_file" accept=".xlsx,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" />
          <button type="submit" name="action" value="upload_portfolio_xlsx">Subir Portfolio (.xlsx)</button>
          <input type="file" name="forward_file" accept=".xlsx,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" />
          <button type="submit" name="action" value="upload_forward_xlsx">Subir Curva Forward (.xlsx)</button>
        </div>
        <div class="formgrid">
          <div class="field"><label>Empresa</label><input name="company_name" value="{escape(state["company"]["name"])}" /></div>
          <div class="field"><label>CNPJ</label><input name="company_cnpj" value="{escape(state["company"]["cnpj"])}" /></div>
          <div class="field"><label>Analista</label><input name="company_analyst" value="{escape(state["company"]["analyst"])}" /></div>
          <div class="field"><label>PLA ajustado (R$)</label><input name="pla" value="{pla_input}" /></div>
          <div class="field" style="grid-column:1 / -1"><label>Caminho planilha (opcional)</label><input name="xlsx_path" value="{escape(state["xlsxPath"])}" /></div>
          <div class="field" style="grid-column:1 / -1"><label>Nota da analise</label><textarea name="company_note" rows="2">{escape(state["company"]["note"])}</textarea></div>
        </div>
      </section>

      <section class="grid5">
        <article class="card"><span class="muted">FA divulgado</span><div class="big">{ratio(metrics["fa"])}</div></article>
        <article class="card"><span class="muted">FA risco</span><div class="big">{ratio(metrics["faRisk"])}</div></article>
        <article class="card"><span class="muted">RWA total</span><div class="big">{br_money(metrics["rwa"])}</div></article>
        <article class="card"><span class="muted">Score / Rating</span><div class="big">{br_number(metrics["score"],1)}</div><span class="muted">{escape(metrics["rating"])}</span></article>
        <article class="card"><span class="muted">Parecer</span><div class="big">{escape(metrics["verdict"])}</div><span class="tag">{escape(metrics["notes"][0])}</span></article>
      </section>

      <section class="layout">
        <article class="card">
          <h2>Premissas por vertice (M+0 a M+6)</h2>
          <div class="table-wrap">
            <table>
              <thead><tr><th>Parametro</th><th>M+0</th><th>M+1</th><th>M+2</th><th>M+3</th><th>M+4</th><th>M+5</th><th>M+6</th></tr></thead>
              <tbody>{''.join(vertex_rows)}</tbody>
            </table>
          </div>
        </article>

        {sections_html}

        <article class="card">
          <h2>Efeito financeiro do mercado regulado</h2>
          <div class="table-wrap">
            <table>
              <thead><tr><th>Linha</th><th>M+0</th><th>M+1</th><th>M+2</th><th>M+3</th><th>M+4</th><th>M+5</th><th>M+6</th></tr></thead>
              <tbody>{regulated_row}</tbody>
            </table>
          </div>
        </article>

        <article class="card">
          <h2>Curva Forward (ACL) por fonte/submercado</h2>
          <p class="muted">Usada na marcacao a mercado e no risco de cada fonte por mes.</p>
          <div class="table-wrap">
            <table>
              <thead><tr><th>R$/MWh</th>{forward_headers}</tr></thead>
              <tbody>{''.join(rows_forward)}</tbody>
            </table>
          </div>
          <div class="kpis">
            <div class="card"><span class="muted">Stress/PLA</span><div class="big">{pct(metrics["stressRatio"])}</div></div>
            <div class="card"><span class="muted">Maior fonte de risco/PLA</span><div class="big">{pct(metrics["topSourceRatio"])}</div></div>
            <div class="card"><span class="muted">VaR total</span><div class="big">{br_money(metrics["varTotal"])}</div></div>
            <div class="card"><span class="muted">Stress total</span><div class="big">{br_money(metrics["stressTotal"])}</div></div>
          </div>
        </article>

        <article class="card">
          <h2>Parametros Gerais (modelo planilha auxiliar)</h2>
          <div class="formgrid">
            <div class="field"><label>Intervalo de Confianca (%)</label><input name="parameters_confidence" value="{p["confidence"]}" /></div>
            <div class="field"><label>Dias para liquidacao</label><input name="parameters_liquidationDays" value="{p["liquidationDays"]}" /></div>
            <div class="field"><label>Peso Theta (adicional de stress)</label><input name="parameters_theta" value="{p["theta"]}" /></div>
            <div class="field"><label>Phi normal (z)</label><input name="parameters_phiZ" value="{p["phiZ"]}" /></div>
            <div class="field"><label>Correlacao media (da matriz)</label><input name="parameters_correlation" value="{p["correlation"]}" readonly /></div>
            <div class="field"><label>Referencia M (criterio interno)</label><input name="parameters_faReference" value="{p["faReference"]}" /></div>
          </div>
          <div style="margin-top:12px">
            <h3 style="margin:0 0 8px 0;font-size:1rem">Matriz de Correlacao (aba Premissas)</h3>
            <div class="table-wrap">
              <table style="min-width:760px">
                <thead><tr><th>Correlacao</th>{corr_month_headers}</tr></thead>
                <tbody>{''.join(corr_rows)}</tbody>
              </table>
            </div>
          </div>
        </article>

        <article class="card">
          <h2>Leitura metodologica e parecer</h2>
          <p><strong>Formula base:</strong> FA = max(0,(RWA - RES_FIN)/PLA), FA_Risco = RWA/PLA, RWA = VaR + theta*Stress.</p>
          <p><strong>Apuracao atual:</strong> VaR {br_money(metrics["varTotal"])}, Stress {br_money(metrics["stressTotal"])}, RWA {br_money(metrics["rwa"])}, RES_FIN {br_money(metrics["resFin"])}.</p>
          <p><strong>Checks do template:</strong> {escape(checks_text)}</p>
          <p><strong>Regras do parecer:</strong> {' '.join(escape(n) for n in metrics["notes"])}</p>
          <p class="muted">A analise detalha risco por mes e por fonte, para apoiar a gestao do portfolio ACL.</p>
        </article>

        <article class="card">
          <h2>Resumo de risco por mes (detalhado)</h2>
          <div class="table-wrap">
            <table>
              <thead><tr><th>Mes</th><th>Risco agregado (VaR + theta*Stress)</th><th>Principais fontes/submercados do risco</th></tr></thead>
              <tbody>{''.join(month_detail_rows)}</tbody>
            </table>
          </div>
        </article>

        <article class="card">
          <h2>Resumo de risco por fonte (horizonte total)</h2>
          <div class="table-wrap">
            <table>
              <thead><tr><th>Fonte/Submercado</th><th>Direcao</th><th>Exposicao MWm</th><th>MtM</th><th>VaR</th><th>Stress</th><th>Risco combinado</th></tr></thead>
              <tbody>{''.join(source_total_rows)}</tbody>
            </table>
          </div>
        </article>
      </section>
    </form>
  </div>
</body>
</html>"""


def parse_form_payload(handler):
    content_type = handler.headers.get("Content-Type", "")
    length = int(handler.headers.get("Content-Length", "0"))
    raw = handler.rfile.read(length)

    if "multipart/form-data" not in content_type.lower():
        body = raw.decode("utf-8", errors="replace")
        return parse_qs(body, keep_blank_values=True), {}

    environ = {"REQUEST_METHOD": "POST", "CONTENT_TYPE": content_type}
    form = cgi.FieldStorage(fp=io.BytesIO(raw), headers=handler.headers, environ=environ, keep_blank_values=True)
    params = {}
    files = {}
    if not form.list:
        return params, files

    for item in form.list:
        if item.filename:
            files[item.name] = {
                "filename": Path(item.filename).name,
                "content": item.file.read(),
            }
        else:
            params.setdefault(item.name, []).append(item.value)
    return params, files


def save_uploaded_xlsx(file_obj):
    if not file_obj:
        raise ValueError("Nenhum arquivo recebido.")
    filename = file_obj["filename"]
    if not filename.lower().endswith(".xlsx"):
        raise ValueError("Arquivo invalido. Envie um .xlsx no formato da declaracao.")
    content = file_obj["content"]
    if not content:
        raise ValueError("Arquivo vazio.")
    UPLOADS_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_name = f"portfolio_{stamp}_{filename}"
    save_path = UPLOADS_DIR / safe_name
    save_path.write_bytes(content)
    return str(save_path)


def workbook_to_bytes(wb):
    buff = io.BytesIO()
    wb.save(buff)
    return buff.getvalue()


def build_portfolio_template_bytes():
    wb = Workbook()
    ws = wb.active
    ws.title = "Declaração Portfólio"

    ws.cell(1, 1, "Declaração Portfólio")
    ws.cell(3, 1, "Patrimônio Líquido Ajustado")
    ws.cell(5, 1, "Exposições")
    ws.cell(5, 2, "Portifólio")
    ws.cell(5, 3, "Unid.")
    for i, month in enumerate(MONTHS):
        ws.cell(5, 4 + i, month)

    section_names = {
        "fixed": "Preço Fixo, Consumo e Geração",
        "variable": "Preço Variável",
        "derivatives": "Derivativos",
    }
    for section in SECTION_DEFS:
        sec_key = section["key"]
        sec_name = section_names[sec_key]
        row_start = section["sheet"]["source_start"]

        for offset, source in enumerate(section["sources"]):
            row = row_start + offset
            ws.cell(row, 1, SOURCE_DISPLAY[source])
            ws.cell(row, 2, sec_name)
            ws.cell(row, 3, "MWm")
            for i in range(7):
                ws.cell(row, 4 + i, 0.0)

        labels = [
            (section["sheet"]["resource"], "RECURSO", "MWm"),
            (section["sheet"]["pm_resource"], "PREÇO MÉDIO RECURSO", "R$/MWh"),
            (section["sheet"]["requirement"], "REQUISITO", "MWm"),
        ]
        if section["has_pm_requirement"]:
            labels.append((section["sheet"]["pm_requirement"], "PREÇO MÉDIO REQUISITO", "R$/MWh"))
        labels.append((section["sheet"]["net"], "NET ENERGÉTICO", "MWm"))

        for row, label, unit in labels:
            ws.cell(row, 1, label)
            ws.cell(row, 2, sec_name)
            ws.cell(row, 3, unit)
            for i in range(7):
                ws.cell(row, 4 + i, 0.0)

    ws.cell(48, 1, "EFEITOS FINANCEIROS DO MERCADO REGULADO")
    ws.cell(48, 2, "Receitas")
    ws.cell(48, 3, "R$")
    for i in range(7):
        ws.cell(48, 4 + i, 0.0)

    return workbook_to_bytes(wb)


def build_forward_template_bytes():
    wb = Workbook()
    ws = wb.active
    ws.title = "Curva Forward"

    ws.cell(1, 1, "Curva Forward - Modelo")
    ws.cell(2, 1, "Preencha valores no formato da CCEE (base e spreads por vertice).")
    ws.cell(14, 1, "R$/MWh")
    for i in range(7):
        ws.cell(14, 2 + i, MONTHS[i])

    rows = [
        (15, "SECO/CONV"),
        (16, "S"),
        (17, "NE"),
        (18, "N"),
        (19, "I0"),
        (20, "I5/CQ5"),
        (21, "I8"),
        (22, "I1"),
    ]
    for row, label in rows:
        ws.cell(row, 1, label)
        for i in range(7):
            ws.cell(row, 2 + i, 0.0)

    return workbook_to_bytes(wb)


class Handler(BaseHTTPRequestHandler):
    def do_GET(self):
        global APP_STATE
        parsed = urlparse(self.path)
        if parsed.path == "/ranking":
            self.send_response(404)
            self.end_headers()
            return
        if parsed.path == "/download/modelo_portfolio.xlsx":
            payload = build_portfolio_template_bytes()
            self.send_response(200)
            self.send_header(
                "Content-Type",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            self.send_header(
                "Content-Disposition",
                'attachment; filename="modelo_portfolio_prudencial.xlsx"',
            )
            self.send_header("Content-Length", str(len(payload)))
            self.end_headers()
            self.wfile.write(payload)
            return

        if parsed.path == "/download/modelo_forward.xlsx":
            payload = build_forward_template_bytes()
            self.send_response(200)
            self.send_header(
                "Content-Type",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            self.send_header(
                "Content-Disposition",
                'attachment; filename="modelo_curva_forward.xlsx"',
            )
            self.send_header("Content-Length", str(len(payload)))
            self.end_headers()
            self.wfile.write(payload)
            return

        if parsed.path == "/historico":
            content = render_history_page().encode("utf-8")
            self.send_response(200)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.send_header("Content-Length", str(len(content)))
            self.end_headers()
            self.wfile.write(content)
            return

        state = APP_STATE
        metrics = calculate(state)
        content = render_main_page(
            state,
            metrics,
            f"Planilha semanal carregada: {state['xlsxPath']}",
        ).encode("utf-8")
        self.send_response(200)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(content)))
        self.end_headers()
        self.wfile.write(content)

    def do_POST(self):
        global APP_STATE
        parsed = urlparse(self.path)
        if parsed.path == "/ranking":
            self.send_response(404)
            self.end_headers()
            return
        if parsed.path == "/historico":
            self.send_response(405)
            self.end_headers()
            return
        params, files = parse_form_payload(self)
        action = params.get("action", ["calculate"])[0]

        state = get_post_state(params)
        flash_msg = "Campos recalculados."

        if action == "upload_new_sheet_xlsx":
            try:
                file_data = files.get("new_sheet_file")
                uploaded_path = save_uploaded_xlsx(file_data)
                state["xlsxPath"] = uploaded_path
                imported = load_calc_state_from_xlsx(uploaded_path)
                state = apply_imported_calc_data(state, imported)
                flash_msg = f"Nova planilha importada com carga completa: {uploaded_path}"
            except Exception as exc:
                flash_msg = f"Falha no upload/importacao da nova planilha: {exc}"
        elif action == "upload_portfolio_xlsx":
            try:
                file_data = files.get("portfolio_file")
                uploaded_path = save_uploaded_xlsx(file_data)
                state["xlsxPath"] = uploaded_path
                imported = load_calc_state_from_xlsx(uploaded_path)
                state = apply_imported_portfolio_only(state, imported)
                flash_msg = f"Portfolio carregado e importado: {uploaded_path}"
            except Exception as exc:
                flash_msg = f"Falha no upload/importacao do portfolio: {exc}"
        elif action == "upload_forward_xlsx":
            try:
                file_data = files.get("forward_file")
                uploaded_path = save_uploaded_xlsx(file_data)
                state["xlsxPath"] = uploaded_path
                imported = load_calc_state_from_xlsx(uploaded_path)
                state = apply_imported_forward_only(state, imported)
                flash_msg = f"Curva forward carregada e importada: {uploaded_path}"
            except Exception as exc:
                flash_msg = f"Falha no upload/importacao da curva forward: {exc}"
        elif action == "import_xlsx":
            try:
                imported = load_calc_state_from_xlsx(state["xlsxPath"])
                state = apply_imported_calc_data(state, imported)
                flash_msg = f"Planilha importada com sucesso: {state['xlsxPath']}"
            except Exception as exc:
                flash_msg = f"Falha ao importar planilha: {exc}"
        elif action == "save_analysis":
            metrics_tmp = calculate(state)
            rec = append_history_record(state, metrics_tmp)
            flash_msg = "Analise salva no historico para " + (rec["company_name"] or "empresa sem nome") + "."

        metrics = calculate(state)
        APP_STATE = state
        content = render_main_page(state, metrics, flash_msg).encode("utf-8")
        self.send_response(200)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(content)))
        self.end_headers()
        self.wfile.write(content)

    def log_message(self, fmt, *args):
        return


def run():
    server = HTTPServer((HOST, PORT), Handler)
    print(f"Calculadora FA prudencial em http://{HOST}:{PORT}")
    server.serve_forever()


if __name__ == "__main__":
    run()
