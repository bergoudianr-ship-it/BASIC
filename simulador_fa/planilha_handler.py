import io
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

VERTICES = ["M+0", "M+1", "M+2", "M+3", "M+4", "M+5", "M+6"]
SUBMERCADOS = ["SE/CO", "SUL", "NE", "N"]
FONTES = ["CONVENCIONAL", "I0", "I5", "I8", "I1"]

# --- Upload: planilha semanal CCEE ---

def parse_planilha_ccee(file_bytes):
    """
    Estrutura real da planilha CCEE (2026.05.27):
    Aba 'Premissas':
      L4 B: Intervalo de Confiança (ex. 0.95)
      L5 B: Dias para liquidação
      L7 B-H: labels M0-M6
      L9 B-H: Volatilidade (já em decimal, ex. 0.0734)
      L10 B-H: Preço Estresse long
      L11 B-H: Preço Estresse short
      L13 B: PLD Min
      L14 B: PLD Max
    Aba 'Curva Forward' (tabela resumo):
      L14 A-H: labels (R$/MWh, M0..M6)
      L15: SECO/CONV  cols B-H
      L16: S (SUL spread) cols B-H
      L17: NE cols B-H
      L18: N cols B-H
      L19: I0 cols B-H
      L20: I5/CQ5 cols B-H
      L21: I8 cols B-H
      L22: I1 cols B-H
    Aba 'Cálculo VaR e Teste de Estresse':
      L5-L11 col B: horas/mês para M0-M6
      L5 col L: phi_norm (percentil, ex. 1.6449)
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    result = {}
    errors = []

    # Aba Premissas
    try:
        ws = wb["Premissas"]
        ic = _safe_float(ws["B4"].value)  # Intervalo de Confiança
        result["dias_liquidacao"] = int(_safe_float(ws["B5"].value) or 1)
        result["volatilidades"] = {}
        result["stress_long"] = {}
        result["stress_short"] = {}
        for i, v in enumerate(VERTICES):
            col = get_column_letter(2 + i)
            result["volatilidades"][v] = _safe_float(ws[f"{col}9"].value)
            result["stress_long"][v] = _safe_float(ws[f"{col}10"].value)
            result["stress_short"][v] = _safe_float(ws[f"{col}11"].value)
        result["pld_min"] = _safe_float(ws["B13"].value)
        result["pld_max"] = _safe_float(ws["B14"].value)
        # Derivar phi_norm do intervalo de confiança sem scipy
        # ic = 0.95 -> phi = 1.6449 -> phi_norm = -1.6449
        phi_table = {0.90: -1.2816, 0.95: -1.6449, 0.975: -1.9600, 0.99: -2.3263}
        if ic and 0 < ic <= 1:
            result["phi_norm"] = phi_table.get(round(ic, 3), -1.6449)
        else:
            result["phi_norm"] = -1.6449
    except Exception as e:
        errors.append(f"Aba 'Premissas': {e}")

    # Aba Curva Forward — usa tabela resumo L15-L22 (linhas 15-22, colunas B-H)
    try:
        ws = wb["Curva Forward"]
        # row 15=SECO, 16=SUL, 17=NE, 18=N, 19=I0, 20=I5, 21=I8, 22=I1
        subm_rows = [("SECO",15),("SUL",16),("NE",17),("N",18),("I0",19),("I5",20),("I8",21),("I1",22)]
        result["forward"] = {v: {} for v in VERTICES}
        for key, row in subm_rows:
            for i, v in enumerate(VERTICES):
                col = get_column_letter(2 + i)
                result["forward"][v][key] = _safe_float(ws[f"{col}{row}"].value)
        # data de coleta da linha 5 col A
        date_val = ws["A5"].value
        if hasattr(date_val, 'strftime'):
            result["data_referencia"] = date_val.strftime("%Y-%m-%d")
    except Exception as e:
        errors.append(f"Aba 'Curva Forward': {e}")

    # Aba Cálculo VaR e Teste de Estresse — horas col B linhas 5-11
    try:
        calc_name = next((n for n in wb.sheetnames if "lculo" in n and "aR" in n), None)
        if not calc_name:
            calc_name = next((n for n in wb.sheetnames if "var" in n.lower() or "estresse" in n.lower()), None)
        if calc_name:
            ws3 = wb[calc_name]
            result["horas"] = {}
            for i, v in enumerate(VERTICES):
                result["horas"][v] = int(_safe_float(ws3.cell(row=5+i, column=2).value) or 720)
            # phi_norm col L linha 5 (percentil direto, ex. 1.6449)
            phi_val = _safe_float(ws3.cell(row=5, column=12).value)
            if phi_val and phi_val > 0:
                result["phi_norm"] = -round(phi_val, 4)
    except Exception as e:
        errors.append(f"Aba VaR/horas: {e}")

    return result, errors


def _safe_float(val, pct=False):
    if val is None:
        return 0.0
    try:
        f = float(val)
        if pct and f > 1:
            f = f / 100
        return f
    except Exception:
        return 0.0


# --- Upload: portfólio preenchido ---

def parse_portfolio(file_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    empresa = {}
    errors = []

    # PLA
    try:
        ws = wb["PLA"]
        pl_bruto = _safe_float(ws["B2"].value)
        deducoes = {}
        itens = ["I","II","III","IV","V","VI","VII","VIII"]
        for idx, it in enumerate(itens):
            deducoes[it] = _safe_float(ws[f"B{3+idx}"].value)
        empresa["pla"] = {"pl_bruto": pl_bruto, "deducoes": deducoes}
    except Exception as e:
        errors.append(f"Aba 'PLA': {e}")
        empresa["pla"] = {"pl_bruto": 0, "deducoes": {k: 0 for k in ["I","II","III","IV","V","VI","VII","VIII"]}}

    # Portfólio Preço Fixo
    try:
        ws = wb["Portfólio Preço Fixo"]
        preco_fixo = _empty_preco_fixo()
        row = 2
        while ws.cell(row=row, column=1).value:
            v = str(ws.cell(row=row, column=1).value).strip()
            s = str(ws.cell(row=row, column=2).value).strip()
            f = str(ws.cell(row=row, column=3).value).strip()
            if v in VERTICES and s in SUBMERCADOS and f in FONTES:
                preco_fixo[v][s][f] = {
                    "recurso":    _safe_float(ws.cell(row=row, column=4).value),
                    "pm_rec":     _safe_float(ws.cell(row=row, column=5).value),
                    "requisito":  _safe_float(ws.cell(row=row, column=6).value),
                    "pm_req":     _safe_float(ws.cell(row=row, column=7).value),
                }
            row += 1
        empresa["preco_fixo"] = preco_fixo
    except Exception as e:
        errors.append(f"Aba 'Portfólio Preço Fixo': {e}")
        empresa["preco_fixo"] = _empty_preco_fixo()

    # Derivativos
    try:
        ws = wb["Derivativos"]
        derivativos = {}
        for i, v in enumerate(VERTICES):
            derivativos[v] = {
                "compra": _safe_float(ws.cell(row=2+i, column=2).value),
                "venda":  _safe_float(ws.cell(row=2+i, column=3).value),
            }
        empresa["derivativos"] = derivativos
    except Exception as e:
        errors.append(f"Aba 'Derivativos': {e}")
        empresa["derivativos"] = {v: {"compra": 0, "venda": 0} for v in VERTICES}

    # Preço Variável
    try:
        ws = wb["Preço Variável"]
        preco_variavel = {}
        for i, v in enumerate(VERTICES):
            preco_variavel[v] = {
                "recurso_pv":   _safe_float(ws.cell(row=2+i, column=2).value),
                "pm_rec_pv":    _safe_float(ws.cell(row=2+i, column=3).value),
                "requisito_pv": _safe_float(ws.cell(row=2+i, column=4).value),
                "pm_req_pv":    _safe_float(ws.cell(row=2+i, column=5).value),
            }
        empresa["preco_variavel"] = preco_variavel
    except Exception as e:
        errors.append(f"Aba 'Preço Variável': {e}")
        empresa["preco_variavel"] = {v: {"recurso_pv":0,"pm_rec_pv":0,"requisito_pv":0,"pm_req_pv":0} for v in VERTICES}

    # ACR
    try:
        ws = wb["ACR"]
        acr = {}
        for i, v in enumerate(VERTICES):
            acr[v] = _safe_float(ws.cell(row=2+i, column=2).value)
        empresa["acr"] = acr
    except Exception as e:
        errors.append(f"Aba 'ACR': {e}")
        empresa["acr"] = {v: 0 for v in VERTICES}

    return empresa, errors


def _empty_preco_fixo():
    pf = {}
    for v in VERTICES:
        pf[v] = {}
        for s in SUBMERCADOS:
            pf[v][s] = {}
            for f in FONTES:
                pf[v][s][f] = {"recurso": 0, "pm_rec": 0, "requisito": 0, "pm_req": 0}
    return pf


# --- Download: modelo de portfólio ---

HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(color="FFFFFF", bold=True)
INPUT_FILL = PatternFill("solid", fgColor="FFFACD")
THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _hdr(ws, row, col, val, width=None):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = HDR_FILL
    c.font = HDR_FONT
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width


def _inp(ws, row, col, val=None):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = INPUT_FILL
    c.border = BORDER
    return c


def gerar_modelo_portfolio():
    wb = openpyxl.Workbook()

    # --- Instruções ---
    ws = wb.active
    ws.title = "Instruções"
    instrucoes = [
        ("SIMULADOR FA CCEE — Modelo de Portfólio", ""),
        ("", ""),
        ("Aba PLA", "Preencha o PL bruto e os 8 itens de dedução do Anexo I do Manual CCEE v2023.2.0."),
        ("Aba Portfólio Preço Fixo", "Uma linha por combinação de Vértice × Submercado × Fonte. Recurso e Requisito em MWm médio. PM em R$/MWh. Sinal: positivo para recurso (venda), positivo para requisito (compra). Não altere as colunas A, B, C — são chave de leitura."),
        ("Aba Derivativos", "Compra e Venda em MWm médio por vértice. Compra = posição long, Venda = posição short."),
        ("Aba Preço Variável", "Recurso PV e Requisito PV com preços médios em R$/MWh por vértice."),
        ("Aba ACR", "Receita líquida de ACR em R$ por vértice. Não incluir CCEAR-Q (Quadro 9 do Manual)."),
        ("", ""),
        ("Unidades", "MWm = MW médio; R$/MWh = reais por megawatt-hora; R$ = reais"),
        ("Período sombra", "K=0 e θ=0 conforme Manual CCEE v2023.2.0, período sombra vigente."),
    ]
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 90
    for r, (a, b) in enumerate(instrucoes, 1):
        ws.cell(row=r, column=1, value=a).font = Font(bold=True)
        ws.cell(row=r, column=2, value=b)

    # --- PLA ---
    ws2 = wb.create_sheet("PLA")
    _hdr(ws2, 1, 1, "Item", 20)
    _hdr(ws2, 1, 2, "Valor (R$)", 20)
    ws2.cell(row=2, column=1, value="PL Bruto")
    _inp(ws2, 2, 2, 0)
    itens_ded = [
        ("I", "Prejuízos acumulados"),
        ("II", "Ativo intangível"),
        ("III", "Créditos tributários diferidos"),
        ("IV", "Participações em outras empresas"),
        ("V", "Outros ativos de longa maturação"),
        ("VI", "Contas a receber de partes relacionadas"),
        ("VII", "Garantias prestadas a partes relacionadas"),
        ("VIII", "Outros ajustes aprovados pela CCEE"),
    ]
    for idx, (it, desc) in enumerate(itens_ded):
        r = 3 + idx
        ws2.cell(row=r, column=1, value=f"Dedução {it} — {desc}")
        _inp(ws2, r, 2, 0)
    r_pla = 3 + len(itens_ded)
    ws2.cell(row=r_pla, column=1, value="PLA Calculado").font = Font(bold=True)
    ws2.cell(row=r_pla, column=2, value=f"=B2-SUM(B3:B{r_pla-1})").font = Font(bold=True)

    # --- Portfólio Preço Fixo ---
    ws3 = wb.create_sheet("Portfólio Preço Fixo")
    headers = ["Vértice", "Submercado", "Fonte", "Recurso (MWm)", "PM Recurso (R$/MWh)", "Requisito (MWm)", "PM Requisito (R$/MWh)"]
    col_widths = [8, 10, 14, 16, 22, 16, 22]
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        _hdr(ws3, 1, ci, h, w)
    row = 2
    for v in VERTICES:
        for s in SUBMERCADOS:
            for f in FONTES:
                ws3.cell(row=row, column=1, value=v)
                ws3.cell(row=row, column=2, value=s)
                ws3.cell(row=row, column=3, value=f)
                for ci in range(4, 8):
                    _inp(ws3, row, ci, 0)
                row += 1

    # --- Derivativos ---
    ws4 = wb.create_sheet("Derivativos")
    for ci, (h, w) in enumerate(zip(["Vértice", "Compra (MWm)", "Venda (MWm)"], [8, 14, 14]), 1):
        _hdr(ws4, 1, ci, h, w)
    for i, v in enumerate(VERTICES):
        ws4.cell(row=2+i, column=1, value=v)
        _inp(ws4, 2+i, 2, 0)
        _inp(ws4, 2+i, 3, 0)

    # --- Preço Variável ---
    ws5 = wb.create_sheet("Preço Variável")
    pv_hdrs = ["Vértice", "Recurso PV (MWm)", "PM Rec PV (R$/MWh)", "Requisito PV (MWm)", "PM Req PV (R$/MWh)"]
    pv_widths = [8, 18, 22, 18, 22]
    for ci, (h, w) in enumerate(zip(pv_hdrs, pv_widths), 1):
        _hdr(ws5, 1, ci, h, w)
    for i, v in enumerate(VERTICES):
        ws5.cell(row=2+i, column=1, value=v)
        for ci in range(2, 6):
            _inp(ws5, 2+i, ci, 0)

    # --- ACR ---
    ws6 = wb.create_sheet("ACR")
    for ci, (h, w) in enumerate(zip(["Vértice", "Receita ACR (R$)"], [8, 20]), 1):
        _hdr(ws6, 1, ci, h, w)
    ws6.cell(row=1, column=3, value="Nota: NÃO incluir CCEAR-Q (Quadro 9 Manual CCEE)").font = Font(italic=True, color="FF0000")
    for i, v in enumerate(VERTICES):
        ws6.cell(row=2+i, column=1, value=v)
        _inp(ws6, 2+i, 2, 0)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── Portfólio Extra — modelo e parser CSV ───────────────────────────────────
import csv as _csv
import unicodedata

EXTRA_SECOES = [
    ("Preco Fixo", "preco_fixo"),
    ("Preco Variavel", "preco_variavel"),
    ("Derivativos", "derivativos"),
]
EXTRA_CAMPOS = [
    ("Exposicao Submercado", "subm"),
    ("Recurso", "recurso"),
    ("PM Recurso", "pm_recurso"),
    ("Requisito", "requisito"),
    ("PM Requisito", "pm_requisito"),
]


def _norm(s):
    if s is None:
        return ""
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode("ascii")
    return s.strip().lower()


def _empty_extra_section():
    return {
        "subm": {s: {v: 0 for v in VERTICES} for s in SUBMERCADOS},
        "recurso": {v: 0 for v in VERTICES},
        "pm_recurso": {v: 0 for v in VERTICES},
        "requisito": {v: 0 for v in VERTICES},
        "pm_requisito": {v: 0 for v in VERTICES},
    }


def gerar_modelo_extra_csv():
    """Gera o CSV modelo do Portfólio Extra (separador ';', decimal '.')."""
    out = io.StringIO()
    w = _csv.writer(out, delimiter=";")
    w.writerow(["Secao", "Campo", "Submercado"] + VERTICES)
    for sec_label, _ in EXTRA_SECOES:
        for campo_label, campo_key in EXTRA_CAMPOS:
            if campo_key == "subm":
                for s in SUBMERCADOS:
                    w.writerow([sec_label, campo_label, s] + [0]*len(VERTICES))
            else:
                w.writerow([sec_label, campo_label, ""] + [0]*len(VERTICES))
    # EFM
    w.writerow(["EFM", "Efeitos Financeiros Mercado Regulado", ""] + [0]*len(VERTICES))
    return out.getvalue()


def _parse_num(x):
    if x is None or str(x).strip() == "":
        return 0.0
    s = str(x).strip().replace(" ", "")
    # aceita decimal com vírgula ou ponto
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return 0.0


def parse_extra_csv(file_bytes):
    """Lê o CSV do Portfólio Extra (separador ';' ou ',') e devolve o dict."""
    errors = []
    extra = {
        "ativo": False,
        "nome": "Portfólio Extra (Simulação)",
        "preco_fixo": _empty_extra_section(),
        "preco_variavel": _empty_extra_section(),
        "derivativos": _empty_extra_section(),
        "efm_regulado": {v: 0 for v in VERTICES},
    }
    try:
        text = file_bytes.decode("utf-8-sig", errors="replace")
    except Exception:
        text = file_bytes.decode("latin-1", errors="replace")

    # detecta separador
    first = text.splitlines()[0] if text.splitlines() else ""
    delim = ";" if first.count(";") >= first.count(",") else ","

    sec_map = {_norm(lbl): key for lbl, key in EXTRA_SECOES}
    campo_map = {_norm(lbl): key for lbl, key in EXTRA_CAMPOS}
    subm_map = {_norm(s): s for s in SUBMERCADOS}
    subm_map[_norm("SUDESTE/CENTRO-OESTE")] = "SE/CO"

    try:
        reader = _csv.reader(io.StringIO(text), delimiter=delim)
        rows = list(reader)
        if not rows:
            return extra, ["CSV vazio."]
        # pula cabeçalho se for texto
        start = 1 if rows and _norm(rows[0][0]) in ("secao", "seção") else 0
        for r in rows[start:]:
            if not r or all((c is None or str(c).strip() == "") for c in r):
                continue
            sec = _norm(r[0]) if len(r) > 0 else ""
            campo = _norm(r[1]) if len(r) > 1 else ""
            subm = _norm(r[2]) if len(r) > 2 else ""
            vals = [_parse_num(r[3+i]) if len(r) > 3+i else 0.0 for i in range(len(VERTICES))]

            if sec == "efm" or "efeitos financeiros" in campo or "efm" in sec:
                for i, v in enumerate(VERTICES):
                    extra["efm_regulado"][v] = vals[i]
                continue

            sec_key = sec_map.get(sec)
            campo_key = campo_map.get(campo)
            if not sec_key or not campo_key:
                continue
            target = extra[sec_key]
            if campo_key == "subm":
                s_real = subm_map.get(subm)
                if s_real:
                    for i, v in enumerate(VERTICES):
                        target["subm"][s_real][v] = vals[i]
            else:
                for i, v in enumerate(VERTICES):
                    target[campo_key][v] = vals[i]
    except Exception as e:
        errors.append(f"Erro ao ler CSV: {e}")

    return extra, errors
