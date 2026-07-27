"""
Simulador FA CCEE — aplicacao completa em um unico arquivo Python.
====================================================================
Backend em Python (calculo + planilhas) servindo a MESMA interface do HTML.

Como rodar:
    pip install openpyxl        # opcional (modelos .xlsx e import CCEE)
    python app.py               # abre em http://localhost:8765
    python app.py 9000          # porta customizada

Deploy (Render / Railway / Fly / qualquer host Python):
    Comando de start:  python app.py
    O servidor escuta em 0.0.0.0 e usa a variavel de ambiente PORT se existir.

Tudo (HTML, dados iniciais, calculo e leitura/escrita de planilhas) esta
embutido neste arquivo. O estado e mantido em memoria por processo.
"""
import sys
import json
import csv
import io
import re
import os
import traceback
from http.server import HTTPServer, BaseHTTPRequestHandler
from socketserver import ThreadingMixIn
from urllib.parse import urlparse


# ===== MOTOR DE CALCULO =====
import math

VERTICES = ["M+0", "M+1", "M+2", "M+3", "M+4", "M+5", "M+6"]
SUBMERCADOS = ["SE/CO", "SUL", "NE", "N"]
SUBM_SPREAD = {"SE/CO": "SECO", "SUL": "SUL", "NE": "NE", "N": "N"}


def _pla(empresa):
    pl = float(empresa["pla"].get("pl_bruto", 0) or 0)
    ded = sum(float(d.get("valor", 0) or 0) for d in empresa["pla"].get("deducoes", []))
    return pl - ded


def _forward_subm(premissas, vertice, subm):
    fwd = premissas["forward"][vertice]
    seco = float(fwd.get("SECO", 0) or 0)
    if subm == "SE/CO":
        return seco
    spread_key = SUBM_SPREAD[subm]
    return seco + float(fwd.get(spread_key, 0) or 0)


def _v(d, key, vertice):
    return float((d.get(key) or {}).get(vertice, 0) or 0)


# ─── Combinação Portfólio Real + Portfólio Extra ─────────────────────────────
def _zeros():
    return {v: 0.0 for v in VERTICES}


def _merge_section(base, extra):
    """Combina uma seção (preco_fixo / preco_variavel / derivativos).
    Volumes (subm, recurso, requisito) são SOMADOS.
    Preços médios (pm_recurso, pm_requisito) são MÉDIA PONDERADA pelo volume.
    """
    base = base or {}
    extra = extra or {}
    out = {"subm": {}, "recurso": {}, "pm_recurso": {},
           "requisito": {}, "pm_requisito": {}}

    # Exposição por submercado: soma
    for s in SUBMERCADOS:
        out["subm"][s] = {}
        for v in VERTICES:
            bv = float(((base.get("subm") or {}).get(s) or {}).get(v, 0) or 0)
            ev = float(((extra.get("subm") or {}).get(s) or {}).get(v, 0) or 0)
            out["subm"][s][v] = bv + ev

    # Recurso/Requisito: soma de volume + média ponderada de preço
    for vol, pm in (("recurso", "pm_recurso"), ("requisito", "pm_requisito")):
        out[vol] = {}
        out[pm] = {}
        for v in VERTICES:
            bvol = _v(base, vol, v)
            evol = _v(extra, vol, v)
            bpm = _v(base, pm, v)
            epm = _v(extra, pm, v)
            tot = bvol + evol
            out[vol][v] = tot
            out[pm][v] = ((bvol * bpm + evol * epm) / tot) if tot != 0 else 0.0
    return out


def combinar_portfolios(empresa, extra):
    """Retorna uma cópia da empresa com o portfólio extra incorporado.
    Mantém o PLA da empresa real. Soma volumes e pondera preços médios.
    """
    import copy
    comb = copy.deepcopy(empresa)
    comb["preco_fixo"] = _merge_section(empresa.get("preco_fixo"), extra.get("preco_fixo"))
    comb["preco_variavel"] = _merge_section(empresa.get("preco_variavel"), extra.get("preco_variavel"))
    comb["derivativos"] = _merge_section(empresa.get("derivativos"), extra.get("derivativos"))
    # EFM: soma
    efm = {}
    for v in VERTICES:
        efm[v] = float((empresa.get("efm_regulado") or {}).get(v, 0) or 0) \
               + float((extra.get("efm_regulado") or {}).get(v, 0) or 0)
    comb["efm_regulado"] = efm
    return comb


def calcular_fa(empresa, premissas):
    phi = abs(float(premissas.get("phi_norm", -1.6449) or -1.6449))
    D = float(premissas.get("dias_liquidacao", 1) or 1)
    sigma = premissas.get("volatilidades", {})
    stress_long = premissas.get("stress_long", {})
    stress_short = premissas.get("stress_short", {})
    horas = premissas.get("horas", {})

    pla = _pla(empresa)
    pf = empresa.get("preco_fixo", {})
    pv = empresa.get("preco_variavel", {})
    der = empresa.get("derivativos", {})

    por_vertice = {}
    res_contr_total = 0.0
    fin_pv_total = 0.0
    res_der_total = 0.0
    var_tot = 0.0
    stest_tot = 0.0
    efm_total = sum(float((empresa.get("efm_regulado") or {}).get(v, 0) or 0) for v in VERTICES)
    mtm_total_all = 0.0

    for vi in VERTICES:
        h = float((horas or {}).get(vi, 720) or 720)
        sig = float((sigma or {}).get(vi, 0) or 0)
        sl = float((stress_long or {}).get(vi, 0) or 0)
        ss = float((stress_short or {}).get(vi, 0) or 0)
        fwd_seco = float((premissas.get("forward") or {}).get(vi, {}).get("SECO", 0) or 0)

        # NET por submercado (soma PF + PV + DER)
        net_subm = {}
        for subm in SUBMERCADOS:
            n_pf  = float(((pf.get("subm") or {}).get(subm) or {}).get(vi, 0) or 0)
            n_pv  = float(((pv.get("subm") or {}).get(subm) or {}).get(vi, 0) or 0)
            n_der = float(((der.get("subm") or {}).get(subm) or {}).get(vi, 0) or 0)
            net_subm[subm] = n_pf + n_pv + n_der

        # MtM por submercado
        mtm_vi = 0.0
        for subm in SUBMERCADOS:
            fwd_s = _forward_subm(premissas, vi, subm)
            mtm_vi += net_subm[subm] * fwd_s * h

        mtm_total_all += mtm_vi

        # NET total (para VaR e Stress)
        net_total = sum(net_subm.values())

        # VaR: usa net_total × SECO forward (conforme planilha VaR)
        exp_var = net_total * fwd_seco * h
        var_vi = phi * abs(exp_var) * sig * math.sqrt(D)
        var_tot += var_vi

        # Stress Test
        if net_total > 0:
            stest_vi = sl * net_total * h - mtm_vi
        elif net_total < 0:
            stest_vi = ss * net_total * h - mtm_vi
        else:
            stest_vi = 0.0
        stest_tot += stest_vi

        # Resultado Contratual (Preço Fixo)
        rec_pf  = _v(pf, "recurso", vi)
        pmr_pf  = _v(pf, "pm_recurso", vi)
        req_pf  = _v(pf, "requisito", vi)
        pmq_pf  = _v(pf, "pm_requisito", vi)
        res_contr_vi = (req_pf * pmq_pf - rec_pf * pmr_pf) * h
        res_contr_total += res_contr_vi

        # Resultado Derivativos
        rec_der  = _v(der, "recurso", vi)
        pmr_der  = _v(der, "pm_recurso", vi)
        req_der  = _v(der, "requisito", vi)
        pmq_der  = _v(der, "pm_requisito", vi)
        res_der_vi = (req_der * pmq_der - rec_der * pmr_der) * h
        res_der_total += res_der_vi

        # Financeiro PLD+ (Preço Variável)
        rec_pv  = _v(pv, "recurso", vi)
        pmr_pv  = _v(pv, "pm_recurso", vi)
        req_pv  = _v(pv, "requisito", vi)
        pmq_pv  = _v(pv, "pm_requisito", vi)
        fin_pv_vi = (req_pv * pmq_pv - rec_pv * pmr_pv) * h
        fin_pv_total += fin_pv_vi

        por_vertice[vi] = {
            "net_seco":  round(net_subm["SE/CO"], 4),
            "net_sul":   round(net_subm["SUL"], 4),
            "net_ne":    round(net_subm["NE"], 4),
            "net_n":     round(net_subm["N"], 4),
            "net_total": round(net_total, 4),
            "mtm":       round(mtm_vi, 2),
            "res_contr": round(res_contr_vi + res_der_vi, 2),
            "fin_pv":    round(fin_pv_vi, 2),
            "efm":       round(float((empresa.get("efm_regulado") or {}).get(vi, 0) or 0), 2),
            "var":       round(var_vi, 2),
            "stest":     round(stest_vi, 2),
        }

    rwa = var_tot  # K=0, θ=0
    pnl = res_contr_total + res_der_total + mtm_total_all
    res_fin = pnl + fin_pv_total + efm_total

    pla_neg = pla <= 0
    fa_ris = None if pla_neg else rwa / pla
    fa_div = None if pla_neg else max(0.0, (rwa - res_fin) / pla)

    return {
        "por_vertice": por_vertice,
        "totais": {
            "res_contr":    round(res_contr_total + res_der_total, 2),
            "fin_pv":       round(fin_pv_total, 2),
            "efm":          round(efm_total, 2),
            "mtm":          round(mtm_total_all, 2),
            "pnl":          round(pnl, 2),
            "res_fin":      round(res_fin, 2),
            "var_tot":      round(var_tot, 2),
            "stest_tot":    round(stest_tot, 2),
            "rwa":          round(rwa, 2),
            "pla":          round(pla, 2),
            "fa_ris":       round(fa_ris, 4) if fa_ris is not None else None,
            "fa_divulgado": round(fa_div, 4) if fa_div is not None else None,
            "pla_negativo": pla_neg,
        }
    }


# ===== LEITURA/ESCRITA DE PLANILHAS =====
import io
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except Exception:
    _HAS_OPENPYXL = False

VERTICES = ["M+0", "M+1", "M+2", "M+3", "M+4", "M+5", "M+6"]
SUBMERCADOS = ["SE/CO", "SUL", "NE", "N"]
FONTES = ["CONVENCIONAL", "I0", "I5", "I8", "I1"]

# --- Upload: planilha semanal CCEE ---

def parse_planilha_ccee(file_bytes):
    """
    Estrutura real da planilha CCEE (2026.05.27):
    Aba 'Premissas':
      L4 B: Intervalo de Confiança (ex. 0.95)
      L5 B: Dias para liquidação
      L7 B-H: labels M0-M6
      L9 B-H: Volatilidade (já em decimal, ex. 0.0734)
      L10 B-H: Preço Estresse long
      L11 B-H: Preço Estresse short
      L13 B: PLD Min
      L14 B: PLD Max
    Aba 'Curva Forward' (tabela resumo):
      L14 A-H: labels (R$/MWh, M0..M6)
      L15: SECO/CONV  cols B-H
      L16: S (SUL spread) cols B-H
      L17: NE cols B-H
      L18: N cols B-H
      L19: I0 cols B-H
      L20: I5/CQ5 cols B-H
      L21: I8 cols B-H
      L22: I1 cols B-H
    Aba 'Cálculo VaR e Teste de Estresse':
      L5-L11 col B: horas/mês para M0-M6
      L5 col L: phi_norm (percentil, ex. 1.6449)
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    result = {}
    errors = []

    # Aba Premissas
    try:
        ws = wb["Premissas"]
        ic = _safe_float(ws["B4"].value)  # Intervalo de Confiança
        result["dias_liquidacao"] = int(_safe_float(ws["B5"].value) or 1)
        result["volatilidades"] = {}
        result["stress_long"] = {}
        result["stress_short"] = {}
        for i, v in enumerate(VERTICES):
            col = get_column_letter(2 + i)
            result["volatilidades"][v] = _safe_float(ws[f"{col}9"].value)
            result["stress_long"][v] = _safe_float(ws[f"{col}10"].value)
            result["stress_short"][v] = _safe_float(ws[f"{col}11"].value)
        result["pld_min"] = _safe_float(ws["B13"].value)
        result["pld_max"] = _safe_float(ws["B14"].value)
        # Derivar phi_norm do intervalo de confiança sem scipy
        # ic = 0.95 -> phi = 1.6449 -> phi_norm = -1.6449
        phi_table = {0.90: -1.2816, 0.95: -1.6449, 0.975: -1.9600, 0.99: -2.3263}
        if ic and 0 < ic <= 1:
            result["phi_norm"] = phi_table.get(round(ic, 3), -1.6449)
        else:
            result["phi_norm"] = -1.6449
    except Exception as e:
        errors.append(f"Aba 'Premissas': {e}")

    # Aba Curva Forward — usa tabela resumo L15-L22 (linhas 15-22, colunas B-H)
    try:
        ws = wb["Curva Forward"]
        # row 15=SECO, 16=SUL, 17=NE, 18=N, 19=I0, 20=I5, 21=I8, 22=I1
        subm_rows = [("SECO",15),("SUL",16),("NE",17),("N",18),("I0",19),("I5",20),("I8",21),("I1",22)]
        result["forward"] = {v: {} for v in VERTICES}
        for key, row in subm_rows:
            for i, v in enumerate(VERTICES):
                col = get_column_letter(2 + i)
                result["forward"][v][key] = _safe_float(ws[f"{col}{row}"].value)
        # data de coleta da linha 5 col A
        date_val = ws["A5"].value
        if hasattr(date_val, 'strftime'):
            result["data_referencia"] = date_val.strftime("%Y-%m-%d")
    except Exception as e:
        errors.append(f"Aba 'Curva Forward': {e}")

    # Aba Cálculo VaR e Teste de Estresse — horas col B linhas 5-11
    try:
        calc_name = next((n for n in wb.sheetnames if "lculo" in n and "aR" in n), None)
        if not calc_name:
            calc_name = next((n for n in wb.sheetnames if "var" in n.lower() or "estresse" in n.lower()), None)
        if calc_name:
            ws3 = wb[calc_name]
            result["horas"] = {}
            for i, v in enumerate(VERTICES):
                result["horas"][v] = int(_safe_float(ws3.cell(row=5+i, column=2).value) or 720)
            # phi_norm col L linha 5 (percentil direto, ex. 1.6449)
            phi_val = _safe_float(ws3.cell(row=5, column=12).value)
            if phi_val and phi_val > 0:
                result["phi_norm"] = -round(phi_val, 4)
    except Exception as e:
        errors.append(f"Aba VaR/horas: {e}")

    return result, errors


def _safe_float(val, pct=False):
    if val is None:
        return 0.0
    try:
        f = float(val)
        if pct and f > 1:
            f = f / 100
        return f
    except Exception:
        return 0.0


# --- Upload: portfólio preenchido ---

def parse_portfolio(file_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    empresa = {}
    errors = []

    # PLA
    try:
        ws = wb["PLA"]
        pl_bruto = _safe_float(ws["B2"].value)
        deducoes = {}
        itens = ["I","II","III","IV","V","VI","VII","VIII"]
        for idx, it in enumerate(itens):
            deducoes[it] = _safe_float(ws[f"B{3+idx}"].value)
        empresa["pla"] = {"pl_bruto": pl_bruto, "deducoes": deducoes}
    except Exception as e:
        errors.append(f"Aba 'PLA': {e}")
        empresa["pla"] = {"pl_bruto": 0, "deducoes": {k: 0 for k in ["I","II","III","IV","V","VI","VII","VIII"]}}

    # Portfólio Preço Fixo
    try:
        ws = wb["Portfólio Preço Fixo"]
        preco_fixo = _empty_preco_fixo()
        row = 2
        while ws.cell(row=row, column=1).value:
            v = str(ws.cell(row=row, column=1).value).strip()
            s = str(ws.cell(row=row, column=2).value).strip()
            f = str(ws.cell(row=row, column=3).value).strip()
            if v in VERTICES and s in SUBMERCADOS and f in FONTES:
                preco_fixo[v][s][f] = {
                    "recurso":    _safe_float(ws.cell(row=row, column=4).value),
                    "pm_rec":     _safe_float(ws.cell(row=row, column=5).value),
                    "requisito":  _safe_float(ws.cell(row=row, column=6).value),
                    "pm_req":     _safe_float(ws.cell(row=row, column=7).value),
                }
            row += 1
        empresa["preco_fixo"] = preco_fixo
    except Exception as e:
        errors.append(f"Aba 'Portfólio Preço Fixo': {e}")
        empresa["preco_fixo"] = _empty_preco_fixo()

    # Derivativos
    try:
        ws = wb["Derivativos"]
        derivativos = {}
        for i, v in enumerate(VERTICES):
            derivativos[v] = {
                "compra": _safe_float(ws.cell(row=2+i, column=2).value),
                "venda":  _safe_float(ws.cell(row=2+i, column=3).value),
            }
        empresa["derivativos"] = derivativos
    except Exception as e:
        errors.append(f"Aba 'Derivativos': {e}")
        empresa["derivativos"] = {v: {"compra": 0, "venda": 0} for v in VERTICES}

    # Preço Variável
    try:
        ws = wb["Preço Variável"]
        preco_variavel = {}
        for i, v in enumerate(VERTICES):
            preco_variavel[v] = {
                "recurso_pv":   _safe_float(ws.cell(row=2+i, column=2).value),
                "pm_rec_pv":    _safe_float(ws.cell(row=2+i, column=3).value),
                "requisito_pv": _safe_float(ws.cell(row=2+i, column=4).value),
                "pm_req_pv":    _safe_float(ws.cell(row=2+i, column=5).value),
            }
        empresa["preco_variavel"] = preco_variavel
    except Exception as e:
        errors.append(f"Aba 'Preço Variável': {e}")
        empresa["preco_variavel"] = {v: {"recurso_pv":0,"pm_rec_pv":0,"requisito_pv":0,"pm_req_pv":0} for v in VERTICES}

    # ACR
    try:
        ws = wb["ACR"]
        acr = {}
        for i, v in enumerate(VERTICES):
            acr[v] = _safe_float(ws.cell(row=2+i, column=2).value)
        empresa["acr"] = acr
    except Exception as e:
        errors.append(f"Aba 'ACR': {e}")
        empresa["acr"] = {v: 0 for v in VERTICES}

    return empresa, errors


def _empty_preco_fixo():
    pf = {}
    for v in VERTICES:
        pf[v] = {}
        for s in SUBMERCADOS:
            pf[v][s] = {}
            for f in FONTES:
                pf[v][s][f] = {"recurso": 0, "pm_rec": 0, "requisito": 0, "pm_req": 0}
    return pf


# --- Download: modelo de portfólio ---

HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(color="FFFFFF", bold=True)
INPUT_FILL = PatternFill("solid", fgColor="FFFACD")
THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _hdr(ws, row, col, val, width=None):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = HDR_FILL
    c.font = HDR_FONT
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width


def _inp(ws, row, col, val=None):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = INPUT_FILL
    c.border = BORDER
    return c


def gerar_modelo_portfolio():
    wb = openpyxl.Workbook()

    # --- Instruções ---
    ws = wb.active
    ws.title = "Instruções"
    instrucoes = [
        ("SIMULADOR FA CCEE — Modelo de Portfólio", ""),
        ("", ""),
        ("Aba PLA", "Preencha o PL bruto e os 8 itens de dedução do Anexo I do Manual CCEE v2023.2.0."),
        ("Aba Portfólio Preço Fixo", "Uma linha por combinação de Vértice × Submercado × Fonte. Recurso e Requisito em MWm médio. PM em R$/MWh. Sinal: positivo para recurso (venda), positivo para requisito (compra). Não altere as colunas A, B, C — são chave de leitura."),
        ("Aba Derivativos", "Compra e Venda em MWm médio por vértice. Compra = posição long, Venda = posição short."),
        ("Aba Preço Variável", "Recurso PV e Requisito PV com preços médios em R$/MWh por vértice."),
        ("Aba ACR", "Receita líquida de ACR em R$ por vértice. Não incluir CCEAR-Q (Quadro 9 do Manual)."),
        ("", ""),
        ("Unidades", "MWm = MW médio; R$/MWh = reais por megawatt-hora; R$ = reais"),
        ("Período sombra", "K=0 e θ=0 conforme Manual CCEE v2023.2.0, período sombra vigente."),
    ]
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 90
    for r, (a, b) in enumerate(instrucoes, 1):
        ws.cell(row=r, column=1, value=a).font = Font(bold=True)
        ws.cell(row=r, column=2, value=b)

    # --- PLA ---
    ws2 = wb.create_sheet("PLA")
    _hdr(ws2, 1, 1, "Item", 20)
    _hdr(ws2, 1, 2, "Valor (R$)", 20)
    ws2.cell(row=2, column=1, value="PL Bruto")
    _inp(ws2, 2, 2, 0)
    itens_ded = [
        ("I", "Prejuízos acumulados"),
        ("II", "Ativo intangível"),
        ("III", "Créditos tributários diferidos"),
        ("IV", "Participações em outras empresas"),
        ("V", "Outros ativos de longa maturação"),
        ("VI", "Contas a receber de partes relacionadas"),
        ("VII", "Garantias prestadas a partes relacionadas"),
        ("VIII", "Outros ajustes aprovados pela CCEE"),
    ]
    for idx, (it, desc) in enumerate(itens_ded):
        r = 3 + idx
        ws2.cell(row=r, column=1, value=f"Dedução {it} — {desc}")
        _inp(ws2, r, 2, 0)
    r_pla = 3 + len(itens_ded)
    ws2.cell(row=r_pla, column=1, value="PLA Calculado").font = Font(bold=True)
    ws2.cell(row=r_pla, column=2, value=f"=B2-SUM(B3:B{r_pla-1})").font = Font(bold=True)

    # --- Portfólio Preço Fixo ---
    ws3 = wb.create_sheet("Portfólio Preço Fixo")
    headers = ["Vértice", "Submercado", "Fonte", "Recurso (MWm)", "PM Recurso (R$/MWh)", "Requisito (MWm)", "PM Requisito (R$/MWh)"]
    col_widths = [8, 10, 14, 16, 22, 16, 22]
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        _hdr(ws3, 1, ci, h, w)
    row = 2
    for v in VERTICES:
        for s in SUBMERCADOS:
            for f in FONTES:
                ws3.cell(row=row, column=1, value=v)
                ws3.cell(row=row, column=2, value=s)
                ws3.cell(row=row, column=3, value=f)
                for ci in range(4, 8):
                    _inp(ws3, row, ci, 0)
                row += 1

    # --- Derivativos ---
    ws4 = wb.create_sheet("Derivativos")
    for ci, (h, w) in enumerate(zip(["Vértice", "Compra (MWm)", "Venda (MWm)"], [8, 14, 14]), 1):
        _hdr(ws4, 1, ci, h, w)
    for i, v in enumerate(VERTICES):
        ws4.cell(row=2+i, column=1, value=v)
        _inp(ws4, 2+i, 2, 0)
        _inp(ws4, 2+i, 3, 0)

    # --- Preço Variável ---
    ws5 = wb.create_sheet("Preço Variável")
    pv_hdrs = ["Vértice", "Recurso PV (MWm)", "PM Rec PV (R$/MWh)", "Requisito PV (MWm)", "PM Req PV (R$/MWh)"]
    pv_widths = [8, 18, 22, 18, 22]
    for ci, (h, w) in enumerate(zip(pv_hdrs, pv_widths), 1):
        _hdr(ws5, 1, ci, h, w)
    for i, v in enumerate(VERTICES):
        ws5.cell(row=2+i, column=1, value=v)
        for ci in range(2, 6):
            _inp(ws5, 2+i, ci, 0)

    # --- ACR ---
    ws6 = wb.create_sheet("ACR")
    for ci, (h, w) in enumerate(zip(["Vértice", "Receita ACR (R$)"], [8, 20]), 1):
        _hdr(ws6, 1, ci, h, w)
    ws6.cell(row=1, column=3, value="Nota: NÃO incluir CCEAR-Q (Quadro 9 Manual CCEE)").font = Font(italic=True, color="FF0000")
    for i, v in enumerate(VERTICES):
        ws6.cell(row=2+i, column=1, value=v)
        _inp(ws6, 2+i, 2, 0)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── Portfólio Extra — modelo e parser CSV ───────────────────────────────────
import csv as _csv
import unicodedata

EXTRA_SECOES = [
    ("Preco Fixo", "preco_fixo"),
    ("Preco Variavel", "preco_variavel"),
    ("Derivativos", "derivativos"),
]
EXTRA_CAMPOS = [
    ("Exposicao Submercado", "subm"),
    ("Recurso", "recurso"),
    ("PM Recurso", "pm_recurso"),
    ("Requisito", "requisito"),
    ("PM Requisito", "pm_requisito"),
]


def _norm(s):
    if s is None:
        return ""
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode("ascii")
    return s.strip().lower()


def _empty_extra_section():
    return {
        "subm": {s: {v: 0 for v in VERTICES} for s in SUBMERCADOS},
        "recurso": {v: 0 for v in VERTICES},
        "pm_recurso": {v: 0 for v in VERTICES},
        "requisito": {v: 0 for v in VERTICES},
        "pm_requisito": {v: 0 for v in VERTICES},
    }


def gerar_modelo_extra_csv():
    """Gera o CSV modelo do Portfólio Extra (separador ';', decimal '.')."""
    out = io.StringIO()
    w = _csv.writer(out, delimiter=";")
    w.writerow(["Secao", "Campo", "Submercado"] + VERTICES)
    for sec_label, _ in EXTRA_SECOES:
        for campo_label, campo_key in EXTRA_CAMPOS:
            if campo_key == "subm":
                for s in SUBMERCADOS:
                    w.writerow([sec_label, campo_label, s] + [0]*len(VERTICES))
            else:
                w.writerow([sec_label, campo_label, ""] + [0]*len(VERTICES))
    # EFM
    w.writerow(["EFM", "Efeitos Financeiros Mercado Regulado", ""] + [0]*len(VERTICES))
    return out.getvalue()


def _parse_num(x):
    if x is None or str(x).strip() == "":
        return 0.0
    s = str(x).strip().replace(" ", "")
    # aceita decimal com vírgula ou ponto
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return 0.0


def parse_extra_csv(file_bytes):
    """Lê o CSV do Portfólio Extra (separador ';' ou ',') e devolve o dict."""
    errors = []
    extra = {
        "ativo": False,
        "nome": "Portfólio Extra (Simulação)",
        "preco_fixo": _empty_extra_section(),
        "preco_variavel": _empty_extra_section(),
        "derivativos": _empty_extra_section(),
        "efm_regulado": {v: 0 for v in VERTICES},
    }
    try:
        text = file_bytes.decode("utf-8-sig", errors="replace")
    except Exception:
        text = file_bytes.decode("latin-1", errors="replace")

    # detecta separador
    first = text.splitlines()[0] if text.splitlines() else ""
    delim = ";" if first.count(";") >= first.count(",") else ","

    sec_map = {_norm(lbl): key for lbl, key in EXTRA_SECOES}
    campo_map = {_norm(lbl): key for lbl, key in EXTRA_CAMPOS}
    subm_map = {_norm(s): s for s in SUBMERCADOS}
    subm_map[_norm("SUDESTE/CENTRO-OESTE")] = "SE/CO"

    try:
        reader = _csv.reader(io.StringIO(text), delimiter=delim)
        rows = list(reader)
        if not rows:
            return extra, ["CSV vazio."]
        # pula cabeçalho se for texto
        start = 1 if rows and _norm(rows[0][0]) in ("secao", "seção") else 0
        for r in rows[start:]:
            if not r or all((c is None or str(c).strip() == "") for c in r):
                continue
            sec = _norm(r[0]) if len(r) > 0 else ""
            campo = _norm(r[1]) if len(r) > 1 else ""
            subm = _norm(r[2]) if len(r) > 2 else ""
            vals = [_parse_num(r[3+i]) if len(r) > 3+i else 0.0 for i in range(len(VERTICES))]

            if sec == "efm" or "efeitos financeiros" in campo or "efm" in sec:
                for i, v in enumerate(VERTICES):
                    extra["efm_regulado"][v] = vals[i]
                continue

            sec_key = sec_map.get(sec)
            campo_key = campo_map.get(campo)
            if not sec_key or not campo_key:
                continue
            target = extra[sec_key]
            if campo_key == "subm":
                s_real = subm_map.get(subm)
                if s_real:
                    for i, v in enumerate(VERTICES):
                        target["subm"][s_real][v] = vals[i]
            else:
                for i, v in enumerate(VERTICES):
                    target[campo_key][v] = vals[i]
    except Exception as e:
        errors.append(f"Erro ao ler CSV: {e}")

    return extra, errors


# ===== INTERFACE (HTML embutido, identico ao original) =====
EMBEDDED_HTML = r'''<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>Simulador FA CCEE</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
:root{
  /* Superfícies — papel quente, não o navy genérico de dashboard IA */
  --bg:#f3f1ea;--surf:#fffdf8;--surf2:#efebe0;--surf3:#f8f5ee;
  --border:#e2ddcf;
  /* Marca — verde-petróleo + terracota */
  --acc:#1a6a63;--acc-d:#135751;--amb:#b16a3a;
  /* Tinta */
  --text:#222b2a;--muted:#6c756f;
  /* Semáforo — tons terrosos, não as cores puras de IA */
  --green:#2f7d54;--yellow:#b3852b;--red:#b23a31;
  --r:9px;
  --font:ui-sans-serif,'Segoe UI',Roboto,system-ui,sans-serif;
  --serif:'Iowan Old Style',Georgia,'Times New Roman',serif;
  --mono:ui-monospace,'Cascadia Mono','Consolas',monospace;
  --sh:0 1px 2px rgba(34,43,42,.04),0 2px 6px rgba(34,43,42,.05);
  --sh-lg:0 2px 8px rgba(34,43,42,.06),0 8px 24px rgba(34,43,42,.06);
}
*{box-sizing:border-box;margin:0;padding:0;}
body{background:var(--bg);color:var(--text);font-family:var(--font);font-size:13px;
  -webkit-font-smoothing:antialiased;}
.num,td,.metric .val,.fa-big .num,.hbadge{font-variant-numeric:tabular-nums;}

/* ---- HEADER ---- */
header{background:var(--surf);border-bottom:1px solid var(--border);padding:14px 22px;
  display:flex;align-items:center;gap:14px;box-shadow:var(--sh);}
header h1{font-family:var(--serif);font-size:21px;color:var(--text);font-weight:600;
  white-space:nowrap;letter-spacing:.2px;position:relative;padding-bottom:2px;}
header h1::after{content:"";position:absolute;left:0;bottom:-3px;width:34px;height:2px;
  background:var(--acc);border-radius:2px;}
#hdr-info{color:var(--muted);font-size:11px;letter-spacing:.2px;}
#hdr-badges{margin-left:auto;display:flex;gap:8px;align-items:center;}
.hbadge{padding:7px 14px;border-radius:7px;font-weight:600;font-size:13px;border:1px solid;
  font-family:var(--mono);}
.hb-gray{background:#eceee9;color:var(--muted);border-color:var(--border);}
.hb-green{background:rgba(47,125,84,.12);color:var(--green);border-color:rgba(47,125,84,.35);}
.hb-yellow{background:rgba(179,133,43,.12);color:var(--yellow);border-color:rgba(179,133,43,.35);}
.hb-red{background:rgba(178,58,49,.12);color:var(--red);border-color:rgba(178,58,49,.35);}

/* ---- NAV ---- */
nav{background:var(--surf);border-bottom:1px solid var(--border);display:flex;
  padding:0 12px;box-shadow:var(--sh);}
nav button{background:none;border:none;color:var(--muted);padding:12px 18px;cursor:pointer;
  font-size:13px;font-weight:500;border-bottom:2px solid transparent;margin-bottom:-1px;
  transition:.15s;}
nav button.active{color:var(--acc);border-bottom-color:var(--acc);font-weight:600;}
nav button:hover:not(.active){color:var(--text);}

main{padding:22px;max-width:1500px;margin:0 auto;}
.tab{display:none;}.tab.active{display:block;}

/* ---- CARDS ---- */
.card{background:var(--surf);border:1px solid var(--border);border-radius:var(--r);
  padding:18px;margin-bottom:16px;box-shadow:var(--sh);}
.card-title{font-size:11px;font-weight:700;color:var(--acc);margin-bottom:14px;
  display:flex;align-items:center;gap:9px;text-transform:uppercase;letter-spacing:.9px;}
.card-title::before{content:"";width:3px;height:14px;background:var(--acc);
  border-radius:2px;flex:none;}
.section-label{background:rgba(177,106,58,.12);color:var(--amb);padding:3px 10px;
  border-radius:5px;font-size:11px;font-weight:700;letter-spacing:.5px;}

/* ---- GRID ---- */
.g2{display:grid;grid-template-columns:1fr 1fr;gap:16px;}
.g3{display:grid;grid-template-columns:1fr 1fr 1fr;gap:16px;}
.g4{display:grid;grid-template-columns:repeat(4,1fr);gap:12px;}
.g6{display:grid;grid-template-columns:repeat(6,1fr);gap:12px;}

/* ---- FORM ---- */
label{display:block;color:var(--muted);font-size:11px;margin-bottom:4px;letter-spacing:.2px;}
input[type=number],input[type=text],input[type=date]{
  width:100%;background:var(--surf3);border:1px solid var(--border);
  color:var(--text);padding:7px 9px;border-radius:6px;font-size:12px;
  transition:border-color .15s,box-shadow .15s;}
input:focus{outline:none;border-color:var(--acc);box-shadow:0 0 0 3px rgba(26,106,99,.12);}
/* Remove as setinhas (spin buttons) dos campos numericos — mais limpo para digitar */
input[type=number]::-webkit-inner-spin-button,
input[type=number]::-webkit-outer-spin-button{-webkit-appearance:none;margin:0;}
input[type=number]{-moz-appearance:textfield;appearance:textfield;}

/* ---- BUTTONS ---- */
.btn{background:var(--acc);color:#fff;border:none;padding:8px 17px;
  border-radius:7px;cursor:pointer;font-size:12px;font-weight:600;
  box-shadow:var(--sh);transition:background .15s;}
.btn:hover{background:var(--acc-d);}
.btn-out{background:var(--surf);color:var(--acc);border:1px solid var(--border);
  padding:7px 15px;border-radius:7px;cursor:pointer;font-size:12px;font-weight:600;
  transition:.15s;}
.btn-out:hover{border-color:var(--acc);background:rgba(26,106,99,.06);}
.btn-amb{background:var(--amb);color:#fff;border:none;padding:8px 17px;
  border-radius:7px;cursor:pointer;font-size:12px;font-weight:600;box-shadow:var(--sh);
  transition:filter .15s;}
.btn-amb:hover{filter:brightness(.93);}
.btn-row{display:flex;gap:9px;align-items:center;flex-wrap:wrap;margin-top:12px;}

/* ---- TOGGLE SWITCH ---- */
.switch{position:relative;display:inline-block;width:50px;height:26px;flex:none;}
.switch input{opacity:0;width:0;height:0;}
.slider{position:absolute;cursor:pointer;inset:0;background:#cdc6b6;border-radius:26px;
  transition:.2s;}
.slider::before{content:"";position:absolute;height:20px;width:20px;left:3px;bottom:3px;
  background:#fff;border-radius:50%;transition:.2s;box-shadow:0 1px 2px rgba(0,0,0,.2);}
.switch input:checked + .slider{background:var(--acc);}
.switch input:checked + .slider::before{transform:translateX(24px);}

/* ---- TABLES ---- */
.tbl-wrap{overflow-x:auto;border:1px solid var(--border);border-radius:var(--r);}
table{width:100%;border-collapse:collapse;font-size:12px;}
thead tr{background:var(--surf2);}
th{padding:8px 10px;text-align:right;font-weight:600;color:var(--muted);
  border-bottom:1px solid var(--border);white-space:nowrap;font-size:11px;
  letter-spacing:.3px;}
th:first-child,th:nth-child(2){text-align:left;}
td{padding:6px 9px;border-bottom:1px solid #eee8db;text-align:right;vertical-align:middle;}
td:first-child,td:nth-child(2){text-align:left;color:var(--muted);}
tbody tr:nth-child(even) td{background:rgba(239,235,224,.4);}
tbody tr:hover td{background:rgba(26,106,99,.06);}
td input{width:104px;background:var(--surf);border:1px solid var(--border);
  color:var(--text);padding:6px 9px;border-radius:6px;font-size:12px;text-align:right;
  font-variant-numeric:tabular-nums;transition:border-color .15s,box-shadow .15s,background .15s;}
td input:hover{border-color:#c9c2b0;}
td input:focus{border-color:var(--acc);outline:none;background:var(--surf3);
  box-shadow:0 0 0 3px rgba(26,106,99,.14);}
.row-section td{background:rgba(177,106,58,.1)!important;color:var(--amb)!important;
  font-weight:700;font-size:11px;letter-spacing:.6px;text-align:left!important;
  text-transform:uppercase;}
.row-total td{background:rgba(26,106,99,.09)!important;font-weight:700;}
.row-check td{background:rgba(47,125,84,.1)!important;color:var(--green)!important;}
.row-pmedio td{background:rgba(177,106,58,.05)!important;}
.col-v{min-width:100px;}

/* ---- METRICS ---- */
.metric{background:var(--surf3);border:1px solid var(--border);border-radius:var(--r);
  padding:13px;text-align:center;}
.metric .val{font-size:18px;font-weight:700;font-family:var(--mono);letter-spacing:-.3px;}
.metric .lbl{font-size:10px;color:var(--muted);margin-top:4px;letter-spacing:.2px;}

/* ---- FA BADGE ---- */
.fa-big{display:flex;flex-direction:column;align-items:center;justify-content:center;
  padding:18px 26px;border-radius:var(--r);border:1px solid;min-width:160px;}
.fa-big .num{font-size:30px;font-weight:700;font-family:var(--mono);letter-spacing:-.5px;}
.fa-big .sub{font-size:11px;font-weight:500;opacity:.85;margin-top:3px;letter-spacing:.3px;}
.fa-gray{background:#eceee9;color:var(--muted);border-color:var(--border);}
.fa-green{background:rgba(47,125,84,.1);color:var(--green);border-color:rgba(47,125,84,.35);}
.fa-yellow{background:rgba(179,133,43,.1);color:var(--yellow);border-color:rgba(179,133,43,.35);}
.fa-red{background:rgba(178,58,49,.1);color:var(--red);border-color:rgba(178,58,49,.35);}

/* ---- MESSAGES ---- */
.msg{padding:9px 13px;border-radius:7px;margin:8px 0;font-size:12px;}
.msg-ok{background:rgba(47,125,84,.1);color:var(--green);border:1px solid rgba(47,125,84,.3);}
.msg-err{background:rgba(178,58,49,.1);color:var(--red);border:1px solid rgba(178,58,49,.3);}
.msg-info{background:rgba(26,106,99,.1);color:var(--acc);border:1px solid rgba(26,106,99,.3);}

/* ---- UPLOAD BOX ---- */
.upload-drop{border:1.5px dashed #cbc3b0;border-radius:var(--r);padding:18px;
  text-align:center;cursor:pointer;transition:.2s;background:var(--surf3);}
.upload-drop:hover{border-color:var(--acc);background:rgba(26,106,99,.04);}
.upload-drop input{display:none;}

/* ---- CHART ---- */
.chart-h{position:relative;height:230px;}

/* ---- SEMAFORO ---- */
.sema{width:11px;height:11px;border-radius:50%;display:inline-block;margin-right:4px;
  vertical-align:middle;}
.s-green{background:var(--green);}.s-yellow{background:var(--yellow);}.s-red{background:var(--red);}

footer{text-align:center;padding:20px;color:var(--muted);font-size:10px;
  border-top:1px solid var(--border);margin-top:28px;letter-spacing:.2px;}
</style>
</head>
<body>

<header>
  <h1>Simulador FA CCEE</h1>
  <span id="hdr-info">—</span>
  <div id="hdr-badges">
    <span id="hb-ris" class="hbadge hb-gray">FA Risco —</span>
    <span id="hb-div" class="hbadge hb-gray">FA Divulgado —</span>
    <span id="hb-rwa" class="hbadge hb-gray">RWA —</span>
    <span id="hb-pla" class="hbadge hb-gray">PLA —</span>
  </div>
</header>

<nav>
  <button class="active" onclick="showTab('premissas')">Premissas</button>
  <button onclick="showTab('pla')">Patrimônio Líquido Ajustado</button>
  <button onclick="showTab('portfolio')">Declaração Portfólio</button>
  <button onclick="showTab('extra')">Portfólio Extra</button>
  <button onclick="showTab('calculo')">Cálculo VaR / Consolidado</button>
  <button onclick="showTab('historico')">Histórico</button>
</nav>

<main>

<!-- ===================== PREMISSAS ===================== -->
<div id="tab-premissas" class="tab active">
  <div class="g2">
    <div class="card">
      <div class="card-title">📊 Atualização Semanal — Planilha CCEE/DCIDE</div>
      <div class="upload-drop" onclick="document.getElementById('up-ccee').click()">
        <input type="file" id="up-ccee" accept=".xlsx" onchange="uploadCCEE(this)">
        <div style="font-size:28px;margin-bottom:6px;">⬆</div>
        <div style="font-weight:600;">Carregar planilha semanal CCEE (.xlsx)</div>
        <div style="color:var(--muted);font-size:11px;margin-top:4px;">
          Atualiza Forward, Volatilidades, Stress, Horas e PLD automaticamente
        </div>
      </div>
      <div id="msg-ccee"></div>
    </div>
    <div class="card">
      <div class="card-title">⚙ Parâmetros</div>
      <div class="g2">
        <div><label>Data de Referência</label><input type="date" id="p-data" onchange="premChanged()"></div>
        <div><label>Intervalo de Confiança (φ)</label><input type="number" id="p-phi" step="0.0001" onchange="premChanged()"></div>
        <div><label>Dias para Liquidação (D)</label><input type="number" id="p-d" step="1" min="1" onchange="premChanged()"></div>
        <div><label>PLD Mínimo (R$/MWh)</label><input type="number" id="p-pldmin" step="0.01" onchange="premChanged()"></div>
        <div><label>PLD Máximo (R$/MWh)</label><input type="number" id="p-pldmax" step="0.01" onchange="premChanged()"></div>
        <div><label>Correlação (ρ)</label><input type="number" id="p-rho" step="0.01" value="1" onchange="premChanged()"></div>
      </div>
      <div class="btn-row"><button class="btn" onclick="salvarPremissas()">Salvar Premissas</button></div>
    </div>
  </div>

  <div class="card">
    <div class="card-title">📈 Curva Forward DCIDE (R$/MWh)</div>
    <div class="tbl-wrap" id="tbl-forward"></div>
  </div>

  <div class="card">
    <div class="card-title">📉 Volatilidades, Preços de Estresse e Horas</div>
    <div class="tbl-wrap" id="tbl-vsh"></div>
  </div>
</div>

<!-- ===================== PLA ===================== -->
<div id="tab-pla" class="tab">
  <div class="card">
    <div class="card-title">🏦 Patrimônio Líquido Ajustado — Anexo I Manual CCEE v2023.2.0
      <a href="https://www2.aneel.gov.br/cedoc/aren20231072_2.pdf" target="_blank"
         style="font-size:10px;color:var(--acc);margin-left:8px;">Ver Manual ↗</a>
    </div>
    <table>
      <thead><tr>
        <th style="width:50px;">Item</th>
        <th style="width:40%;">Descrição</th>
        <th style="width:200px;">Valor (R$)</th>
      </tr></thead>
      <tbody>
        <tr>
          <td style="font-weight:700;">PL Bruto</td>
          <td style="color:var(--text);">Patrimônio Líquido Bruto</td>
          <td><input type="number" id="pla-bruto" step="1000" oninput="plaChanged()" placeholder="0"></td>
        </tr>
        <tr class="row-section"><td colspan="3">DEDUÇÕES — Anexo I do Manual de Monitoramento Prudencial</td></tr>
        <tr id="ded-row-0"><td>I</td><td><input type="text" id="ded-desc-0" placeholder="Descrição (opcional)" style="width:100%;"></td><td><input type="number" id="ded-val-0" step="1000" oninput="plaChanged()" placeholder="0"></td></tr>
        <tr id="ded-row-1"><td>II</td><td><input type="text" id="ded-desc-1" placeholder="Descrição (opcional)" style="width:100%;"></td><td><input type="number" id="ded-val-1" step="1000" oninput="plaChanged()" placeholder="0"></td></tr>
        <tr id="ded-row-2"><td>III</td><td><input type="text" id="ded-desc-2" placeholder="Descrição (opcional)" style="width:100%;"></td><td><input type="number" id="ded-val-2" step="1000" oninput="plaChanged()" placeholder="0"></td></tr>
        <tr id="ded-row-3"><td>IV</td><td><input type="text" id="ded-desc-3" placeholder="Descrição (opcional)" style="width:100%;"></td><td><input type="number" id="ded-val-3" step="1000" oninput="plaChanged()" placeholder="0"></td></tr>
        <tr id="ded-row-4"><td>V</td><td><input type="text" id="ded-desc-4" placeholder="Descrição (opcional)" style="width:100%;"></td><td><input type="number" id="ded-val-4" step="1000" oninput="plaChanged()" placeholder="0"></td></tr>
        <tr id="ded-row-5"><td>VI</td><td><input type="text" id="ded-desc-5" placeholder="Descrição (opcional)" style="width:100%;"></td><td><input type="number" id="ded-val-5" step="1000" oninput="plaChanged()" placeholder="0"></td></tr>
        <tr id="ded-row-6"><td>VII</td><td><input type="text" id="ded-desc-6" placeholder="Descrição (opcional)" style="width:100%;"></td><td><input type="number" id="ded-val-6" step="1000" oninput="plaChanged()" placeholder="0"></td></tr>
        <tr id="ded-row-7"><td>VIII</td><td><input type="text" id="ded-desc-7" placeholder="Descrição (opcional)" style="width:100%;"></td><td><input type="number" id="ded-val-7" step="1000" oninput="plaChanged()" placeholder="0"></td></tr>
        <tr class="row-total">
          <td colspan="2" style="text-align:right;font-weight:700;color:var(--text);">PLA = PL Bruto − Σ Deduções</td>
          <td style="text-align:right;font-size:16px;font-weight:900;" id="pla-result">R$ 0</td>
        </tr>
      </tbody>
    </table>
    <div class="btn-row"><button class="btn" onclick="salvarEmpresa()">Salvar PLA</button></div>
  </div>
</div>

<!-- ===================== PORTFÓLIO ===================== -->
<div id="tab-portfolio" class="tab">
  <div class="card" style="margin-bottom:10px;">
    <div class="card-title">📋 Declaração Portfólio</div>
    <div style="display:flex;gap:10px;align-items:center;flex-wrap:wrap;">
      <button class="btn-out" onclick="downloadModelo()">⬇ Baixar Modelo .xlsx</button>
      <div class="upload-drop" style="padding:8px 14px;display:inline-block;" onclick="document.getElementById('up-port').click()">
        <input type="file" id="up-port" accept=".xlsx" onchange="uploadPortfolio(this)">
        ⬆ Carregar Portfólio Preenchido
      </div>
      <div id="msg-port"></div>
      <button class="btn" style="margin-left:auto;" onclick="salvarEmpresa()">Salvar & Recalcular</button>
      <button class="btn-amb" onclick="salvarHistorico()">Registrar no Histórico</button>
    </div>
    <div id="msg-emp"></div>
  </div>

  <!-- SEÇÃO 1: PREÇO FIXO -->
  <div class="card">
    <div class="card-title"><span class="section-label">SEÇÃO 1</span> Preço Fixo, Consumo e Geração</div>
    <p style="color:var(--muted);font-size:11px;margin-bottom:8px;">
      Exposições vendidas são negativas. Recurso físico: marcar ao PM dos contratos de venda. Requisito físico: marcar ao PM dos contratos de compra.
    </p>
    <div class="tbl-wrap" id="tbl-pf"></div>
  </div>

  <!-- SEÇÃO 2: PREÇO VARIÁVEL -->
  <div class="card">
    <div class="card-title"><span class="section-label">SEÇÃO 2</span> Preço Variável (PLD)</div>
    <p style="color:var(--muted);font-size:11px;margin-bottom:8px;">
      Contratos atrelados ao PLD. PM deve ser apenas o spread ponderado pelo volume. Apenas o resultado financeiro é utilizado no FA.
    </p>
    <div class="tbl-wrap" id="tbl-pv"></div>
  </div>

  <!-- SEÇÃO 3: DERIVATIVOS -->
  <div class="card">
    <div class="card-title"><span class="section-label">SEÇÃO 3</span> Derivativos de Energia</div>
    <p style="color:var(--muted);font-size:11px;margin-bottom:8px;">
      Apenas derivativos de energia devem ser declarados nesta seção.
    </p>
    <div class="tbl-wrap" id="tbl-der"></div>
  </div>

  <!-- SEÇÃO 4: EFM REGULADO -->
  <div class="card">
    <div class="card-title"><span class="section-label">SEÇÃO 4</span> Efeitos Financeiros do Mercado Regulado</div>
    <p style="color:var(--muted);font-size:11px;margin-bottom:8px;">
      Receitas de ACR e outros efeitos financeiros do mercado regulado (R$). Não incluir CCEAR-Q.
    </p>
    <div class="tbl-wrap" id="tbl-efm"></div>
  </div>
</div>

<!-- ===================== PORTFÓLIO EXTRA ===================== -->
<div id="tab-extra" class="tab">
  <div class="card" style="margin-bottom:10px;">
    <div class="card-title">Portfólio Extra — Simulação de Hedges e Trocas</div>
    <p style="color:var(--muted);font-size:12px;margin-bottom:12px;">
      Lance aqui operações que <b>não estão no portfólio real</b> para testar hedges e trocas de resultado.
      Os volumes são <b>somados</b> e os preços médios entram por <b>média ponderada</b> com o portfólio real.
      Use o botão abaixo para incluir (ou não) este portfólio no cálculo do Fator de Alavancagem.
    </p>

    <!-- TOGGLE de ativação -->
    <div id="extra-toggle-box" style="display:flex;align-items:center;gap:14px;flex-wrap:wrap;
         padding:14px 16px;border-radius:9px;border:1px solid var(--border);background:var(--surf3);">
      <label class="switch" style="margin:0;">
        <input type="checkbox" id="extra-ativo" onchange="toggleExtra()">
        <span class="slider"></span>
      </label>
      <div style="flex:1;min-width:240px;">
        <div style="font-weight:700;font-size:13px;color:var(--text);" id="extra-toggle-label">
          Portfólio Extra DESATIVADO — FA usa só o portfólio real
        </div>
        <div style="font-size:11px;color:var(--muted);" id="extra-toggle-sub">
          Ative para somar este portfólio ao real no cálculo do Fator de Alavancagem.
        </div>
      </div>
      <span id="extra-status-badge" class="hbadge hb-gray">Inativo</span>
    </div>

    <div style="display:flex;gap:10px;align-items:center;flex-wrap:wrap;margin-top:12px;">
      <button class="btn-out" onclick="downloadModeloExtra()">⬇ Baixar Modelo CSV</button>
      <div class="upload-drop" style="padding:8px 14px;display:inline-block;" onclick="document.getElementById('up-extra').click()">
        <input type="file" id="up-extra" accept=".csv" onchange="uploadExtraCSV(this)">
        ⬆ Carregar CSV Preenchido
      </div>
      <button class="btn" style="margin-left:auto;" onclick="salvarExtra()">Salvar Portfólio Extra</button>
      <button class="btn-out" onclick="limparExtra()">Limpar Tudo</button>
    </div>
    <div id="msg-extra"></div>
  </div>

  <!-- COMPARATIVO Real vs Combinado -->
  <div class="card">
    <div class="card-title">Comparativo — Real vs Real + Extra</div>
    <div class="tbl-wrap" id="extra-compare"></div>
  </div>

  <!-- SEÇÃO 1: PREÇO FIXO -->
  <div class="card">
    <div class="card-title"><span class="section-label">SEÇÃO 1</span> Preço Fixo, Consumo e Geração</div>
    <p style="color:var(--muted);font-size:11px;margin-bottom:8px;">
      Exposições vendidas são negativas. Recurso = venda; Requisito = compra.
    </p>
    <div class="tbl-wrap" id="tbl-pf-x"></div>
  </div>

  <!-- SEÇÃO 2: PREÇO VARIÁVEL -->
  <div class="card">
    <div class="card-title"><span class="section-label">SEÇÃO 2</span> Preço Variável (PLD)</div>
    <div class="tbl-wrap" id="tbl-pv-x"></div>
  </div>

  <!-- SEÇÃO 3: DERIVATIVOS -->
  <div class="card">
    <div class="card-title"><span class="section-label">SEÇÃO 3</span> Derivativos de Energia</div>
    <div class="tbl-wrap" id="tbl-der-x"></div>
  </div>

  <!-- SEÇÃO 4: EFM REGULADO -->
  <div class="card">
    <div class="card-title"><span class="section-label">SEÇÃO 4</span> Efeitos Financeiros do Mercado Regulado</div>
    <div class="tbl-wrap" id="tbl-efm-x"></div>
  </div>
</div>

<!-- ===================== CÁLCULO ===================== -->
<div id="tab-calculo" class="tab">
  <!-- FA Result -->
  <div class="card">
    <div class="card-title">🏆 Fator de Alavancagem — Resultado</div>
    <div style="display:flex;gap:16px;align-items:flex-start;flex-wrap:wrap;">
      <div id="fa-ris-badge" class="fa-big fa-gray"><div class="num">—</div><div class="sub">FA Risco (RWA / PLA)</div></div>
      <div id="fa-div-badge" class="fa-big fa-gray"><div class="num">—</div><div class="sub">FA Divulgado</div></div>
      <div style="flex:1;min-width:300px;">
        <div class="g4">
          <div class="metric"><div class="val" id="r-rwa">—</div><div class="lbl">RWA (R$)</div></div>
          <div class="metric"><div class="val" id="r-resfin">—</div><div class="lbl">Resultado Financeiro (R$)</div></div>
          <div class="metric"><div class="val" id="r-pla">—</div><div class="lbl">PLA (R$)</div></div>
          <div class="metric"><div class="val" id="r-var">—</div><div class="lbl">VaR Total (R$)</div></div>
        </div>
        <div class="g4" style="margin-top:8px;">
          <div class="metric"><div class="val" id="r-mtm">—</div><div class="lbl">MtM (R$)</div></div>
          <div class="metric"><div class="val" id="r-rescontr">—</div><div class="lbl">Res. Contratual (R$)</div></div>
          <div class="metric"><div class="val" id="r-finpv">—</div><div class="lbl">Fin. PLD+ (R$)</div></div>
          <div class="metric"><div class="val" id="r-efm">—</div><div class="lbl">EFM Regulado (R$)</div></div>
        </div>
      </div>
    </div>
    <div style="margin-top:10px;font-size:11px;color:var(--muted);">
      <span class="sema s-green"></span>FA &lt; 1 — Normal &nbsp;|&nbsp;
      <span class="sema s-yellow"></span>1 ≤ FA ≤ 3 — Atenção &nbsp;|&nbsp;
      <span class="sema s-red"></span>FA &gt; 3 — Crítico &nbsp;|&nbsp;
      K=0 e θ=0 (período sombra) &nbsp;|&nbsp; Manual CCEE v2023.2.0
    </div>
  </div>

  <!-- Cálculo VaR por vértice -->
  <div class="card">
    <div class="card-title">📊 Cálculo VaR e Teste de Estresse — por Vértice</div>
    <div class="tbl-wrap" id="tbl-var"></div>
  </div>

  <!-- Consolidado -->
  <div class="card">
    <div class="card-title">📋 Consolidado — Exposições e Resultados por Vértice</div>
    <div class="tbl-wrap" id="tbl-consolidado"></div>
  </div>

  <!-- Gráficos -->
  <div class="g2" style="margin-top:14px;">
    <div class="card"><div class="card-title">Exposição NET por Vértice (MWm)</div><div class="chart-h"><canvas id="ch-exp"></canvas></div></div>
    <div class="card"><div class="card-title">MtM por Vértice (R$)</div><div class="chart-h"><canvas id="ch-mtm"></canvas></div></div>
    <div class="card"><div class="card-title">VaR por Vértice (R$)</div><div class="chart-h"><canvas id="ch-var"></canvas></div></div>
    <div class="card"><div class="card-title">Stress Test por Vértice (R$)</div><div class="chart-h"><canvas id="ch-stest"></canvas></div></div>
  </div>
</div>

<!-- ===================== HISTÓRICO ===================== -->
<div id="tab-historico" class="tab">
  <div class="card">
    <div class="card-title">📅 Série Temporal Semanal</div>
    <div class="btn-row" style="margin-bottom:12px;">
      <button class="btn-out" onclick="exportCSV()">⬇ Exportar CSV</button>
    </div>
    <div class="tbl-wrap" id="tbl-hist"></div>
  </div>
  <div class="g2">
    <div class="card"><div class="card-title">FA Risco e FA Divulgado</div><div class="chart-h"><canvas id="ch-hfa"></canvas></div></div>
    <div class="card"><div class="card-title">RWA × PLA (R$)</div><div class="chart-h"><canvas id="ch-hrwa"></canvas></div></div>
  </div>
</div>

</main>

<footer>Simulação baseada no Manual CCEE v2023.2.0 — K=0 e θ=0 (período sombra) — Curva Forward DCIDE/CCEE — Uso interno — Valores oficiais apurados pela CCEE</footer>

<script>
// ============================================================
// CONSTANTS
// ============================================================
const VX = ["M+0","M+1","M+2","M+3","M+4","M+5","M+6"];
const SUBMS = ["SE/CO","SUL","NE","N"];
const SUBM_LABELS = {"SE/CO":"SUDESTE/CENTRO-OESTE","SUL":"SUL","NE":"NORDESTE","N":"NORTE"};
const DED_ITENS = ["I","II","III","IV","V","VI","VII","VIII"];

// ============================================================
// STATE
// ============================================================
let premissas = {}, emp = {}, lastEmp = null;
let extraData = null, lastExtraPreview = null;
let charts = {};

// ============================================================
// HELPERS
// ============================================================
const fmtN = (v,d=2) => v==null||v===undefined ? '—' : Number(v).toLocaleString('pt-BR',{minimumFractionDigits:d,maximumFractionDigits:d});
const fmtR = (v,d=0) => v==null ? '—' : 'R$ '+fmtN(v,d);
const fmtM = v => v==null ? '—' : fmtR(v/1e6,2)+' mi';
const faClass = v => v==null ? 'fa-gray' : v<1 ? 'fa-green' : v<=3 ? 'fa-yellow' : 'fa-red';
const hbClass = v => v==null ? 'hb-gray' : v<1 ? 'hb-green' : v<=3 ? 'hb-yellow' : 'hb-red';
function showMsg(id, txt, type='ok'){
  const el=document.getElementById(id); if(!el) return;
  el.innerHTML=`<div class="msg msg-${type}">${txt}</div>`;
  if(type==='ok') setTimeout(()=>el.innerHTML='',5000);
}
function inp(id, def=0){ const e=document.getElementById(id); return e ? (parseFloat(e.value)||def) : def; }

// ============================================================
// TABS
// ============================================================
const TABS = ['premissas','pla','portfolio','extra','calculo','historico'];
function showTab(name){
  document.querySelectorAll('.tab').forEach(e=>e.classList.remove('active'));
  document.querySelectorAll('nav button').forEach(b=>b.classList.remove('active'));
  document.getElementById('tab-'+name).classList.add('active');
  document.querySelectorAll('nav button')[TABS.indexOf(name)].classList.add('active');
  if(name==='calculo') renderCalculo();
  if(name==='historico') loadHistorico();
  if(name==='extra') calcExtra();
}
// ============================================================
// INIT
// ============================================================
async function init(){
  premissas = await fetch('/api/premissas').then(r=>r.json());
  emp = await fetch('/api/empresa').then(r=>r.json());
  extraData = await fetch('/api/portfolio_extra').then(r=>r.json());
  renderPremissas();
  loadEmpresaForm();
  renderPortfolioTables('emp', emp, triggerCalcEmp);
  renderPortfolioTables('extra', extraData, triggerCalcExtra, '-x');
  syncExtraToggle();
  await calcularEmp();
}

// ============================================================
// PREMISSAS
// ============================================================
function renderPremissas(){
  const p = premissas;
  document.getElementById('p-data').value = p.data_referencia||'';
  document.getElementById('p-phi').value = p.phi_norm||'-1.6449';
  document.getElementById('p-d').value = p.dias_liquidacao||1;
  document.getElementById('p-pldmin').value = p.pld_min||57.31;
  document.getElementById('p-pldmax').value = p.pld_max||785.27;
  document.getElementById('hdr-info').textContent = `Premissas: ${p.data_referencia||'—'} | φ=${p.phi_norm||''} | D=${p.dias_liquidacao||''} | PLD [${p.pld_min||''}–${p.pld_max||''}]`;

  // Forward table
  const fwd = p.forward||{};
  const keys = ['SECO','SUL','NE','N','I0','I5','I8','I1'];
  const labels = ['SECO/CONV (R$/MWh)','SUL – spread (R$/MWh)','NORDESTE – spread (R$/MWh)','NORTE – spread (R$/MWh)','INCENTIVADA 0% (R$/MWh)','INCENTIVADA 50%/CQ5 (R$/MWh)','INCENTIVADA 80% (R$/MWh)','INCENTIVADA 100% (R$/MWh)'];
  let h = `<table><thead><tr><th>Curva Forward DCIDE</th>`+VX.map(v=>`<th class="col-v">${v}</th>`).join('')+`</tr></thead><tbody>`;
  keys.forEach((k,i)=>{ h+=`<tr><td>${labels[i]}</td>`+VX.map(v=>`<td>${fmtN(fwd[v]?.[k])}</td>`).join('')+`</tr>`; });
  h+=`</tbody></table>`;
  document.getElementById('tbl-forward').innerHTML=h;

  // Vol/Stress/Horas table
  const vol=p.volatilidades||{}, sl=p.stress_long||{}, ss=p.stress_short||{}, hr=p.horas||{};
  let h2=`<table><thead><tr><th>Parâmetro / Vértice</th>`+VX.map(v=>`<th>${v}</th>`).join('')+`</tr></thead><tbody>`;
  h2+=`<tr><td>Volatilidade σ (%)</td>`+VX.map(v=>`<td>${fmtN((vol[v]||0)*100,3)}</td>`).join('')+`</tr>`;
  h2+=`<tr><td>Preço de Estresse Long (R$/MWh)</td>`+VX.map(v=>`<td>${fmtN(sl[v])}</td>`).join('')+`</tr>`;
  h2+=`<tr><td>Preço de Estresse Short (R$/MWh)</td>`+VX.map(v=>`<td>${fmtN(ss[v])}</td>`).join('')+`</tr>`;
  h2+=`<tr><td>Horas / Mês</td>`+VX.map(v=>`<td>${hr[v]||0}</td>`).join('')+`</tr>`;
  h2+=`</tbody></table>`;
  document.getElementById('tbl-vsh').innerHTML=h2;
}

function premChanged(){
  premissas.data_referencia = document.getElementById('p-data').value;
  premissas.phi_norm = parseFloat(document.getElementById('p-phi').value)||0;
  premissas.dias_liquidacao = parseInt(document.getElementById('p-d').value)||1;
  premissas.pld_min = parseFloat(document.getElementById('p-pldmin').value)||0;
  premissas.pld_max = parseFloat(document.getElementById('p-pldmax').value)||0;
}

async function salvarPremissas(){
  premChanged();
  await fetch('/api/premissas',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(premissas)});
  showMsg('msg-ccee','Premissas salvas.');
  renderPremissas();
  await calcularEmp();
}

async function uploadCCEE(input){
  const file=input.files[0]; if(!file) return;
  showMsg('msg-ccee','Carregando...','info');
  try {
    const buf = await file.arrayBuffer();
    const res = await fetch('/api/upload/premissas',{
      method:'POST',
      headers:{'Content-Type':'application/octet-stream'},
      body: buf
    }).then(r=>r.json());
    if(res.ok){
      premissas=await fetch('/api/premissas').then(r=>r.json());
      renderPremissas();
      showMsg('msg-ccee',`✔ Planilha CCEE importada — Referência: ${res.data.data_referencia||'?'} | φ=${res.data.phi_norm||''}`,'ok');
      await calcularEmp();
    } else {
      showMsg('msg-ccee','Erro na importação: '+(res.errors||[]).join('; '),'err');
    }
  } catch(e) {
    showMsg('msg-ccee','Falha na requisição: '+e.message,'err');
  }
  input.value='';
}

// ============================================================
// PLA FORM
// ============================================================
function loadEmpresaForm(){
  const p=emp.pla||{};
  document.getElementById('pla-bruto').value = p.pl_bruto||0;
  (p.deducoes||[]).forEach((d,i)=>{
    const di=document.getElementById(`ded-desc-${i}`); if(di) di.value=d.descricao||'';
    const vi=document.getElementById(`ded-val-${i}`); if(vi) vi.value=d.valor||0;
  });
  plaChanged();
}

function plaChanged(){
  const pl = inp('pla-bruto');
  const ded = DED_ITENS.reduce((s,_,i)=>s+inp(`ded-val-${i}`),0);
  const pla = pl-ded;
  document.getElementById('pla-result').textContent = fmtR(pla);
  document.getElementById('pla-result').style.color = pla>0 ? 'var(--green)' : pla<0 ? 'var(--red)' : 'var(--muted)';
  // update emp object
  emp.pla = { pl_bruto:pl, deducoes: DED_ITENS.map((_,i)=>({item:DED_ITENS[i],valor:inp(`ded-val-${i}`),descricao:document.getElementById(`ded-desc-${i}`).value})) };
  triggerCalcEmp();
}

// ============================================================
// PORTFOLIO TABLES — reusuable for emp and sim
// ============================================================
function renderPortfolioTables(ctx, data, onChange, sfx){
  sfx = sfx || '';
  renderSection(ctx, 'pf', data.preco_fixo, 'tbl-pf'+sfx, onChange);
  renderSection(ctx, 'pv', data.preco_variavel, 'tbl-pv'+sfx, onChange);
  renderDerSection(ctx, data.derivativos, 'tbl-der'+sfx, onChange);
  renderEFM(ctx, data.efm_regulado, 'tbl-efm'+sfx, onChange);
}

function renderSection(ctx, sec, secData, containerId, onChange){
  const el=document.getElementById(containerId); if(!el) return;
  const hdr=`<table><thead><tr>
    <th style="min-width:220px;">Exposições</th>
    <th style="width:80px;">Unid.</th>
    ${VX.map(v=>`<th class="col-v">${v}</th>`).join('')}
  </tr></thead><tbody>`;

  let rows='';
  // Submercado NET rows
  rows+=`<tr class="row-section"><td colspan="${2+VX.length}">EXPOSIÇÃO LÍQUIDA POR SUBMERCADO (MWm) — posição vendida = negativa</td></tr>`;
  SUBMS.forEach(s=>{
    rows+=`<tr><td>${SUBM_LABELS[s]}</td><td>MWm</td>`;
    VX.forEach(v=>{
      const val=(secData?.subm?.[s]?.[v])||0;
      rows+=`<td><input type="number" step="0.1" value="${val}" data-ctx="${ctx}" data-sec="${sec}" data-field="subm" data-sub="${s}" data-v="${v}" onchange="pfChange(this)"></td>`;
    });
    rows+=`</tr>`;
  });

  // PM Recurso / Requisito
  rows+=`<tr class="row-section"><td colspan="${2+VX.length}">RECURSO E REQUISITO — totais consolidados</td></tr>`;
  const fields=[
    {key:'recurso',      label:'RECURSO',            unit:'MWm',     cls:''},
    {key:'pm_recurso',   label:'PREÇO MÉDIO RECURSO', unit:'R$/MWh',  cls:'row-pmedio'},
    {key:'requisito',    label:'REQUISITO',           unit:'MWm',     cls:''},
    {key:'pm_requisito', label:'PREÇO MÉDIO REQUISITO',unit:'R$/MWh', cls:'row-pmedio'},
  ];
  fields.forEach(f=>{
    rows+=`<tr class="${f.cls}"><td>${f.label}</td><td>${f.unit}</td>`;
    VX.forEach(v=>{
      const val=(secData?.[f.key]?.[v])||0;
      rows+=`<td><input type="number" step="${f.unit==='R$/MWh'?'0.01':'0.1'}" value="${val}" data-ctx="${ctx}" data-sec="${sec}" data-field="${f.key}" data-v="${v}" onchange="pfChange(this)"></td>`;
    });
    rows+=`</tr>`;
  });

  // NET (computed)
  rows+=`<tr class="row-total" id="${ctx}-${sec}-net"><td>NET ENERGÉTICO</td><td>MWm</td>`+VX.map(v=>`<td id="${ctx}-${sec}-net-${v}">0,00</td>`).join('')+`</tr>`;
  // CHECK
  rows+=`<tr class="row-check" id="${ctx}-${sec}-chk"><td>CHECK</td><td></td>`+VX.map(v=>`<td id="${ctx}-${sec}-chk-${v}">—</td>`).join('')+`</tr>`;

  el.innerHTML=hdr+rows+`</tbody></table>`;
  updateNET(ctx, sec, secData);
}

function renderDerSection(ctx, derData, containerId, onChange){
  const el=document.getElementById(containerId); if(!el) return;
  const hdr=`<table><thead><tr>
    <th style="min-width:220px;">Exposições</th>
    <th style="width:80px;">Unid.</th>
    ${VX.map(v=>`<th class="col-v">${v}</th>`).join('')}
  </tr></thead><tbody>`;

  let rows='';
  rows+=`<tr class="row-section"><td colspan="${2+VX.length}">EXPOSIÇÃO LÍQUIDA POR SUBMERCADO (MWm)</td></tr>`;
  SUBMS.forEach(s=>{
    rows+=`<tr><td>${SUBM_LABELS[s]}</td><td>MWm</td>`;
    VX.forEach(v=>{
      const val=(derData?.subm?.[s]?.[v])||0;
      rows+=`<td><input type="number" step="0.1" value="${val}" data-ctx="${ctx}" data-sec="der" data-field="subm" data-sub="${s}" data-v="${v}" onchange="pfChange(this)"></td>`;
    });
    rows+=`</tr>`;
  });
  rows+=`<tr class="row-section"><td colspan="${2+VX.length}">RECURSO E REQUISITO</td></tr>`;
  [{key:'recurso',label:'RECURSO',unit:'MWm'},{key:'pm_recurso',label:'PREÇO MÉDIO RECURSO',unit:'R$/MWh',cls:'row-pmedio'},
   {key:'requisito',label:'REQUISITO',unit:'MWm'},{key:'pm_requisito',label:'PREÇO MÉDIO REQUISITO',unit:'R$/MWh',cls:'row-pmedio'}
  ].forEach(f=>{
    rows+=`<tr class="${f.cls||''}"><td>${f.label}</td><td>${f.unit}</td>`;
    VX.forEach(v=>{
      const val=(derData?.[f.key]?.[v])||0;
      rows+=`<td><input type="number" step="${f.unit==='R$/MWh'?'0.01':'0.1'}" value="${val}" data-ctx="${ctx}" data-sec="der" data-field="${f.key}" data-v="${v}" onchange="pfChange(this)"></td>`;
    });
    rows+=`</tr>`;
  });
  el.innerHTML=hdr+rows+`</tbody></table>`;
}

function renderEFM(ctx, efmData, containerId, onChange){
  const el=document.getElementById(containerId); if(!el) return;
  const fn = onChange.name||'triggerCalcEmp';
  let h=`<table><thead><tr><th style="min-width:220px;">Efeitos Financeiros</th><th>Unid.</th>${VX.map(v=>`<th class="col-v">${v}</th>`).join('')}</tr></thead><tbody>`;
  h+=`<tr><td>Efeitos Financeiros do Mercado Regulado (ACR, etc.)</td><td>R$</td>`;
  VX.forEach(v=>{
    h+=`<td><input type="number" step="1000" value="${(efmData?.[v])||0}" data-ctx="${ctx}" data-field="efm" data-v="${v}" onchange="efmChange(this)"></td>`;
  });
  h+=`</tr></tbody></table>`;
  el.innerHTML=h;
}

// ============================================================
// INPUT CHANGE HANDLERS
// ============================================================
function pfChange(inp){
  const {ctx,sec,field,sub,v} = inp.dataset;
  const target = ctx==='extra' ? extraData : emp;
  const val = parseFloat(inp.value)||0;
  if(sec==='der'){
    if(field==='subm') target.derivativos.subm[sub][v]=val;
    else target.derivativos[field][v]=val;
  } else {
    const s = sec==='pf' ? 'preco_fixo' : 'preco_variavel';
    if(field==='subm') target[s].subm[sub][v]=val;
    else target[s][field][v]=val;
  }
  if(sec!=='der') updateNET(ctx, sec, sec==='pf'?target.preco_fixo:target.preco_variavel);
  if(ctx==='extra') triggerCalcExtra(); else triggerCalcEmp();
}

function efmChange(inp){
  const {ctx,v} = inp.dataset;
  const target = ctx==='extra' ? extraData : emp;
  target.efm_regulado[v] = parseFloat(inp.value)||0;
  if(ctx==='extra') triggerCalcExtra(); else triggerCalcEmp();
}

function updateNET(ctx, sec, secData){
  const netId = `${ctx}-${sec}-net`;
  if(!document.getElementById(netId)) return;
  VX.forEach(v=>{
    const net = (secData?.recurso?.[v]||0) - (secData?.requisito?.[v]||0);
    const netEl = document.getElementById(`${ctx}-${sec}-net-${v}`);
    if(netEl) netEl.textContent = fmtN(net,2);

    // CHECK: sum subm == net
    const sumSubm = SUBMS.reduce((s,sm)=>s+(secData?.subm?.[sm]?.[v]||0),0);
    const diff = Math.abs(sumSubm - net);
    const chkEl = document.getElementById(`${ctx}-${sec}-chk-${v}`);
    if(chkEl){
      if(net===0 && sumSubm===0){ chkEl.textContent='OK'; chkEl.style.color='var(--green)'; }
      else if(diff<0.01){ chkEl.textContent='OK'; chkEl.style.color='var(--green)'; }
      else { chkEl.textContent=`Δ${fmtN(diff,2)}`; chkEl.style.color='var(--red)'; }
    }
  });
}

// ============================================================
// CALCULATE EMPRESA
// ============================================================
let calcTimer=null;
function triggerCalcEmp(){ clearTimeout(calcTimer); calcTimer=setTimeout(calcularEmp,300); }

async function calcularEmp(){
  lastEmp = await fetch('/api/calcular').then(r=>r.json());
  updateHeader(lastEmp);
  renderResultados(lastEmp, 'emp');
  if(document.getElementById('tab-calculo').classList.contains('active')) renderCalculo();
}

async function salvarEmpresa(){
  emp.pla = { pl_bruto:inp('pla-bruto'), deducoes: DED_ITENS.map((_,i)=>({item:DED_ITENS[i],valor:inp(`ded-val-${i}`),descricao:document.getElementById(`ded-desc-${i}`).value})) };
  lastEmp = await fetch('/api/empresa',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(emp)}).then(r=>r.json());
  showMsg('msg-emp','Portfólio salvo e recalculado.');
  updateHeader(lastEmp);
  renderResultados(lastEmp,'emp');
  renderCalculo();
}

async function salvarHistorico(){
  const res=await fetch('/api/historico/salvar',{method:'POST'}).then(r=>r.json());
  showMsg('msg-emp',`Salvo no histórico: ${res.entry?.data||'?'} — FA_RIS=${fmtN(res.entry?.fa_ris,4)}`);
}

// ============================================================
// PORTFÓLIO EXTRA
// ============================================================
function emptyExtraSection(){
  const z=()=>Object.fromEntries(VX.map(v=>[v,0]));
  return {
    subm: Object.fromEntries(SUBMS.map(s=>[s,z()])),
    recurso:z(), pm_recurso:z(), requisito:z(), pm_requisito:z()
  };
}

let calcExtraTimer=null;
function triggerCalcExtra(){ clearTimeout(calcExtraTimer); calcExtraTimer=setTimeout(calcExtra,300); }

async function calcExtra(){
  if(!extraData) return;
  lastExtraPreview = await fetch('/api/calcular/extra-preview',{
    method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify(extraData)
  }).then(r=>r.json());
  renderExtraCompare();
}

function renderExtraCompare(){
  if(!lastExtraPreview) return;
  const r=lastExtraPreview.real.totais, c=lastExtraPreview.combinado.totais;
  const faStr=(t)=>t.pla_negativo?'PLA≤0':(t.fa_ris!=null?fmtN(t.fa_ris,4):'—');
  const faDivStr=(t)=>t.pla_negativo?'PLA≤0':(t.fa_divulgado!=null?fmtN(t.fa_divulgado,4):'—');
  const delta=(a,b)=>{ if(a==null||b==null) return '—'; const d=b-a; const cor=d>0?'var(--red)':d<0?'var(--green)':'var(--muted)'; return `<span style="color:${cor}">${d>0?'+':''}${fmtN(d,4)}</span>`; };
  const deltaR=(a,b)=>{ const d=(b||0)-(a||0); const cor=d<0?'var(--red)':d>0?'var(--green)':'var(--muted)'; return `<span style="color:${cor}">${d>0?'+':''}${fmtN(d/1e6,2)} mi</span>`; };
  let h=`<table><thead><tr><th>Indicador</th><th>Portfólio Real</th><th>Real + Extra</th><th>Variação</th></tr></thead><tbody>`;
  h+=`<tr class="row-total"><td>FA Risco</td><td>${faStr(r)}</td><td>${faStr(c)}</td><td>${delta(r.fa_ris,c.fa_ris)}</td></tr>`;
  h+=`<tr class="row-total"><td>FA Divulgado</td><td>${faDivStr(r)}</td><td>${faDivStr(c)}</td><td>${delta(r.fa_divulgado,c.fa_divulgado)}</td></tr>`;
  h+=`<tr><td>RWA (R$)</td><td>${fmtN(r.rwa,0)}</td><td>${fmtN(c.rwa,0)}</td><td>${deltaR(r.rwa,c.rwa)}</td></tr>`;
  h+=`<tr><td>Resultado Financeiro (R$)</td><td>${fmtN(r.res_fin,0)}</td><td>${fmtN(c.res_fin,0)}</td><td>${deltaR(r.res_fin,c.res_fin)}</td></tr>`;
  h+=`<tr><td>VaR Total (R$)</td><td>${fmtN(r.var_tot,0)}</td><td>${fmtN(c.var_tot,0)}</td><td>${deltaR(r.var_tot,c.var_tot)}</td></tr>`;
  h+=`<tr><td>MtM (R$)</td><td>${fmtN(r.mtm,0)}</td><td>${fmtN(c.mtm,0)}</td><td>${deltaR(r.mtm,c.mtm)}</td></tr>`;
  h+=`<tr><td>Stress Total (R$)</td><td>${fmtN(r.stest_tot,0)}</td><td>${fmtN(c.stest_tot,0)}</td><td>${deltaR(r.stest_tot,c.stest_tot)}</td></tr>`;
  h+=`</tbody></table>`;
  document.getElementById('extra-compare').innerHTML=h;
}

function syncExtraToggle(){
  const on = !!(extraData && extraData.ativo);
  const cb=document.getElementById('extra-ativo'); if(cb) cb.checked=on;
  const lbl=document.getElementById('extra-toggle-label');
  const sub=document.getElementById('extra-toggle-sub');
  const badge=document.getElementById('extra-status-badge');
  if(lbl) lbl.textContent = on
    ? 'Portfólio Extra ATIVADO — FA usa Real + Extra'
    : 'Portfólio Extra DESATIVADO — FA usa só o portfólio real';
  if(sub) sub.textContent = on
    ? 'O cálculo do Fator de Alavancagem está somando este portfólio ao real.'
    : 'Ative para somar este portfólio ao real no cálculo do Fator de Alavancagem.';
  if(badge){ badge.textContent = on?'Ativo no FA':'Inativo'; badge.className = 'hbadge '+(on?'hb-green':'hb-gray'); }
}

async function toggleExtra(){
  if(!extraData) return;
  extraData.ativo = document.getElementById('extra-ativo').checked;
  syncExtraToggle();
  await fetch('/api/portfolio_extra',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(extraData)});
  await calcularEmp();   // header + Cálculo passam a refletir (ou não) o extra
  showMsg('msg-extra', extraData.ativo
    ? '✔ Portfólio Extra ATIVADO no cálculo do FA.'
    : 'Portfólio Extra desativado — FA voltou a usar só o real.', 'ok');
}

async function salvarExtra(){
  await fetch('/api/portfolio_extra',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(extraData)});
  showMsg('msg-extra','Portfólio Extra salvo.');
  if(extraData.ativo) await calcularEmp();
}

function limparExtra(){
  extraData.preco_fixo=emptyExtraSection();
  extraData.preco_variavel=emptyExtraSection();
  extraData.derivativos=emptyExtraSection();
  extraData.efm_regulado=Object.fromEntries(VX.map(v=>[v,0]));
  renderPortfolioTables('extra',extraData,triggerCalcExtra,'-x');
  calcExtra();
  showMsg('msg-extra','Campos do Portfólio Extra zerados (lembre de salvar).','info');
}

async function uploadExtraCSV(input){
  const file=input.files[0]; if(!file) return;
  showMsg('msg-extra','Carregando CSV...','info');
  try{
    const buf=await file.arrayBuffer();
    const res=await fetch('/api/upload/portfolio_extra_csv',{
      method:'POST', headers:{'Content-Type':'application/octet-stream'}, body:buf
    }).then(r=>r.json());
    if(res.ok){
      const wasAtivo=extraData.ativo;
      extraData=Object.assign(extraData,res.data);
      extraData.ativo=wasAtivo;   // upload não altera o estado do toggle
      renderPortfolioTables('extra',extraData,triggerCalcExtra,'-x');
      calcExtra();
      showMsg('msg-extra','✔ CSV importado com sucesso.');
    } else { showMsg('msg-extra','Erro: '+(res.errors||[]).join('; '),'err'); }
  }catch(e){ showMsg('msg-extra','Falha na requisição: '+e.message,'err'); }
  input.value='';
}

function downloadModeloExtra(){ window.open('/api/download/modelo_extra_csv','_blank'); }


// ============================================================
// UPLOAD PORTFOLIO
// ============================================================
async function uploadPortfolio(input){
  const file=input.files[0]; if(!file) return;
  showMsg('msg-port','Carregando...','info');
  try {
    const buf = await file.arrayBuffer();
    const res = await fetch('/api/upload/portfolio',{
      method:'POST',
      headers:{'Content-Type':'application/octet-stream'},
      body: buf
    }).then(r=>r.json());
    if(res.ok){
      emp=Object.assign(emp, res.data);
      loadEmpresaForm();
      renderPortfolioTables('emp',emp,triggerCalcEmp);
      showMsg('msg-port','✔ Portfólio importado com sucesso.');
      await calcularEmp();
    } else { showMsg('msg-port','Erro: '+(res.errors||[]).join('; '),'err'); }
  } catch(e) {
    showMsg('msg-port','Falha na requisição: '+e.message,'err');
  }
  input.value='';
}


function downloadModelo(){ window.open('/api/download/modelo','_blank'); }

// ============================================================
// RENDER RESULTADOS
// ============================================================
function updateHeader(res){
  const t=res.totais;
  const ri=t.fa_ris, di=t.fa_divulgado, neg=t.pla_negativo;
  const rs=neg?'PLA≤0':(ri!=null?fmtN(ri,4):'—');
  const ds=neg?'PLA≤0':(di!=null?fmtN(di,4):'—');
  document.getElementById('hb-ris').textContent='FA Risco '+rs;
  document.getElementById('hb-ris').className='hbadge '+(neg?'hb-gray':hbClass(ri));
  document.getElementById('hb-div').textContent='FA Divulgado '+ds;
  document.getElementById('hb-div').className='hbadge '+(neg?'hb-gray':hbClass(di));
  document.getElementById('hb-rwa').textContent='RWA '+fmtM(t.rwa);
  document.getElementById('hb-pla').textContent='PLA '+fmtM(t.pla);
  // Indicador de portfólio extra ativo
  let exEl=document.getElementById('hb-extra');
  if(res.extra_ativo){
    if(!exEl){ exEl=document.createElement('span'); exEl.id='hb-extra'; document.getElementById('hdr-badges').prepend(exEl); }
    exEl.className='hbadge hb-yellow'; exEl.textContent='+ EXTRA';
    exEl.title='Portfólio Extra incluído no cálculo do FA';
  } else if(exEl){ exEl.remove(); }
}

function renderResultados(res, ctx){
  const t=res.totais;
  if(ctx==='emp'){
    const neg=t.pla_negativo;
    const ri=t.fa_ris, di=t.fa_divulgado;
    const rs=neg?'PLA ≤ 0':(ri!=null?fmtN(ri,4):'—');
    const ds=neg?'PLA ≤ 0':(di!=null?fmtN(di,4):'—');
    setBadge('fa-ris-badge',rs,neg?'fa-gray':faClass(ri),'FA Risco');
    setBadge('fa-div-badge',ds,neg?'fa-gray':faClass(di),'FA Divulgado');
    document.getElementById('r-rwa').textContent=fmtM(t.rwa);
    document.getElementById('r-resfin').textContent=fmtM(t.res_fin);
    document.getElementById('r-pla').textContent=fmtM(t.pla);
    document.getElementById('r-var').textContent=fmtM(t.var_tot);
    document.getElementById('r-mtm').textContent=fmtM(t.mtm);
    document.getElementById('r-rescontr').textContent=fmtM(t.res_contr);
    document.getElementById('r-finpv').textContent=fmtM(t.fin_pv);
    document.getElementById('r-efm').textContent=fmtM(t.efm);
  }
}

function setBadge(id, val, cls, sub){
  const el=document.getElementById(id); if(!el) return;
  el.className='fa-big '+cls;
  el.innerHTML=`<div class="num">${val}</div><div class="sub">${sub}</div>`;
}

// ============================================================
// CALCULO TAB
// ============================================================
function renderCalculo(){
  if(!lastEmp) return;
  const pv = lastEmp.por_vertice;
  const t = lastEmp.totais;
  const p = premissas;

  // VaR table
  let h=`<table><thead><tr>
    <th>Mês</th><th>Horas</th><th>NET Total (MWm)</th>
    <th>Forward SE/CO (R$/MWh)</th><th>Exposição (R$)</th>
    <th>Percentil</th><th>φ (Dist. Normal)</th>
    <th>Volatilidade (%)</th><th>Dias Liq.</th>
    <th>VaR Paramétrico (R$)</th>
  </tr></thead><tbody>`;
  VX.forEach(v=>{
    const d=pv[v]||{};
    const hr=(p.horas||{})[v]||0;
    const fwd=(p.forward||{})[v]?.SECO||0;
    const exp=d.net_total*fwd*hr;
    const sig=((p.volatilidades||{})[v]||0)*100;
    h+=`<tr>
      <td style="text-align:left;">${v}</td>
      <td>${hr}</td>
      <td>${fmtN(d.net_total,3)}</td>
      <td>${fmtN(fwd)}</td>
      <td>${fmtN(exp,0)}</td>
      <td>${fmtN((1+Math.abs(p.phi_norm||1.6449))/2*100,2)}%</td>
      <td>${fmtN(Math.abs(p.phi_norm||1.6449),4)}</td>
      <td>${fmtN(sig,3)}%</td>
      <td>${p.dias_liquidacao||1}</td>
      <td style="font-weight:700;">${fmtN(d.var,0)}</td>
    </tr>`;
  });
  h+=`<tr class="row-total"><td colspan="9" style="text-align:right;">VaR Total (ρ=1)</td><td>${fmtN(t.var_tot,0)}</td></tr>`;
  h+=`</tbody></table>`;
  document.getElementById('tbl-var').innerHTML=h;

  // Consolidado
  let h2=`<table><thead><tr>
    <th style="min-width:220px;">Item</th><th>Unid.</th>
    ${VX.map(v=>`<th>${v}</th>`).join('')}<th>Total</th>
  </tr></thead><tbody>`;

  const grp=(label,unit,rows,total,bold)=>{
    let s=`<tr class="row-section"><td colspan="${3+VX.length}">${label}</td></tr>`;
    rows.forEach(r=>{ s+=`<tr${bold?' class="row-total"':''}><td>${r.label}</td><td>${unit}</td>`+VX.map(v=>`<td>${fmtN(r.vals[v],r.dec)}</td>`).join('')+`<td>${total!=null?fmtN(total,r.dec):'—'}</td></tr>`; });
    return s;
  };

  // Exposições por submercado
  h2+=`<tr class="row-section"><td colspan="${3+VX.length}">EXPOSIÇÕES LÍQUIDAS — MWm</td></tr>`;
  [{s:'net_seco',l:'SE/CO'},{s:'net_sul',l:'SUL'},{s:'net_ne',l:'NORDESTE'},{s:'net_n',l:'NORTE'}].forEach(r=>{
    h2+=`<tr><td>${r.l}</td><td>MWm</td>`+VX.map(v=>`<td>${fmtN(pv[v][r.s],3)}</td>`).join('')+`<td>—</td></tr>`;
  });
  h2+=`<tr class="row-total"><td>NET ENERGÉTICO TOTAL</td><td>MWm</td>`+VX.map(v=>`<td>${fmtN(pv[v].net_total,3)}</td>`).join('')+`<td>—</td></tr>`;

  // Acréscimos
  h2+=`<tr class="row-section"><td colspan="${3+VX.length}">ACRÉSCIMOS FINANCEIROS (R$)</td></tr>`;
  [
    {l:'Resultado Contratual (PF+DER)',k:'res_contr',dec:0},
    {l:'Financeiro PLD+',k:'fin_pv',dec:0},
    {l:'EFM Regulado',k:'efm',dec:0},
    {l:'MtM',k:'mtm',dec:0},
  ].forEach(r=>{
    h2+=`<tr><td>${r.l}</td><td>R$</td>`+VX.map(v=>`<td>${fmtN(pv[v][r.k],r.dec)}</td>`).join('')+`<td>—</td></tr>`;
  });
  h2+=`<tr class="row-total"><td>Resultado Financeiro Total (Fin)</td><td>R$</td>`+VX.map(()=>`<td>—</td>`).join('')+`<td style="font-weight:900;">${fmtN(t.res_fin,0)}</td></tr>`;

  // VaR/Stress/RWA
  h2+=`<tr class="row-section"><td colspan="${3+VX.length}">RISCO</td></tr>`;
  [{l:'VaR Paramétrico',k:'var',dec:0},{l:'Teste de Estresse',k:'stest',dec:0}].forEach(r=>{
    h2+=`<tr><td>${r.l}</td><td>R$</td>`+VX.map(v=>`<td>${fmtN(pv[v][r.k],r.dec)}</td>`).join('')+`<td>—</td></tr>`;
  });
  h2+=`<tr class="row-total"><td>RWA</td><td>R$</td>`+VX.map(()=>`<td>—</td>`).join('')+`<td>${fmtN(t.rwa,0)}</td></tr>`;

  // FA
  h2+=`<tr class="row-section"><td colspan="${3+VX.length}">FATOR DE ALAVANCAGEM</td></tr>`;
  h2+=`<tr class="row-total"><td>PLA</td><td>R$</td>`+VX.map(()=>`<td>—</td>`).join('')+`<td>${fmtN(t.pla,0)}</td></tr>`;
  const risStr=t.pla_negativo?'PLA ≤ 0':(t.fa_ris!=null?fmtN(t.fa_ris,4):'—');
  const divStr=t.pla_negativo?'PLA ≤ 0':(t.fa_divulgado!=null?fmtN(t.fa_divulgado,4):'—');
  h2+=`<tr class="row-total"><td style="font-size:14px;">FA Risco (RWA/PLA)</td><td></td>`+VX.map(()=>`<td>—</td>`).join('')+`<td style="font-size:16px;font-weight:900;color:${faColorCSS(t.fa_ris,t.pla_negativo)}">${risStr}</td></tr>`;
  h2+=`<tr class="row-total"><td style="font-size:14px;">FA Divulgado</td><td></td>`+VX.map(()=>`<td>—</td>`).join('')+`<td style="font-size:16px;font-weight:900;color:${faColorCSS(t.fa_divulgado,t.pla_negativo)}">${divStr}</td></tr>`;
  h2+=`</tbody></table>`;
  document.getElementById('tbl-consolidado').innerHTML=h2;

  // Charts
  mkChart('ch-exp','bar',VX,
    SUBMS.map((s,i)=>({label:s,data:VX.map(v=>pv[v][['net_seco','net_sul','net_ne','net_n'][i]]||0),backgroundColor:['#1a6a63','#4f9a8f','#b16a3a','#d0a05f'][i]})));
  mkChart('ch-mtm','bar',VX,[{label:'MtM R$',data:VX.map(v=>pv[v].mtm||0),backgroundColor:'rgba(177,106,58,.75)'}]);
  mkChart('ch-var','bar',VX,[{label:'VaR R$',data:VX.map(v=>pv[v].var||0),backgroundColor:'rgba(26,106,99,.75)'}]);
  mkChart('ch-stest','bar',VX,[{label:'Stress R$',data:VX.map(v=>pv[v].stest||0),backgroundColor:'rgba(178,58,49,.7)'}]);
}

function faColorCSS(v, neg){ return neg?'var(--muted)':v==null?'var(--muted)':v<1?'var(--green)':v<=3?'var(--yellow)':'var(--red)'; }


// ============================================================
// HISTORICO
// ============================================================
async function loadHistorico(){
  const hist=await fetch('/api/historico').then(r=>r.json())||[];
  if(!hist.length){ document.getElementById('tbl-hist').innerHTML='<p style="color:var(--muted);padding:10px;">Nenhum registro ainda. Use "Registrar no Histórico" na aba Portfólio.</p>'; return; }
  let h=`<table><thead><tr><th>Data</th><th>Semana</th><th>FA Risco</th><th>FA Divulgado</th><th>RWA (R$)</th><th>PnL (R$)</th><th>Res.Fin. (R$)</th><th>PLA (R$)</th><th>VaR (R$)</th><th>Stress (R$)</th></tr></thead><tbody>`;
  [...hist].reverse().forEach(h2=>{
    h+=`<tr><td style="text-align:left;">${h2.data}</td><td>${h2.semana}</td>
    <td>${h2.fa_ris!=null?fmtN(h2.fa_ris,4):'—'}</td>
    <td>${h2.fa_divulgado!=null?fmtN(h2.fa_divulgado,4):'—'}</td>
    <td>${fmtN(h2.rwa,0)}</td><td>${fmtN(h2.pnl,0)}</td><td>${fmtN(h2.res_fin,0)}</td>
    <td>${fmtN(h2.pla,0)}</td><td>${fmtN(h2.var_tot,0)}</td><td>${fmtN(h2.stest_tot,0)}</td></tr>`;
  });
  document.getElementById('tbl-hist').innerHTML=h+`</tbody></table>`;
  const labels=hist.map(h=>h.data);
  mkChart('ch-hfa','line',labels,[
    {label:'FA Risco',data:hist.map(h=>h.fa_ris),borderColor:'#1a6a63',backgroundColor:'#1a6a63',tension:.3,fill:false,pointRadius:4},
    {label:'FA Divulgado',data:hist.map(h=>h.fa_divulgado),borderColor:'#b16a3a',backgroundColor:'#b16a3a',tension:.3,fill:false,pointRadius:4},
  ]);
  mkChart('ch-hrwa','line',labels,[
    {label:'RWA',data:hist.map(h=>h.rwa),borderColor:'#b23a31',backgroundColor:'#b23a31',tension:.3,fill:false,pointRadius:4},
    {label:'PLA',data:hist.map(h=>h.pla),borderColor:'#2f7d54',backgroundColor:'#2f7d54',tension:.3,fill:false,pointRadius:4},
  ]);
}
function exportCSV(){ window.open('/api/download/historico_csv','_blank'); }

// ============================================================
// CHARTS
// ============================================================
function mkChart(id,type,labels,datasets){
  const canvas=document.getElementById(id); if(!canvas) return;
  if(charts[id]) charts[id].destroy();
  charts[id]=new Chart(canvas,{type,data:{labels,datasets},options:{
    responsive:true,maintainAspectRatio:false,
    plugins:{legend:{labels:{color:'#222b2a',font:{size:11}}}},
    scales:{x:{ticks:{color:'#6c756f'},grid:{color:'rgba(34,43,42,.07)'}},
            y:{ticks:{color:'#6c756f'},grid:{color:'rgba(34,43,42,.07)'}}}
  }});
}

init();
</script>
</body>
</html>
'''

# ===== DADOS INICIAIS =====
EMBEDDED_DATA = json.loads(r'''{
  "premissas": {
    "data_referencia": "2026-05-25",
    "forward": {
      "M+0": {
        "SECO": 222.72,
        "SUL": 23.8,
        "NE": -61.59,
        "N": -60.5,
        "I0": 1.69,
        "I5": 30.17,
        "I8": 172.09,
        "I1": 172.09
      },
      "M+1": {
        "SECO": 227.35,
        "SUL": 5.28,
        "NE": -36.7,
        "N": -32.64,
        "I0": 1.83,
        "I5": 28.61,
        "I8": 170.29,
        "I1": 170.29
      },
      "M+2": {
        "SECO": 229.93,
        "SUL": 2.72,
        "NE": -30.27,
        "N": -15.07,
        "I0": 1.83,
        "I5": 29.4,
        "I8": 171.12,
        "I1": 171.12
      },
      "M+3": {
        "SECO": 263.05,
        "SUL": 2.16,
        "NE": -28.37,
        "N": -11.42,
        "I0": 1.65,
        "I5": 29.28,
        "I8": 169.97,
        "I1": 169.97
      },
      "M+4": {
        "SECO": 292.63,
        "SUL": 2.58,
        "NE": -23.71,
        "N": -7.36,
        "I0": 1.61,
        "I5": 27.84,
        "I8": 167.61,
        "I1": 167.61
      },
      "M+5": {
        "SECO": 313.43,
        "SUL": 1.81,
        "NE": -18.07,
        "N": -7.05,
        "I0": 1.61,
        "I5": 28.98,
        "I8": 169.53,
        "I1": 169.53
      },
      "M+6": {
        "SECO": 313.43,
        "SUL": 1.81,
        "NE": -18.07,
        "N": -7.05,
        "I0": 1.61,
        "I5": 28.98,
        "I8": 169.53,
        "I1": 169.53
      }
    },
    "volatilidades": {
      "M+0": 0.07339999999999999,
      "M+1": 0.11894,
      "M+2": 0.10205,
      "M+3": 0.08728,
      "M+4": 0.067,
      "M+5": 0.05623,
      "M+6": 0.048600000000000004
    },
    "stress_long": {
      "M+0": 168.47,
      "M+1": 134.86,
      "M+2": 139.94,
      "M+3": 173.77,
      "M+4": 219.12,
      "M+5": 231.91,
      "M+6": 232.69
    },
    "stress_short": {
      "M+0": 279.76,
      "M+1": 326.52,
      "M+2": 333.03,
      "M+3": 365.56,
      "M+4": 372.34,
      "M+5": 395.11,
      "M+6": 389.62
    },
    "horas": {
      "M+0": 744,
      "M+1": 720,
      "M+2": 744,
      "M+3": 744,
      "M+4": 720,
      "M+5": 744,
      "M+6": 720
    },
    "phi_norm": -1.6449,
    "dias_liquidacao": 1,
    "K": 0,
    "theta": 0,
    "rho": 1,
    "pld_min": 57.31,
    "pld_max": 785.27
  },
  "empresa": {
    "nome": "Minha Empresa",
    "pla": {
      "pl_bruto": 0,
      "deducoes": [
        {
          "item": "I",
          "valor": 0,
          "descricao": ""
        },
        {
          "item": "II",
          "valor": 0,
          "descricao": ""
        },
        {
          "item": "III",
          "valor": 0,
          "descricao": ""
        },
        {
          "item": "IV",
          "valor": 0,
          "descricao": ""
        },
        {
          "item": "V",
          "valor": 0,
          "descricao": ""
        },
        {
          "item": "VI",
          "valor": 0,
          "descricao": ""
        },
        {
          "item": "VII",
          "valor": 0,
          "descricao": ""
        },
        {
          "item": "VIII",
          "valor": 0,
          "descricao": ""
        }
      ]
    },
    "preco_fixo": {
      "subm": {
        "SE/CO": {
          "M+0": 0,
          "M+1": 0,
          "M+2": 0,
          "M+3": 0,
          "M+4": 0,
          "M+5": 0,
          "M+6": 0
        },
        "SUL": {
          "M+0": 0,
          "M+1": 0,
          "M+2": 0,
          "M+3": 0,
          "M+4": 0,
          "M+5": 0,
          "M+6": 0
        },
        "NE": {
          "M+0": 0,
          "M+1": 0,
          "M+2": 0,
          "M+3": 0,
          "M+4": 0,
          "M+5": 0,
          "M+6": 0
        },
        "N": {
          "M+0": 0,
          "M+1": 0,
          "M+2": 0,
          "M+3": 0,
          "M+4": 0,
          "M+5": 0,
          "M+6": 0
        }
      },
      "recurso": {
        "M+0": 0,
        "M+1": 0,
        "M+2": 0,
        "M+3": 0,
        "M+4": 0,
        "M+5": 0,
        "M+6": 0
      },
      "pm_recurso": {
        "M+0": 0,
        "M+1": 0,
        "M+2": 0,
        "M+3": 0,
        "M+4": 0,
        "M+5": 0,
        "M+6": 0
      },
      "requisito": {
        "M+0": 0,
        "M+1": 0,
        "M+2": 0,
        "M+3": 0,
        "M+4": 0,
        "M+5": 0,
        "M+6": 0
      },
      "pm_requisito": {
        "M+0": 0,
        "M+1": 0,
        "M+2": 0,
        "M+3": 0,
        "M+4": 0,
        "M+5": 0,
        "M+6": 0
      }
    },
    "preco_variavel": {
      "subm": {
        "SE/CO": {
          "M+0": 0,
          "M+1": 0,
          "M+2": 0,
          "M+3": 0,
          "M+4": 0,
          "M+5": 0,
          "M+6": 0
        },
        "SUL": {
          "M+0": 0,
          "M+1": 0,
          "M+2": 0,
          "M+3": 0,
          "M+4": 0,
          "M+5": 0,
          "M+6": 0
        },
        "NE": {
          "M+0": 0,
          "M+1": 0,
          "M+2": 0,
          "M+3": 0,
          "M+4": 0,
          "M+5": 0,
          "M+6": 0
        },
        "N": {
          "M+0": 0,
          "M+1": 0,
          "M+2": 0,
          "M+3": 0,
          "M+4": 0,
          "M+5": 0,
          "M+6": 0
        }
      },
      "recurso": {
        "M+0": 0,
        "M+1": 0,
        "M+2": 0,
        "M+3": 0,
        "M+4": 0,
        "M+5": 0,
        "M+6": 0
      },
      "pm_recurso": {
        "M+0": 0,
        "M+1": 0,
        "M+2": 0,
        "M+3": 0,
        "M+4": 0,
        "M+5": 0,
        "M+6": 0
      },
      "requisito": {
        "M+0": 0,
        "M+1": 0,
        "M+2": 0,
        "M+3": 0,
        "M+4": 0,
        "M+5": 0,
        "M+6": 0
      },
      "pm_requisito": {
        "M+0": 0,
        "M+1": 0,
        "M+2": 0,
        "M+3": 0,
        "M+4": 0,
        "M+5": 0,
        "M+6": 0
      }
    },
    "derivativos": {
      "subm": {
        "SE/CO": {
          "M+0": 0,
          "M+1": 0,
          "M+2": 0,
          "M+3": 0,
          "M+4": 0,
          "M+5": 0,
          "M+6": 0
        },
        "SUL": {
          "M+0": 0,
          "M+1": 0,
          "M+2": 0,
          "M+3": 0,
          "M+4": 0,
          "M+5": 0,
          "M+6": 0
        },
        "NE": {
          "M+0": 0,
          "M+1": 0,
          "M+2": 0,
          "M+3": 0,
          "M+4": 0,
          "M+5": 0,
          "M+6": 0
        },
        "N": {
          "M+0": 0,
          "M+1": 0,
          "M+2": 0,
          "M+3": 0,
          "M+4": 0,
          "M+5": 0,
          "M+6": 0
        }
      },
      "recurso": {
        "M+0": 0,
        "M+1": 0,
        "M+2": 0,
        "M+3": 0,
        "M+4": 0,
        "M+5": 0,
        "M+6": 0
      },
      "pm_recurso": {
        "M+0": 0,
        "M+1": 0,
        "M+2": 0,
        "M+3": 0,
        "M+4": 0,
        "M+5": 0,
        "M+6": 0
      },
      "requisito": {
        "M+0": 0,
        "M+1": 0,
        "M+2": 0,
        "M+3": 0,
        "M+4": 0,
        "M+5": 0,
        "M+6": 0
      },
      "pm_requisito": {
        "M+0": 0,
        "M+1": 0,
        "M+2": 0,
        "M+3": 0,
        "M+4": 0,
        "M+5": 0,
        "M+6": 0
      }
    },
    "efm_regulado": {
      "M+0": 0,
      "M+1": 0,
      "M+2": 0,
      "M+3": 0,
      "M+4": 0,
      "M+5": 0,
      "M+6": 0
    }
  },
  "portfolio_extra": {
    "ativo": false,
    "nome": "Portfólio Extra (Simulação)",
    "preco_fixo": {
      "subm": {
        "SE/CO": {
          "M+0": 0,
          "M+1": 0,
          "M+2": 0,
          "M+3": 0,
          "M+4": 0,
          "M+5": 0,
          "M+6": 0
        },
        "SUL": {
          "M+0": 0,
          "M+1": 0,
          "M+2": 0,
          "M+3": 0,
          "M+4": 0,
          "M+5": 0,
          "M+6": 0
        },
        "NE": {
          "M+0": 0,
          "M+1": 0,
          "M+2": 0,
          "M+3": 0,
          "M+4": 0,
          "M+5": 0,
          "M+6": 0
        },
        "N": {
          "M+0": 0,
          "M+1": 0,
          "M+2": 0,
          "M+3": 0,
          "M+4": 0,
          "M+5": 0,
          "M+6": 0
        }
      },
      "recurso": {
        "M+0": 0,
        "M+1": 0,
        "M+2": 0,
        "M+3": 0,
        "M+4": 0,
        "M+5": 0,
        "M+6": 0
      },
      "pm_recurso": {
        "M+0": 0,
        "M+1": 0,
        "M+2": 0,
        "M+3": 0,
        "M+4": 0,
        "M+5": 0,
        "M+6": 0
      },
      "requisito": {
        "M+0": 0,
        "M+1": 0,
        "M+2": 0,
        "M+3": 0,
        "M+4": 0,
        "M+5": 0,
        "M+6": 0
      },
      "pm_requisito": {
        "M+0": 0,
        "M+1": 0,
        "M+2": 0,
        "M+3": 0,
        "M+4": 0,
        "M+5": 0,
        "M+6": 0
      }
    },
    "preco_variavel": {
      "subm": {
        "SE/CO": {
          "M+0": 0,
          "M+1": 0,
          "M+2": 0,
          "M+3": 0,
          "M+4": 0,
          "M+5": 0,
          "M+6": 0
        },
        "SUL": {
          "M+0": 0,
          "M+1": 0,
          "M+2": 0,
          "M+3": 0,
          "M+4": 0,
          "M+5": 0,
          "M+6": 0
        },
        "NE": {
          "M+0": 0,
          "M+1": 0,
          "M+2": 0,
          "M+3": 0,
          "M+4": 0,
          "M+5": 0,
          "M+6": 0
        },
        "N": {
          "M+0": 0,
          "M+1": 0,
          "M+2": 0,
          "M+3": 0,
          "M+4": 0,
          "M+5": 0,
          "M+6": 0
        }
      },
      "recurso": {
        "M+0": 0,
        "M+1": 0,
        "M+2": 0,
        "M+3": 0,
        "M+4": 0,
        "M+5": 0,
        "M+6": 0
      },
      "pm_recurso": {
        "M+0": 0,
        "M+1": 0,
        "M+2": 0,
        "M+3": 0,
        "M+4": 0,
        "M+5": 0,
        "M+6": 0
      },
      "requisito": {
        "M+0": 0,
        "M+1": 0,
        "M+2": 0,
        "M+3": 0,
        "M+4": 0,
        "M+5": 0,
        "M+6": 0
      },
      "pm_requisito": {
        "M+0": 0,
        "M+1": 0,
        "M+2": 0,
        "M+3": 0,
        "M+4": 0,
        "M+5": 0,
        "M+6": 0
      }
    },
    "derivativos": {
      "subm": {
        "SE/CO": {
          "M+0": 0,
          "M+1": 0,
          "M+2": 0,
          "M+3": 0,
          "M+4": 0,
          "M+5": 0,
          "M+6": 0
        },
        "SUL": {
          "M+0": 0,
          "M+1": 0,
          "M+2": 0,
          "M+3": 0,
          "M+4": 0,
          "M+5": 0,
          "M+6": 0
        },
        "NE": {
          "M+0": 0,
          "M+1": 0,
          "M+2": 0,
          "M+3": 0,
          "M+4": 0,
          "M+5": 0,
          "M+6": 0
        },
        "N": {
          "M+0": 0,
          "M+1": 0,
          "M+2": 0,
          "M+3": 0,
          "M+4": 0,
          "M+5": 0,
          "M+6": 0
        }
      },
      "recurso": {
        "M+0": 0,
        "M+1": 0,
        "M+2": 0,
        "M+3": 0,
        "M+4": 0,
        "M+5": 0,
        "M+6": 0
      },
      "pm_recurso": {
        "M+0": 0,
        "M+1": 0,
        "M+2": 0,
        "M+3": 0,
        "M+4": 0,
        "M+5": 0,
        "M+6": 0
      },
      "requisito": {
        "M+0": 0,
        "M+1": 0,
        "M+2": 0,
        "M+3": 0,
        "M+4": 0,
        "M+5": 0,
        "M+6": 0
      },
      "pm_requisito": {
        "M+0": 0,
        "M+1": 0,
        "M+2": 0,
        "M+3": 0,
        "M+4": 0,
        "M+5": 0,
        "M+6": 0
      }
    },
    "efm_regulado": {
      "M+0": 0,
      "M+1": 0,
      "M+2": 0,
      "M+3": 0,
      "M+4": 0,
      "M+5": 0,
      "M+6": 0
    }
  },
  "historico": [
    {
      "data": "2026-05-27",
      "semana": "22/2026",
      "fa_ris": 0.0,
      "fa_divulgado": 0.0,
      "rwa": 0.0,
      "pnl": 0.0,
      "res_fin": 0.0,
      "pla": 0.0,
      "var_tot": 0.0,
      "stest_tot": 0.0
    }
  ]
}''')

# ─────────────────────────────────────────────────────────────────────────────
#  SERVIDOR HTTP — backend Python. Serve a interface (identica ao HTML) e a API.
#  Estado mantido em memoria (DB), a partir dos dados embutidos.
# ─────────────────────────────────────────────────────────────────────────────
import copy as _copy
DB = _copy.deepcopy(EMBEDDED_DATA)


def _get(name):
    return DB.get(name)


def _set(name, value):
    DB[name] = value


class Handler(BaseHTTPRequestHandler):

    def log_message(self, fmt, *args):
        print(f"  [{self.address_string()}] {fmt % args}", flush=True)

    def _send_json(self, data, status=200):
        body = json.dumps(data, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Access-Control-Allow-Origin", "*")
        self.end_headers()
        self.wfile.write(body)

    def _send_bytes(self, data, content_type, filename):
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
        self.send_header("Content-Length", str(len(data)))
        self.send_header("Access-Control-Allow-Origin", "*")
        self.end_headers()
        self.wfile.write(data)

    def _send_html(self):
        body = EMBEDDED_HTML.encode("utf-8")
        self.send_response(200)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _read_body(self):
        length = self.headers.get("Content-Length")
        if not length:
            return b""
        return self.rfile.read(int(length))

    def do_OPTIONS(self):
        self.send_response(204)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET,POST,OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.end_headers()

    def do_GET(self):
        path = urlparse(self.path).path
        try:
            if path in ("/", "/index.html"):
                self._send_html()
            elif path == "/api/premissas":
                self._send_json(_get("premissas"))
            elif path == "/api/empresa":
                self._send_json(_get("empresa"))
            elif path == "/api/historico":
                self._send_json(_get("historico"))
            elif path == "/api/portfolio_extra":
                self._send_json(_get("portfolio_extra"))
            elif path == "/api/calcular":
                empresa = _get("empresa")
                premissas = _get("premissas")
                extra = _get("portfolio_extra") or {}
                extra_ativo = bool(extra.get("ativo"))
                base = combinar_portfolios(empresa, extra) if extra_ativo else empresa
                resultado = calcular_fa(base, premissas)
                resultado["extra_ativo"] = extra_ativo
                self._send_json(resultado)
            elif path == "/api/download/modelo":
                if not _HAS_OPENPYXL:
                    self._send_json({"error": "openpyxl indisponivel"}, 500); return
                self._send_bytes(
                    gerar_modelo_portfolio(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "modelo_portfolio.xlsx")
            elif path == "/api/download/modelo_extra_csv":
                self._send_bytes(gerar_modelo_extra_csv().encode("utf-8-sig"),
                                 "text/csv", "modelo_portfolio_extra.csv")
            elif path == "/api/download/historico_csv":
                hist = _get("historico") or []
                buf = io.StringIO()
                if hist:
                    w = csv.DictWriter(buf, fieldnames=list(hist[0].keys()))
                    w.writeheader(); w.writerows(hist)
                self._send_bytes(buf.getvalue().encode("utf-8-sig"), "text/csv", "historico_fa.csv")
            else:
                self.send_response(404); self.end_headers()
        except Exception:
            traceback.print_exc()
            self._send_json({"error": traceback.format_exc()}, 500)

    def do_POST(self):
        path = urlparse(self.path).path
        try:
            if path == "/api/premissas":
                data = json.loads(self._read_body())
                _set("premissas", data)
                self._send_json({"ok": True})

            elif path == "/api/empresa":
                data = json.loads(self._read_body())
                _set("empresa", data)
                self._send_json(calcular_fa(data, _get("premissas")))

            elif path == "/api/historico/salvar":
                empresa = _get("empresa"); premissas = _get("premissas")
                t = calcular_fa(empresa, premissas)["totais"]
                from datetime import date
                sem = date.today().isocalendar()
                entry = {
                    "data": premissas.get("data_referencia", str(date.today())),
                    "semana": f"{sem[1]}/{sem[0]}",
                    "fa_ris": t["fa_ris"], "fa_divulgado": t["fa_divulgado"],
                    "rwa": t["rwa"], "pnl": t["pnl"], "res_fin": t["res_fin"],
                    "pla": t["pla"], "var_tot": t["var_tot"], "stest_tot": t["stest_tot"],
                }
                hist = _get("historico") or []
                hist = [entry if h["data"] == entry["data"] else h for h in hist]
                if not any(h["data"] == entry["data"] for h in hist):
                    hist.append(entry)
                _set("historico", hist)
                self._send_json({"ok": True, "entry": entry})

            elif path == "/api/upload/premissas":
                if not _HAS_OPENPYXL:
                    self._send_json({"ok": False, "errors": ["openpyxl indisponivel no servidor."], "data": {}}); return
                ct = self.headers.get("Content-Type", "")
                file_bytes = _extract_file(self._read_body(), ct)
                if len(file_bytes) < 100:
                    self._send_json({"ok": False, "errors": [f"Arquivo nao recebido ({len(file_bytes)} bytes)."], "data": {}}); return
                parsed_data, errors = parse_planilha_ccee(file_bytes)
                if errors:
                    self._send_json({"ok": False, "errors": errors, "data": parsed_data})
                else:
                    premissas = _get("premissas"); premissas.update(parsed_data)
                    _set("premissas", premissas)
                    self._send_json({"ok": True, "data": parsed_data, "errors": []})

            elif path == "/api/upload/portfolio":
                if not _HAS_OPENPYXL:
                    self._send_json({"ok": False, "errors": ["openpyxl indisponivel no servidor."], "data": {}}); return
                ct = self.headers.get("Content-Type", "")
                file_bytes = _extract_file(self._read_body(), ct)
                if len(file_bytes) < 100:
                    self._send_json({"ok": False, "errors": [f"Arquivo nao recebido ({len(file_bytes)} bytes)."], "data": {}}); return
                empresa_data, errors = parse_portfolio(file_bytes)
                if errors:
                    self._send_json({"ok": False, "errors": errors, "data": empresa_data})
                else:
                    self._send_json({"ok": True, "data": empresa_data, "errors": []})

            elif path == "/api/calcular/simulador":
                body = json.loads(self._read_body())
                self._send_json(calcular_fa(body, _get("premissas")))

            elif path == "/api/portfolio_extra":
                data = json.loads(self._read_body())
                _set("portfolio_extra", data)
                self._send_json({"ok": True})

            elif path == "/api/calcular/extra-preview":
                extra = json.loads(self._read_body())
                empresa = _get("empresa"); premissas = _get("premissas")
                real = calcular_fa(empresa, premissas)
                comb = calcular_fa(combinar_portfolios(empresa, extra), premissas)
                self._send_json({"real": real, "combinado": comb})

            elif path == "/api/upload/portfolio_extra_csv":
                ct = self.headers.get("Content-Type", "")
                file_bytes = _extract_file(self._read_body(), ct)
                if len(file_bytes) < 20:
                    self._send_json({"ok": False, "errors": [f"Arquivo nao recebido ({len(file_bytes)} bytes)."], "data": {}}); return
                extra_data, errors = parse_extra_csv(file_bytes)
                if errors:
                    self._send_json({"ok": False, "errors": errors, "data": extra_data})
                else:
                    self._send_json({"ok": True, "data": extra_data, "errors": []})

            else:
                self.send_response(404); self.end_headers()
        except Exception:
            traceback.print_exc()
            try:
                self._send_json({"ok": False, "errors": [traceback.format_exc()]}, 500)
            except Exception:
                pass


def _extract_file(body_bytes, content_type):
    if "multipart" not in content_type:
        return body_bytes
    m = re.search(r'boundary=([^\s;]+)', content_type)
    if not m:
        return body_bytes
    boundary = m.group(1).encode('ascii')
    sep = b'--' + boundary
    idx = body_bytes.find(sep)
    if idx == -1:
        return body_bytes
    hdr_end = body_bytes.find(b'\r\n\r\n', idx)
    if hdr_end == -1:
        return body_bytes
    file_start = hdr_end + 4
    idx_next = body_bytes.find(b'\r\n' + sep, file_start)
    return body_bytes[file_start:idx_next] if idx_next != -1 else body_bytes[file_start:]


class ThreadedHTTPServer(ThreadingMixIn, HTTPServer):
    daemon_threads = True


def run(port=None):
    port = port or int(os.environ.get("PORT", 8765))
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass
    server = ThreadedHTTPServer(("0.0.0.0", port), Handler)
    print(f"\n  Simulador FA CCEE rodando em http://localhost:{port}")
    print("  Pressione Ctrl+C para encerrar.\n", flush=True)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nServidor encerrado.")


if __name__ == "__main__":
    p = int(sys.argv[1]) if len(sys.argv) > 1 else None
    run(p)
